import os
import io
import csv
import json
import secrets
import requests
import uuid
import threading
import time as _time
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse, urlencode
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, render_template, send_file, abort, redirect, url_for, flash
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from dotenv import load_dotenv
import google.generativeai as genai
from google.api_core.exceptions import ResourceExhausted
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from fpdf.errors import FPDFException
import re
import click
import base64
from functools import wraps
from email.utils import parseaddr, parsedate_to_datetime
from html import unescape
from db import db as default_db
from flask import current_app
from cryptography.fernet import Fernet, InvalidToken

class DatabaseAdapterProxy:
    def __getattr__(self, name):
        adapter = default_db
        try:
            if current_app and 'DB_ADAPTER' in current_app.config:
                adapter = current_app.config['DB_ADAPTER']
        except RuntimeError:
            pass # Outside application context
        return getattr(adapter, name)

sheets_db = DatabaseAdapterProxy()

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

load_dotenv()

# --- Flask App Setup ---
app = Flask(__name__)
from werkzeug.middleware.proxy_fix import ProxyFix
# Fix for Vercel to handle secure cookies
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

@app.teardown_appcontext
def teardown_db(exception=None):
    try:
        sheets_db.close_scoped_connection()
    except Exception as e:
        app.logger.error(f"Error closing database connection: {e}")


# --- Rate Limiting (Flask-Limiter) ---
def get_real_client_ip():
    try:
        if request.headers.get('CF-Connecting-IP'):
            return request.headers.get('CF-Connecting-IP')
        if request.headers.get('X-Forwarded-For'):
            return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    except Exception:
        pass
    return get_remote_address()

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        get_real_client_ip,
        default_limits=["200 per day", "50 per hour"],
        default_limits_exempt_when=lambda: current_user.is_authenticated,
        storage_uri="memory://"
    )
except ImportError:
    # Fail-fast check outside testing
    if not os.getenv("TESTING") and not os.getenv("FLASK_TESTING"):
        raise ImportError("CRITICAL: Flask-Limiter is required for this application outside testing environments. Please install flask-limiter.")
    class DummyLimiter:
        def __init__(self, *args, **kwargs):
            pass
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
        def init_app(self, app):
            pass
    limiter = DummyLimiter()

# --- SECURITY: SECRET_KEY must be set via environment variable ---
_secret_key = os.getenv("FLASK_SECRET_KEY")
if not _secret_key:
    if os.getenv("FLASK_ENV") == "development" or os.getenv("TESTING"):
        _secret_key = "dev-only-insecure-key-" + secrets.token_hex(16)
        print("WARNING: Using auto-generated SECRET_KEY for development. Set FLASK_SECRET_KEY for production.")
    else:
        raise ValueError("CRITICAL: FLASK_SECRET_KEY environment variable is not set. Refusing to start in production without it.")
app.config['SECRET_KEY'] = _secret_key

# --- SECURITY: Limit upload size to 100MB (user requested revert) ---
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# --- SECURITY: Secure session cookies ---
app.config['SESSION_COOKIE_SECURE'] = True       # Only send over HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True      # JavaScript cannot access cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'    # Prevent CSRF via cross-site requests
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=7)
app.config['REMEMBER_COOKIE_SECURE'] = True
app.config['REMEMBER_COOKIE_HTTPONLY'] = True

bcrypt = Bcrypt()
login_manager = LoginManager()
login_manager.login_view = 'login' 
login_manager.login_message_category = 'info'

# --- Database Models (Adapted for Sheets) ---
class User(UserMixin):
    def __init__(self, user_data):
        self.id = str(user_data.get('id'))
        self.username = user_data.get('username')
        self.password_hash = user_data.get('password_hash')
        self.full_name = user_data.get('full_name')
        self.qualifications = user_data.get('qualifications')
        self.designation = user_data.get('designation')
        self.license_no = user_data.get('license_no')
        self.expiry_date = user_data.get('expiry_date')
        self.membership_no = user_data.get('membership_no')
        self.address_line_1 = user_data.get('address_line_1')
        self.address_line_2 = user_data.get('address_line_2')
        self.address_line_3 = user_data.get('address_line_3')
        self.contact_no = user_data.get('contact_no')
        self.email = user_data.get('email')
        self.gemini_api_key = user_data.get('gemini_api_key')
        self.gemini_model = user_data.get('gemini_model')
        self.role = user_data.get('role') or 'employee'
        self.admin_id = user_data.get('admin_id')
        self.is_locked = bool(user_data.get('is_locked', False))
        self.must_change_password = bool(user_data.get('must_change_password', False))
        permissions = user_data.get('permissions') or {}
        if isinstance(permissions, str):
            try:
                permissions = json.loads(permissions)
            except (ValueError, TypeError):
                permissions = {}
        self.permissions = permissions if isinstance(permissions, dict) else {}

    def get_id(self):
        return self.id

    def __repr__(self):
        return f'<User {self.username}>'

# SavedReport class is no longer needed as an Object, we handle dicts directly or simple wrapper if needed.
# But existing code might rely on it if I don't catch all usages. 
# For now, I will return dicts from sheets_db and adapt logic.

# --- Flask-Login User Loader ---
@login_manager.user_loader
def load_user(user_id):
    user_data = sheets_db.get_user_by_id(user_id)
    if user_data:
        return User(user_data)
    return None


# --- Workspace / RBAC Helpers ---
VALID_CLAIM_STATUSES = {
    'new_appointment', 'inspection_pending', 'documents_awaited',
    'report_under_preparation', 'report_submitted', 'closed'
}

def is_admin_user(user=None):
    user = user or current_user
    return bool(getattr(user, 'is_authenticated', False) and getattr(user, 'role', 'employee') == 'admin')

def workspace_admin_id_for(user=None):
    user = user or current_user
    if not getattr(user, 'is_authenticated', False):
        return None
    if is_admin_user(user):
        return int(user.id)
    admin_id = getattr(user, 'admin_id', None)
    try:
        return int(admin_id) if admin_id is not None else None
    except (TypeError, ValueError):
        return None

def has_user_permission(permission, user=None):
    user = user or current_user
    if is_admin_user(user):
        return True
    return bool((getattr(user, 'permissions', {}) or {}).get(permission, False))

def _api_or_json_request():
    return request.path.startswith('/api/') or request.is_json

def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not getattr(current_user, 'is_authenticated', False):
            return login_manager.unauthorized()
        if not is_admin_user(current_user):
            return jsonify({'error': 'Administrator access is required.'}), 403
        return view(*args, **kwargs)
    return wrapped

def gmail_sync_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not getattr(current_user, 'is_authenticated', False):
            return login_manager.unauthorized()
        if not has_user_permission('gmail_sync', current_user):
            return jsonify({'error': 'Gmail sync permission is required.'}), 403
        if not workspace_admin_id_for(current_user):
            return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
        return view(*args, **kwargs)
    return wrapped

@app.before_request
def reject_locked_accounts():
    if getattr(current_user, 'is_authenticated', False) and getattr(current_user, 'is_locked', False):
        logout_user()
        if _api_or_json_request():
            return jsonify({'error': 'This account is locked.'}), 403
        flash('This account is locked. Please contact your administrator.', 'danger')
        return redirect(url_for('login'))

@app.before_request
def require_password_change_when_flagged():
    """A reset employee may sign in only to change their temporary password."""
    if not getattr(current_user, 'is_authenticated', False) or not getattr(current_user, 'must_change_password', False):
        return None
    allowed_endpoints = {'index', 'logout', 'get_user_profile', 'change_password', 'static'}
    if request.endpoint in allowed_endpoints:
        return None
    if _api_or_json_request():
        return jsonify({'error': 'You must change your password before continuing.'}), 403
    flash('Please change your temporary password before continuing.', 'warning')
    return redirect(url_for('index'))

def _copy_json(value):
    return json.loads(json.dumps(value or {}))

def redact_financial_report_data(report_data):
    """Employees may edit operations but cannot read or replace survey-fee values."""
    clean = _copy_json(report_data)
    clean.pop('fee_breakdown', None)
    page3 = clean.get('assessment', {}).get('page3_details', {})
    protected_page3_fields = {
        'photo_charges', 'fees_subtotal', 'total_before_gst', 'cgst', 'sgst', 'igst',
        'grand_total', 'apply_gst', 'fee_items', 'include_in_consolidated',
    }
    for key in list(page3.keys()):
        if 'fee' in key.lower() or key in protected_page3_fields:
            page3.pop(key, None)
    return clean

def preserve_financial_report_data(incoming, stored):
    """Keep protected fee values server-side when an employee saves an edited report."""
    merged = _copy_json(incoming)
    stored = stored or {}
    if 'fee_breakdown' in stored:
        merged['fee_breakdown'] = _copy_json(stored['fee_breakdown'])
    stored_page3 = stored.get('assessment', {}).get('page3_details', {})
    merged_page3 = merged.setdefault('assessment', {}).setdefault('page3_details', {})
    protected_page3_fields = {
        'photo_charges', 'fees_subtotal', 'total_before_gst', 'cgst', 'sgst', 'igst',
        'grand_total', 'apply_gst', 'fee_items', 'include_in_consolidated',
    }
    for key, value in stored_page3.items():
        if 'fee' in key.lower() or key in protected_page3_fields:
            merged_page3[key] = _copy_json(value)
    return merged

# --- Gemini API Configuration ---
def _score_model_for_intelligence(name):
    """
    Dynamically score models to prioritize pure intelligence/reasoning capabilities.
    Calibrated against Artificial Analysis Intelligence Index hierarchies.
    """
    n = name.lower()
    score = 0
    
    # 1. Base Version Multiplier (The strongest indicator of intelligence)
    # E.g., 3.1 adds 3,100,000; 2.5 adds 2,500,000.
    m = re.search(r'(\d+)\.(\d+)', n)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2))
        score += (major * 1000000) + (minor * 100000)
        
    # 2. Base Tier Capabilities
    # Pro > Flash > Flash-Lite
    if "pro" in n:
        score += 300000
    elif "flash-lite" in n:
        score += 0       # Baseline for its specific generation
    elif "flash" in n:
        score += 100000
        
    # 3. Reasoning / Thinking Modifiers (The Paradigm Shift)
    # Models with native/maximized thinking massively outperform their base counterparts.
    # Checks for historical and future reasoning indicators.
    if any(kw in n for kw in ["thinking", "reasoning", "deep-think", "high"]):
        score += 90000
    elif "low" in n:
        score -= 150000  # Heavily penalize explicitly downgraded reasoning models
        
    # 4. Model State / Stability
    # For pure intelligence, we tolerate experimental/preview to get the smartest model.
    # We only apply very minor bonuses to break ties in favor of stability.
    if "exp" in n or "experimental" in n:
        score += 0
    elif "preview" in n:
        score += 20
    else:
        score += 50      # Stable gets a slight edge if intelligence is equal
        
    # 5. Date Tie-Breaker
    # If two models have the exact same stats, newer date wins.
    date_match = re.search(r'-(\d{4,8})', n)
    if date_match:
        val = int(date_match.group(1))
        score += (val % 20) 
        
    return score

def load_valid_api_key():
    """
    Loads GEMINI_API_KEY from the environment, prioritizing .env.local,
    but falling back to .env if the prioritized key is expired or invalid.
    """
    # Tests must never make a network call merely by importing the application.
    if os.getenv('TESTING') or os.getenv('FLASK_TESTING'):
        return os.getenv('GEMINI_API_KEY')

    # 1. Try loading .env.local first
    if os.path.exists('.env.local'):
        load_dotenv('.env.local', override=True)
        key = os.getenv("GEMINI_API_KEY")
        if key:
            try:
                # Test the key by listing models
                genai.configure(api_key=key)
                list(genai.list_models())
                print("[API-KEY] Successfully verified GEMINI_API_KEY from .env.local")
                return key
            except Exception as e:
                print(f"[API-KEY] Key from .env.local failed verification ({e}). Falling back to .env...")
                
    # 2. Fall back to .env
    load_dotenv('.env', override=True)
    key = os.getenv("GEMINI_API_KEY")
    if key:
        try:
            genai.configure(api_key=key)
            list(genai.list_models())
            print("[API-KEY] Successfully verified GEMINI_API_KEY from .env")
            return key
        except Exception as e:
            print(f"[API-KEY] Key from .env failed verification ({e}).")
            
    return key

# Load the working API key
API_KEY = load_valid_api_key()
if not API_KEY:
    # If both failed or are invalid, fall back to standard os.getenv to avoid complete block
    API_KEY = os.getenv("GEMINI_API_KEY")
    if not API_KEY:
        raise ValueError("GEMINI_API_KEY not found in environment or .env files.")

# Re-configure with verified key
genai.configure(api_key=API_KEY)

generation_config = {
  "temperature": 1,
  "top_p": 0.95,
  "top_k": 64,
  "max_output_tokens": 65536, 
  "response_mime_type": "text/plain",
}
safety_settings = [
    {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
]

def get_best_models():
    """
    Query available models from Gemini API, rank them using intelligence scoring,
    and return a sorted list of model names.
    """
    default_models = ['gemini-1.5-pro', 'gemini-1.5-flash']
    try:
        models = []
        for m in genai.list_models():
            if "generateContent" in m.supported_generation_methods:
                clean_name = m.name.split('/')[-1] if '/' in m.name else m.name
                score = _score_model_for_intelligence(clean_name)
                models.append((clean_name, score))
        
        if not models:
            return default_models
            
        # Sort by score descending
        models.sort(key=lambda x: x[1], reverse=True)
        return [name for name, score in models]
    except Exception as e:
        print(f"[MODEL-SELECT] Error fetching available models: {e}. Using defaults.")
        return default_models

# Resolve best models using intelligence scoring
best_models = get_best_models()
print(f"[MODEL-SELECT] Ranked available models: {best_models}")

PRIMARY_MODEL_NAME = best_models[0] if best_models else 'gemini-1.5-pro'
SECONDARY_MODEL_NAME = best_models[1] if len(best_models) > 1 else PRIMARY_MODEL_NAME

print(f"[MODEL-SELECT] Selecting primary model: {PRIMARY_MODEL_NAME}")
print(f"[MODEL-SELECT] Selecting secondary model: {SECONDARY_MODEL_NAME}")

model = genai.GenerativeModel(
    model_name=PRIMARY_MODEL_NAME,
    safety_settings=safety_settings,
    generation_config=generation_config
)

secondary_model = genai.GenerativeModel(
    model_name=SECONDARY_MODEL_NAME,
    safety_settings=safety_settings,
    generation_config=generation_config
)

def get_user_best_models(api_key):
    default_models = ['gemini-1.5-pro', 'gemini-1.5-flash']
    if not api_key:
        return default_models
    original_key = os.getenv("GEMINI_API_KEY")
    try:
        genai.configure(api_key=api_key)
        models = []
        for m in genai.list_models():
            if "generateContent" in m.supported_generation_methods:
                clean_name = m.name.split('/')[-1] if '/' in m.name else m.name
                score = _score_model_for_intelligence(clean_name)
                models.append((clean_name, score))
        if original_key:
            genai.configure(api_key=original_key)
        if not models:
            return default_models
        models.sort(key=lambda x: x[1], reverse=True)
        return [name for name, score in models]
    except Exception as e:
        print(f"[MODEL-SELECT] Error fetching models for custom key: {e}. Using defaults.")
        if original_key:
            try:
                genai.configure(api_key=original_key)
            except Exception:
                pass
        return default_models

def get_generative_models(user=None):
    user_key = user.gemini_api_key if (user and getattr(user, 'gemini_api_key', None)) else None
    api_key = user_key or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY is not configured.")
    genai.configure(api_key=api_key)
    
    user_model = user.gemini_model if (user and getattr(user, 'gemini_model', None)) else None
    if user_model:
        primary_name = user_model
        secondary_name = user_model
    else:
        ranked = get_user_best_models(api_key)
        primary_name = ranked[0] if ranked else 'gemini-1.5-pro'
        secondary_name = ranked[1] if len(ranked) > 1 else primary_name
        
    primary = genai.GenerativeModel(
        model_name=primary_name,
        safety_settings=safety_settings,
        generation_config=generation_config
    )
    secondary = genai.GenerativeModel(
        model_name=secondary_name,
        safety_settings=safety_settings,
        generation_config=generation_config
    )
    return primary, secondary

class DiskDataStore:
    def __init__(self):
        self.directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'temp_pdfs')
        os.makedirs(self.directory, exist_ok=True)

    def _get_paths(self, key):
        safe_key = "".join(c for c in key if c.isalnum() or c in ('-', '_'))
        meta_path = os.path.join(self.directory, f"{safe_key}.json")
        pdf_path = os.path.join(self.directory, f"{safe_key}.pdf")
        return meta_path, pdf_path

    def __setitem__(self, key, value):
        meta_path, pdf_path = self._get_paths(key)
        metadata = {
            "report_no": value.get("report_no"),
            "vehicle_no": value.get("vehicle_no"),
            "user_id": value.get("user_id")
        }
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f)
        with open(pdf_path, "wb") as f:
            f.write(value.get("pdf_report"))

    def __getitem__(self, key):
        meta_path, pdf_path = self._get_paths(key)
        if not os.path.exists(meta_path) or not os.path.exists(pdf_path):
            raise KeyError(key)
        with open(meta_path, "r", encoding="utf-8") as f:
            metadata = json.load(f)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
        return {
            "pdf_report": pdf_bytes,
            "report_no": metadata.get("report_no") or "",
            "vehicle_no": metadata.get("vehicle_no") or "",
            "user_id": metadata.get("user_id")
        }

    def __contains__(self, key):
        meta_path, pdf_path = self._get_paths(key)
        return os.path.exists(meta_path) and os.path.exists(pdf_path)

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default

    def clear(self):
        import shutil
        if os.path.exists(self.directory):
            for filename in os.listdir(self.directory):
                file_path = os.path.join(self.directory, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f'Failed to delete {file_path}. Reason: {e}')

# --- Persistent storage for generated files (Shared across Gunicorn workers) ---
generated_data_store = DiskDataStore()

# --- Async Task Infrastructure ---
# Avoids Cloudflare Free plan's hard 100-second proxy timeout by returning
# a task_id immediately and processing in a background thread.
_task_store = {}  # { task_id: { "status": ..., "result": ..., "error": ..., "created_at": ... } }
_task_lock = threading.Lock()
_task_executor = ThreadPoolExecutor(max_workers=4)

def _create_task():
    """Create a new task entry and return its ID."""
    task_id = str(uuid.uuid4())
    with _task_lock:
        _task_store[task_id] = {
            "status": "processing",
            "result": None,
            "error": None,
            "created_at": _time.time()
        }
        # Purge tasks older than 30 minutes to prevent memory leaks
        cutoff = _time.time() - 1800
        stale = [k for k, v in _task_store.items() if v["created_at"] < cutoff]
        for k in stale:
            del _task_store[k]
    return task_id

def _complete_task(task_id, result):
    """Mark a task as completed with its result."""
    with _task_lock:
        if task_id in _task_store:
            _task_store[task_id]["status"] = "completed"
            _task_store[task_id]["result"] = result

def _fail_task(task_id, error_msg):
    """Mark a task as failed with an error message."""
    with _task_lock:
        if task_id in _task_store:
            _task_store[task_id]["status"] = "error"
            _task_store[task_id]["error"] = error_msg

def _get_task(task_id):
    """Get a copy of the task state."""
    with _task_lock:
        task = _task_store.get(task_id)
        if task:
            return dict(task)
        return None

# --- Define the expected fields based on the template ---
EXPECTED_FIELDS = [
    "report_no", "report_date", "policy_no", "claim_no", "policy_validity",
    "insurer", "insured", "insured_contact_name", "insured_contact_no", "hypothecation", "idv", "policy_type_label",
    # Vehicle fields
    "vehicle_regn_no", "vehicle_regn_date", "vehicle_chassis_no", "vehicle_engine_no",
    "vehicle_make_model", "vehicle_type_body", "vehicle_cf_validity", "vehicle_seating",
    "vehicle_bhp_cc", "vehicle_pre_accident_condition", "vehicle_ulw", "vehicle_rlw",
    "vehicle_permit_no", "vehicle_permit_type", "vehicle_permit_validity",
    "vehicle_route_area", "vehicle_tax_token", "vehicle_tax_validity",
    "vehicle_odometer", "vehicle_colour", "class_of_vehicle", "regn_cert_no", "vehicle_cc",
    # DL fields
    "dl_name", "dl_no", "dl_issue_date", "dl_validity", "dl_issuing_authority",
    "dl_endorsement", "dl_type", "dl_dob",
    # Docs compared fields
    "doc_regn_cert", "doc_dl", "doc_tax_token", "doc_permit_compared",
    "doc_fitness_certificate", "doc_load_challan",
    # Load/Goods Details
    "load_nature_packing", "load_weight_goods", "load_origin_destination",
    "load_lr_invoice_no", "load_transport_name", "load_date",
    # Accident fields
    "accident_datetime", "accident_assign_received", "accident_survey_date",
    "accident_place", "accident_survey_place",
    # Police fields
    "police_reported_to", "police_diary_case_no", "police_date_reported",
    # Other fields
    "tp_details", "accident_cause", "damages_extent", "remark",
    "tp_injury_loss", "injury_driver_occupant", "damages_consistent"
]

# --- Gemini Prompt Function ---
def build_gemini_prompt():
    """Creates the detailed prompt for Gemini extraction for both Survey Report and Assessment details."""
    # Prompt remains the same as before
    prompt = """
    You are an expert data extraction assistant specializing in Indian motor insurance claim documents.
    Analyze the provided PDF document which contains various supporting documents like Registration Certificate (RC), Driving License (DL), Insurance Policy, Claim Form, Repair Estimate/Pre-Invoice/Tax Invoice, etc.
    Your goal is to extract specific information for BOTH a Motor Final Survey Report AND a Repair Assessment Summary, ensuring all descriptive text is in English.

    **IMPORTANT INSTRUCTIONS:**
    1.  **Prioritize Typed Text & Clarity:** Strongly prefer machine-written/printed text. Choose the most clearly legible and complete value if multiple sources exist. Use handwritten only if typed is missing/illegible/overridden.
    2.  **Multi-line Data:** Combine multi-line data for a single field into one string.
    3.  **Transcribe to English:** Transcribe free-form text (addresses, descriptions, places, remarks, cause, damages, TP details, authority names, route area) to English if originally non-English. Extract names, IDs, technical specs as they are.
    4.  **Missing Information:** Use an empty string "" if information cannot be reliably found. Do NOT guess or provide defaults like 'Average', 'Not Known', 'Not Reported', 'N/A', 'No ( As Per Claim Form )', etc. Return "" for these. The application will handle defaults later. Specifically return "" for `vehicle_pre_accident_condition`, `dl_endorsement`, `police_reported_to`, `police_diary_case_no`, `police_date_reported`, `tp_details`, `accident_cause`, `damages_extent`, `remark`, `tp_injury_loss`, `injury_driver_occupant`, `damages_consistent` and all `load_*` fields if not found.
    5.  **Permit & Load Details:** Look for commercial vehicle permit and goods/load details. If found, populate `vehicle_permit_*`, `doc_permit_compared`, `doc_load_challan`, and all `load_*` fields. If private or no permit/load found, return "" for these.
    6.  **Tax Token:** Look for the Tax Token number, often labeled as Application Number or Receipt Number. Extract this value for `vehicle_tax_token`.
    7.  **Assessment Data Source:** Extract assessment data (Parts, Summary) primarily from the **Job Card Retail - Tax Invoice** or **Pre-Invoice** section of the PDF. Labour totals are NOT required from AI.
    8.  **Labour Extraction (Table 12):** DO NOT extract labour totals (Painting/Denting). This will be handled by user input.
    9.  **Parts Extraction (Table 13):** Extract EACH line item from the "Parts" section of the invoice. Include: Description (as Part Name), Quantity (Qty), Taxable Amount (calculate per unit if total is given), and Tax % (GST Rate).
    10. **Summary Extraction:** Extract "Deductibles" (or similar term like Excess/Compulsory Excess) and "Salvage" amount if mentioned in the invoice summary.
    
    Return the extracted data STRICTLY in JSON format with the following nested structure:
    
    {
      "survey_report_data": {
        "report_no": "...",
        "report_date": "...", // Format DD.MM.YYYY
        "policy_no": "...",
        "claim_no": "...",
        "policy_validity": "...", // Format: "DD.MM.YYYY to DD.MM.YYYY"
        "insurer": "...", // Full name & address (transcribed)
        "insured": "...", // Full name & address (transcribed)
        "insured_contact_name": "...", // Name of contact person for insured, if different from insured name. "" if not found.
        "insured_contact_no": "...", // Contact phone number for insured. "" if not found.
        "hypothecation": "...", // Financer name. "" if none.
        "idv": "...", // Insured Declared Value, number as string
        "policy_type_label": "...", // e.g., "Regular Policy", "Nil Depreciation"
        "vehicle_regn_no": "...", // a.
        "vehicle_regn_date": "...", // b. DD.MM.YYYY
        "vehicle_chassis_no": "...", // c.
        "vehicle_engine_no": "...", // d.
        "vehicle_make_model": "...", // e.
        "vehicle_type_body": "...", // f.
        "vehicle_cf_validity": "...", // g. DD.MM.YYYY
        "vehicle_seating": "...", // h. Number
        "vehicle_bhp_cc": "...", // i. e.g., "1197 cc"
        "vehicle_pre_accident_condition": "...", // j. Transcribed. "" if not found.
        "vehicle_ulw": "...", // k. e.g., "918 kg"
        "vehicle_rlw": "...", // l. e.g., "1355 kg"
        "vehicle_permit_no": "...", // m. "" if none/N/A.
        "vehicle_permit_type": "...", // n. "" if none/N/A.
        "vehicle_permit_validity": "...", // o. "" if none/N/A.
        "vehicle_route_area": "...", // p. Transcribed. "" if none/N/A.
        "vehicle_tax_token": "...", // q/m. Look for Application No., if not available then Receipt No., "" if none.
        "vehicle_tax_validity": "...", // r/n. "" if none.
        "vehicle_odometer": "...", // s/o. "" if not found.
        "vehicle_colour": "...", // t/p. "" if not found.
        "class_of_vehicle": "...", // e.g., "LMV", "Private Car"
        "regn_cert_no": "...", // RC document/certificate number
        "vehicle_cc": "...", // e.g., "2184"
        "dl_name": "...", // a.
        "dl_no": "...", // b.
        "dl_issue_date": "...", // c. DD.MM.YYYY
        "dl_validity": "...", // d. DD.MM.YYYY
        "dl_issuing_authority": "...", // e. Transcribed.
        "dl_endorsement": "...", // f. "" if not found.
        "dl_type": "...", // g.
        "dl_dob": "...", // h. DD.MM.YYYY
        "doc_regn_cert": "...", // a. YES/NO/- or ""
        "doc_dl": "...", // b. YES/NO/- or ""
        "doc_tax_token": "...", // c. YES/NO/- or ""
        "doc_permit_compared": "...", // d. YES/NO/- or "" if none/N/A.
        "doc_fitness_certificate": "...", // YES/NO/- or ""
        "doc_load_challan": "...", // YES/NO/- or ""
        "load_nature_packing": "...", // "" if not applicable
        "load_weight_goods": "...", // "" if not applicable
        "load_origin_destination": "...", // "" if not applicable
        "load_lr_invoice_no": "...", // "" if not applicable
        "load_transport_name": "...", // "" if not applicable
        "load_date": "...", // DD.MM.YYYY or ""
        "accident_datetime": "...", // a. e.g., "26.03.2025 at 07:00 Am"
        "accident_assign_received": "...", // b. DD.MM.YYYY
        "accident_survey_date": "...", // c. DD.MM.YYYY at HH:MM
        "accident_place": "...", // d. Transcribed.
        "accident_survey_place": "...", // e. Workshop address (transcribed).
        "police_reported_to": "...", // a. Transcribed. "" if not found.
        "police_diary_case_no": "...", // b. "" if not found.
        "police_date_reported": "...", // c. DD.MM.YYYY. "" if not found.
        "tp_details": "...", // 9. Transcribed. "" if not found.
        "accident_cause": "...", // 10. Transcribed. "" if not found.
        "damages_extent": "...", // 11. Transcribed. "" if not found.
        "remark": "...", // Transcribed. "" if not found.
        "tp_injury_loss": "...", // Short sentence or ""
        "injury_driver_occupant": "...", // Short sentence or ""
        "damages_consistent": "..." // "yes", "no", or ""
      },
      "assessment_data": {
        "customer_gstin": "...", // Extracted Customer GSTIN/UIN, or "" if not found
        "parts": [ // Array of extracted parts
          {
            "part_name": "...", // Description from invoice
            "qty": "...", // Quantity (numeric string or "")
            "part_amt": "...", // Taxable Amount PER UNIT (numeric string or ""). Calculate if necessary.
            "gst_pc": "..." // GST Percentage (numeric string like "18", "28", or "")
          }
          // ... more parts
        ],
        "deductibles": "...", // Extracted Deductibles/Excess amount (numeric string or "")
        "salvage": "..." // Extracted Salvage amount (numeric string or "")
      }
    }

    Ensure the output is ONLY the JSON object, without any introductory text, explanations, or markdown formatting like ```json ... ```. Use "" for any field where the information cannot be reliably extracted. For numeric fields in assessment_data, return numbers as strings, or "" if not found.
    """
    return prompt

def build_invoice_gemini_prompt():
    """Creates a focused prompt for Gemini to extract parts data AND Customer GSTIN from an invoice."""
    prompt = """
    You are an expert data extraction assistant specializing in Indian motor repair invoices/estimates.
    Analyze the provided PDF document which should be a Tax Invoice, Pre-Invoice, or Repair Estimate.
    Your goal is to extract:
    1.  The line items listed under the "Parts" or "Materials" section.
    2.  The Customer's GSTIN/UIN if available on the invoice (often labeled as "GSTIN", "Customer GSTIN", or similar).

    **IMPORTANT INSTRUCTIONS:**
    1.  **Focus:** Extract ONLY the parts list and the Customer GSTIN/UIN. Ignore labour charges, summaries, other addresses, vehicle details etc.
    2.  **Data Points (Parts):** For EACH part line item, extract:
        *   Part Name/Description
        *   Quantity (Qty)
        *   Taxable Amount (Rate per unit, or calculate if only total is given)
        *   Tax Rate Percentage (GST Rate %, e.g., "18", "28")
    3.  **Customer GSTIN/UIN:** Look for the recipient's (customer's) GSTIN. If multiple GSTINs are present, prioritize the one clearly associated with the customer or "Bill To" party. If not found, return "".
    4.  **Formatting:** Use numeric values where possible for Qty, Amount, and Tax Rate. If a value is not found or unclear, use an empty string "" or 0 where appropriate (0 for numeric fields if missing).
    5.  **Structure:** Return the extracted data STRICTLY in JSON format as follows:

    {
      "customer_gstin": "...", // Extracted Customer GSTIN/UIN, or "" if not found
      "parts": [ // Array of extracted parts
        {
          "part_name": "...", // Description from invoice
          "qty": "...", // Quantity (numeric string or "")
          "part_amt": "...", // Taxable Amount PER UNIT (numeric string or ""). Calculate if necessary.
          "gst_pc": "..." // GST Percentage (numeric string like "18", "28", or "")
        },
        // ... more parts
      ]
    }

    Ensure the output is ONLY the JSON object, without any introductory text, explanations, or markdown formatting like ```json ... ```. If no parts table is found, return {"customer_gstin": "", "parts": []}.
    """
    return prompt

# --- Gemini Response Parser ---
def parse_gemini_response(response_text):
    """
    Attempts to parse JSON from Gemini's text response, handling nested structure,
    performing initial calculations, attempting fixes, and preparing data for editable fields.
    """
    try:
        # Basic cleanup and fixes
        if response_text.strip().startswith("```json"):
            response_text = response_text.strip()[7:-3].strip()
        elif response_text.strip().startswith("```"):
             response_text = response_text.strip()[3:-3].strip()
        response_text = response_text.strip()
        response_text = re.sub(r',\s*([}\]])', r'\1', response_text) 
        response_text = re.sub(r'(\s*)"(\w+)":\s*undefined', r'\1"\2": ""', response_text)
        response_text = re.sub(r'(\s*)"(\w+)":\s*null', r'\1"\2": ""', response_text)

        data = json.loads(response_text)

        survey_data_raw = data.get('survey_report_data', {})
        extracted_survey_data = {key: survey_data_raw.get(key, '') for key in EXPECTED_FIELDS}
        for key, value in extracted_survey_data.items():
            if value is None: extracted_survey_data[key] = ''

        assessment_data_raw = data.get('assessment_data', {})
        
        # Initialize page3_details with customer_gstin from AI if present
        customer_gstin_from_ai = str(assessment_data_raw.get('customer_gstin', '')).strip()

        extracted_assessment_data = {
            'labour_painting_total': 0.0, 
            'labour_denting_total': 0.0, 
            'labour_total_base': 0.0, 
            'labour_cgst': 0.0, 
            'labour_sgst': 0.0, 
            'labour_igst': 0.0, 
            'labour_grand_total': 0.0, 
            'labour_paint_depn': 0.0, 
            'labour_grand_total_adjusted': 0.0, 
            'parts': [],
            'parts_total_base': 0.0,
            'parts_total_gst': 0.0,
            'parts_grand_total': 0.0,
            'parts_net_total': 0.0, 
            'deductibles': 1000.0,
            'impose_excess': 0.0, # New Field
            'salvage': "-",
            'net_liability': 0.0,
            'user_labour_rows': [], 
            'header_gst': '', 
            'header_vehicle_year': '', 
            'policy_type': 'NORMAL', 
            'nd_deduction_pc': 5,
            'nd_deduction_amount': 0,
            'towing_charges': 0,
            'report_type': 'Final Survey Report', 
            'claim_type': 'Cashless', 
            'labour_tax_type': 'CGST/SGST',
            'page3_details': {
                'customer_gstin': customer_gstin_from_ai, 
                'fee_items': [], 
                'estimated_amount': '',
                'photo_copies_count': '',
                'include_in_consolidated': False 
            },
            'note_text': "Note :- The subject policy covered with Depn. waiver", 
            'payment_to_text': "REPAIRER" 
        }

        parts_list_raw = assessment_data_raw.get('parts', [])
        processed_parts = []
        for idx, part_raw in enumerate(parts_list_raw):
            if not isinstance(part_raw, dict): continue
            try:
                qty_str = str(part_raw.get('qty', '1')).strip()
                part_amt_str = str(part_raw.get('part_amt', '0')).strip()
                gst_pc_str = str(part_raw.get('gst_pc', '0')).replace('%','').strip()

                qty = float(qty_str) if qty_str else 1.0
                part_amt = float(part_amt_str) if part_amt_str else 0.0
                original_gst_pc = float(gst_pc_str) if gst_pc_str else 0.0

                gst_applicable = original_gst_pc > 0
                total_parts_amt = qty * part_amt
                total_gst = total_parts_amt * (original_gst_pc / 100.0) if gst_applicable else 0.0
                gross_amt = total_parts_amt + total_gst
                net_amt = gross_amt

                processed_part = {
                    "sl_no": idx + 1,
                    "est_sl_no": "", 
                    "bill_sl_no": "", 
                    "part_name": str(part_raw.get('part_name', '')),
                    "hns_code": "", # New Field
                    "estimate_amt": gross_amt, # Default to Gross
                    "bill_amt": total_parts_amt, # Default to Base
                    "type_part": "", 
                    "qty": qty, 
                    "part_amt": part_amt, 
                    "original_gst_pc": original_gst_pc,
                    "gst_applicable": gst_applicable,
                    "total_parts_amt": total_parts_amt, 
                    "total_gst": total_gst, 
                    "gross_amt": gross_amt, 
                    "depr": 0.0, 
                    "imt_23_amt": 0.0, # New Field
                    "net_amt": net_amt 
                }
                processed_parts.append(processed_part)
            except (ValueError, TypeError) as e:
                print(f"Warning: Could not parse part data for item {idx}: {part_raw}. Error: {e}")
        extracted_assessment_data['parts'] = processed_parts

        try:
            deductibles_raw = str(assessment_data_raw.get('deductibles', '')).strip()
            if deductibles_raw:
                 try:
                     parsed_deductible = float(deductibles_raw)
                     extracted_assessment_data['deductibles'] = parsed_deductible
                 except ValueError:
                     extracted_assessment_data['deductibles'] = 1000.0
            else:
                 extracted_assessment_data['deductibles'] = 1000.0
        except Exception as e:
             extracted_assessment_data['deductibles'] = 1000.0

        try:
            salvage_raw = str(assessment_data_raw.get('salvage', '')).strip()
            if salvage_raw:
                try: extracted_assessment_data['salvage'] = float(salvage_raw)
                except ValueError: extracted_assessment_data['salvage'] = salvage_raw 
            else:
                 extracted_assessment_data['salvage'] = "-" 
        except Exception as e:
            extracted_assessment_data['salvage'] = "-"

        final_data = {
            "survey_report": extracted_survey_data,
            "assessment": extracted_assessment_data
        }
        return final_data

    except json.JSONDecodeError as e:
        print(f"JSON Decode Error after attempting fixes: {e}")
        error_detail = f"Failed to parse JSON response from AI. Check logs. Error: {e}."
        raise ValueError(error_detail)
    except Exception as e:
        print(f"Error during parsing: {e}")
        import traceback
        traceback.print_exc()
        raise ValueError(f"An unexpected error occurred during response parsing. Error: {e}")

def parse_invoice_gemini_response(response_text):
    """
    Attempts to parse JSON containing parts data and customer_gstin from Gemini's invoice response.
    """
    try:
        # Basic cleanup (similar to main parser)
        if response_text.strip().startswith("```json"):
            response_text = response_text.strip()[7:-3].strip()
        elif response_text.strip().startswith("```"):
             response_text = response_text.strip()[3:-3].strip()
        response_text = response_text.strip()
        response_text = re.sub(r',\s*([}\]])', r'\1', response_text) # Remove trailing commas
        response_text = re.sub(r'(\s*)"(\w+)":\s*undefined', r'\1"\2": ""', response_text) # undefined -> ""
        response_text = re.sub(r'(\s*)"(\w+)":\s*null', r'\1"\2": ""', response_text) # null -> ""

        data = json.loads(response_text)

        customer_gstin = str(data.get('customer_gstin', '')).strip()
        parts_list_raw = data.get('parts', [])
        if not isinstance(parts_list_raw, list):
             raise ValueError("Expected 'parts' key to contain a list.")

        extracted_parts = []
        for idx, part_raw in enumerate(parts_list_raw):
            if not isinstance(part_raw, dict):
                print(f"Warning: Skipping non-dict item in parts list: {part_raw}")
                continue
            extracted_part = {
                "part_name": str(part_raw.get('part_name', '')).strip(),
                "qty": str(part_raw.get('qty', '1')).strip(), 
                "part_amt": str(part_raw.get('part_amt', '0')).strip(), 
                "gst_pc": str(part_raw.get('gst_pc', '0')).replace('%','').strip() 
            }
            extracted_parts.append(extracted_part)

        return {"customer_gstin": customer_gstin, "parts": extracted_parts}

    except json.JSONDecodeError as e:
        print(f"Invoice JSON Decode Error: {e}")
        print(f"Processed Invoice Response Text:\n---\n{response_text}\n---")
        raise ValueError(f"Failed to parse JSON from invoice AI response. Error: {e}")
    except Exception as e:
        print(f"Error during invoice response parsing: {e}")
        import traceback
        traceback.print_exc()
        raise ValueError(f"An unexpected error occurred during invoice response parsing: {e}")

def normalize_pdf_text_for_fpdf(text_val):
    """
    Replaces common problematic Unicode characters with their CP1252 equivalents
    or a standard ASCII alternative for FPDF's default fonts.
    This primarily targets characters that look like hyphens but are not the standard ASCII hyphen.
    """
    if not isinstance(text_val, str):
        text_val = str(text_val) # Ensure it's a string

    # Replace various dashes/hyphens with a standard hyphen (U+002D)
    text_val = text_val.replace('–', '-')  # En dash (U+2013)
    text_val = text_val.replace('—', '-')  # Em dash (U+2014)
    text_val = text_val.replace('−', '-')  # Minus sign (U+2212)
    
    # Aggressively replace all characters outside latin-1 to prevent fpdf.errors.FPDFUnicodeEncodingException
    text_val = text_val.encode('latin-1', 'replace').decode('latin-1')
    return text_val

def format_pdf_number(value):
    """Formats a number for PDF output. Shows '0' if zero, else formats to 2 decimal places."""
    try:
        num = float(value)
        if abs(num) < 0.001:
            return '0'
        else:
            return f"{num:.2f}"
    except (ValueError, TypeError):
         val_str = str(value)
         if val_str.strip() == '0':
             return '0'
         return normalize_pdf_text_for_fpdf(val_str)

# --- Security: URL Validation Helper ---
def _is_safe_redirect_url(target):
    """Validate that a redirect URL is safe (relative, same-host only)."""
    if not target:
        return False
    parsed = urlparse(target)
    # Only allow relative URLs (no scheme/netloc) that don't start with //
    return not parsed.scheme and not parsed.netloc and not target.startswith('//')

# --- Security: SSRF Allowlist for Proxy Endpoints ---
_ALLOWED_UPLOAD_DOMAINS = {'www.googleapis.com', 'googleapis.com'}

def _is_allowed_upload_url(url):
    """Validate that the upload URL targets only allowed Google API domains."""
    try:
        parsed = urlparse(url)
        return parsed.scheme == 'https' and parsed.hostname in _ALLOWED_UPLOAD_DOMAINS
    except Exception:
        return False

# --- Security Headers ---
@app.after_request
def set_security_headers(response):
    """Add security headers to every response."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    # HSTS: only set if on HTTPS (Vercel always serves HTTPS)
    if request.is_secure or request.headers.get('X-Forwarded-Proto') == 'https':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# --- Authentication Routes ---
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user_data = sheets_db.get_user_by_username(username)
        
        if user_data and user_data.get('is_locked', False):
            flash('This account is locked. Please contact your administrator.', 'danger')
        elif user_data and bcrypt.check_password_hash(user_data['password_hash'], password):
            user = User(user_data)
            login_user(user, remember=True)
            # SECURITY: Validate 'next' parameter to prevent open redirect (VULN-04)
            next_page = request.args.get('next')
            if next_page and _is_safe_redirect_url(next_page):
                redirect_target = next_page
            else:
                redirect_target = url_for('index')
            flash('Login Successful! Please change your password if your administrator reset it.', 'success')
            return redirect(redirect_target)
        else:
            flash('Login Unsuccessful. Please check username and password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# --- Google OAuth2 Configuration ---
GOOGLE_OAUTH_CLIENT_ID = os.getenv('GOOGLE_OAUTH_CLIENT_ID')
GOOGLE_OAUTH_CLIENT_SECRET = os.getenv('GOOGLE_OAUTH_CLIENT_SECRET')
GOOGLE_OAUTH_SCOPES = ['https://www.googleapis.com/auth/drive.file']
GMAIL_OAUTH_CLIENT_ID = os.getenv('GMAIL_OAUTH_CLIENT_ID') or GOOGLE_OAUTH_CLIENT_ID
GMAIL_OAUTH_CLIENT_SECRET = os.getenv('GMAIL_OAUTH_CLIENT_SECRET') or GOOGLE_OAUTH_CLIENT_SECRET
GMAIL_OAUTH_SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
GMAIL_SYNC_LOOKBACK_DAYS = max(1, int(os.getenv('GMAIL_SYNC_LOOKBACK_DAYS', '30')))

@app.route('/auth/google')
@login_required
def google_auth():
    """Initiate Google OAuth2 flow for Drive access."""
    from flask import session
    if not GOOGLE_OAUTH_CLIENT_ID or not GOOGLE_OAUTH_CLIENT_SECRET:
        return jsonify({'error': 'Google OAuth not configured'}), 500
    
    # Determine redirect URI based on request
    if request.host.startswith('localhost') or request.host.startswith('127.0.0.1'):
        redirect_uri = url_for('google_auth_callback', _external=True)
    else:
        redirect_uri = f"https://{request.host}/auth/google/callback"
    
    # SECURITY: Generate OAuth state token to prevent CSRF on callback (VULN-07)
    oauth_state = secrets.token_urlsafe(32)
    session['oauth_state'] = oauth_state
    
    auth_url = (
        'https://accounts.google.com/o/oauth2/v2/auth?'
        f'client_id={GOOGLE_OAUTH_CLIENT_ID}&'
        f'redirect_uri={redirect_uri}&'
        'response_type=code&'
        f'scope={" ".join(GOOGLE_OAUTH_SCOPES)}&'
        'access_type=offline&'
        'prompt=consent&'
        f'state={oauth_state}'
    )
    return redirect(auth_url)

@app.route('/auth/google/callback')
@login_required
def google_auth_callback():
    """Handle OAuth2 callback and store tokens."""
    from flask import session
    
    code = request.args.get('code')
    error = request.args.get('error')
    returned_state = request.args.get('state')
    
    # SECURITY: Verify OAuth state token to prevent CSRF (VULN-07)
    expected_state = session.pop('oauth_state', None)
    if not returned_state or returned_state != expected_state:
        flash('OAuth security verification failed. Please try again.', 'error')
        return redirect(url_for('index'))
    
    if error:
        flash(f'Google authorization failed: {error}', 'error')
        return redirect(url_for('index'))
    
    if not code:
        flash('No authorization code received', 'error')
        return redirect(url_for('index'))
    
    # Determine redirect URI (must match the one used in auth request)
    if request.host.startswith('localhost') or request.host.startswith('127.0.0.1'):
        redirect_uri = url_for('google_auth_callback', _external=True)
    else:
        redirect_uri = f"https://{request.host}/auth/google/callback"
    
    # Exchange code for tokens
    token_response = requests.post(
        'https://oauth2.googleapis.com/token',
        data={
            'client_id': GOOGLE_OAUTH_CLIENT_ID,
            'client_secret': GOOGLE_OAUTH_CLIENT_SECRET,
            'code': code,
            'grant_type': 'authorization_code',
            'redirect_uri': redirect_uri
        }
    )
    
    if token_response.status_code != 200:
        flash('Failed to get access token from Google', 'error')
        return redirect(url_for('index'))
    
    tokens = token_response.json()
    
    # Store tokens in session
    session['google_access_token'] = tokens.get('access_token')
    session['google_refresh_token'] = tokens.get('refresh_token')
    session['google_token_expiry'] = tokens.get('expires_in', 3600)
    
    flash('Successfully connected to Google Drive!', 'success')
    return redirect(url_for('index'))

@app.route('/auth/google/status')
@login_required
def google_auth_status():
    """Check if user has connected Google Drive."""
    from flask import session
    has_token = 'google_access_token' in session and session['google_access_token']
    return jsonify({'connected': has_token})

@app.route('/auth/google/disconnect', methods=['POST'])
@login_required
def google_auth_disconnect():
    """Disconnect Google Drive."""
    from flask import session
    session.pop('google_access_token', None)
    session.pop('google_refresh_token', None)
    session.pop('google_token_expiry', None)
    return jsonify({'success': True, 'message': 'Disconnected from Google Drive'})


# --- Gmail OAuth and On-Demand Intimation Sync ---
def _gmail_redirect_uri():
    if request.host.startswith('localhost') or request.host.startswith('127.0.0.1'):
        return url_for('gmail_auth_callback', _external=True)
    return f"https://{request.host}/auth/gmail/callback"


def _gmail_token_cipher():
    key = os.getenv('GMAIL_TOKEN_ENCRYPTION_KEY')
    if not key:
        raise ValueError('GMAIL_TOKEN_ENCRYPTION_KEY is not configured.')
    try:
        return Fernet(key.encode('utf-8'))
    except Exception as exc:
        raise ValueError('GMAIL_TOKEN_ENCRYPTION_KEY is invalid.') from exc


def _encrypt_gmail_token(token_data):
    return _gmail_token_cipher().encrypt(json.dumps(token_data).encode('utf-8')).decode('utf-8')


def _decrypt_gmail_token(encrypted_token):
    try:
        return json.loads(_gmail_token_cipher().decrypt(encrypted_token.encode('utf-8')).decode('utf-8'))
    except (InvalidToken, ValueError, TypeError, json.JSONDecodeError) as exc:
        raise ValueError('Stored Gmail authorization is invalid. Reconnect the mailbox.') from exc


@app.route('/auth/gmail')
@login_required
@admin_required
def gmail_auth():
    from flask import session
    if not GMAIL_OAUTH_CLIENT_ID or not GMAIL_OAUTH_CLIENT_SECRET:
        return jsonify({'error': 'Gmail OAuth is not configured.'}), 500
    oauth_state = secrets.token_urlsafe(32)
    session['gmail_oauth_state'] = oauth_state
    session['gmail_workspace_admin_id'] = workspace_admin_id_for(current_user)
    params = {
        'client_id': GMAIL_OAUTH_CLIENT_ID,
        'redirect_uri': _gmail_redirect_uri(),
        'response_type': 'code',
        'scope': ' '.join(GMAIL_OAUTH_SCOPES),
        'access_type': 'offline',
        'prompt': 'consent',
        'state': oauth_state,
    }
    return redirect('https://accounts.google.com/o/oauth2/v2/auth?' + urlencode(params))


@app.route('/auth/gmail/callback')
@login_required
@admin_required
def gmail_auth_callback():
    from flask import session
    returned_state = request.args.get('state')
    expected_state = session.pop('gmail_oauth_state', None)
    workspace_admin_id = session.pop('gmail_workspace_admin_id', None)
    if not returned_state or returned_state != expected_state or str(workspace_admin_id) != str(workspace_admin_id_for(current_user)):
        flash('Gmail authorization security verification failed.', 'danger')
        return redirect(url_for('index'))
    if request.args.get('error') or not request.args.get('code'):
        flash('Gmail authorization was cancelled or denied.', 'danger')
        return redirect(url_for('index'))
    token_response = requests.post('https://oauth2.googleapis.com/token', data={
        'client_id': GMAIL_OAUTH_CLIENT_ID,
        'client_secret': GMAIL_OAUTH_CLIENT_SECRET,
        'code': request.args['code'],
        'grant_type': 'authorization_code',
        'redirect_uri': _gmail_redirect_uri(),
    }, timeout=20)
    if token_response.status_code != 200:
        flash('Gmail authorization token exchange failed.', 'danger')
        return redirect(url_for('index'))
    token_data = token_response.json()
    if not token_data.get('refresh_token'):
        flash('Gmail authorization did not return a refresh token. Reconnect with consent.', 'danger')
        return redirect(url_for('index'))
    mailbox_email = None
    try:
        profile = requests.get('https://gmail.googleapis.com/gmail/v1/users/me/profile', headers={
            'Authorization': f"Bearer {token_data.get('access_token', '')}"
        }, timeout=20)
        if profile.ok:
            mailbox_email = profile.json().get('emailAddress')
    except requests.RequestException:
        pass
    persistent_token = {
        'refresh_token': token_data['refresh_token'],
        'token_uri': 'https://oauth2.googleapis.com/token',
        'client_id': GMAIL_OAUTH_CLIENT_ID,
        'client_secret': GMAIL_OAUTH_CLIENT_SECRET,
        'scopes': GMAIL_OAUTH_SCOPES,
    }
    try:
        encrypted_token = _encrypt_gmail_token(persistent_token)
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('index'))
    if not sheets_db.save_gmail_integration(workspace_admin_id, encrypted_token, mailbox_email):
        flash('Could not save the Gmail mailbox connection.', 'danger')
        return redirect(url_for('index'))
    flash('Gmail mailbox connected successfully.', 'success')
    return redirect(url_for('index'))


@app.route('/auth/gmail/status')
@login_required
def gmail_auth_status():
    workspace_admin_id = workspace_admin_id_for(current_user)
    integration = sheets_db.get_gmail_integration(workspace_admin_id) if workspace_admin_id else None
    return jsonify({
        'connected': bool(integration),
        'mailbox_email': integration.get('mailbox_email') if integration else None,
        'can_manage': is_admin_user(current_user),
        'can_sync': has_user_permission('gmail_sync', current_user),
    })


@app.route('/auth/gmail/disconnect', methods=['POST'])
@login_required
@admin_required
def gmail_auth_disconnect():
    sheets_db.delete_gmail_integration(workspace_admin_id_for(current_user))
    return jsonify({'success': True})


def _extract_gmail_text(payload):
    plain_parts, html_parts = [], []

    def collect(part):
        mime_type = part.get('mimeType', '')
        body = part.get('body') or {}
        data = body.get('data')
        if data and mime_type in {'text/plain', 'text/html'}:
            try:
                text = base64.urlsafe_b64decode(data + '=' * (-len(data) % 4)).decode('utf-8', errors='replace')
                (plain_parts if mime_type == 'text/plain' else html_parts).append(text)
            except (ValueError, TypeError):
                pass
        for child in part.get('parts') or []:
            collect(child)

    collect(payload or {})
    if plain_parts:
        return '\n'.join(plain_parts).strip()
    html_text = '\n'.join(html_parts)
    return unescape(re.sub(r'<[^>]+>', ' ', html_text)).strip()


def _gmail_headers(payload):
    headers = {str(item.get('name', '')).lower(): item.get('value', '') for item in (payload or {}).get('headers', [])}
    sender = parseaddr(headers.get('from', ''))[1].lower()
    received_at = None
    try:
        received_at = parsedate_to_datetime(headers.get('date', ''))
        if received_at and received_at.tzinfo:
            received_at = received_at.astimezone().replace(tzinfo=None)
    except (TypeError, ValueError, IndexError):
        pass
    return sender, headers.get('subject', ''), received_at


def _parse_claim_intimation_with_gemini(email_text, subject, user):
    prompt = f"""Extract only the following motor-insurance intimation fields from this email.
Return strict JSON with exactly these keys: claim_no, vehicle_no, insured_name, policy_no, insurer, date_of_loss.
Use empty strings for unavailable fields. Do not invent values.

Subject: {subject}
Email body:
{email_text[:30000]}
"""
    primary, secondary = get_generative_models(user)
    errors = []
    for selected_model in (primary, secondary):
        try:
            response = selected_model.generate_content(prompt)
            content = getattr(response, 'text', '') or ''
            cleaned = re.sub(r'^\s*```(?:json)?\s*|\s*```\s*$', '', content.strip(), flags=re.IGNORECASE)
            parsed = json.loads(cleaned)
            if not isinstance(parsed, dict):
                raise ValueError('Model returned a non-object response.')
            return {key: str(parsed.get(key, '') or '').strip() for key in (
                'claim_no', 'vehicle_no', 'insured_name', 'policy_no', 'insurer', 'date_of_loss'
            )}
        except Exception as exc:
            errors.append(str(exc))
    raise ValueError(errors[-1] if errors else 'No model response was available.')


def _gmail_service_for_workspace(workspace_admin_id):
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GoogleRequest
    from googleapiclient.discovery import build
    integration = sheets_db.get_gmail_integration(workspace_admin_id)
    if not integration:
        raise ValueError('No Gmail mailbox is connected for this workspace.')
    token_data = _decrypt_gmail_token(integration.get('encrypted_token', ''))
    credentials = Credentials(
        token=None,
        refresh_token=token_data.get('refresh_token'),
        token_uri=token_data.get('token_uri'),
        client_id=token_data.get('client_id'),
        client_secret=token_data.get('client_secret'),
        scopes=token_data.get('scopes') or GMAIL_OAUTH_SCOPES,
    )
    credentials.refresh(GoogleRequest())
    return build('gmail', 'v1', credentials=credentials, cache_discovery=False)


def _list_unread_gmail_messages(gmail):
    """Fetch every unread message in the configured lookback window without changing Gmail state."""
    messages = []
    page_token = None
    while True:
        request_args = {
            'userId': 'me',
            'q': f'is:unread newer_than:{GMAIL_SYNC_LOOKBACK_DAYS}d',
            'maxResults': 500,
        }
        if page_token:
            request_args['pageToken'] = page_token
        page = gmail.users().messages().list(**request_args).execute()
        messages.extend(page.get('messages', []))
        page_token = page.get('nextPageToken')
        if not page_token:
            return messages


@app.route('/api/gmail/sync', methods=['POST'])
@login_required
@gmail_sync_required
def sync_gmail_intimations():
    workspace_admin_id = workspace_admin_id_for(current_user)
    data = request.get_json(silent=True) or {}
    sender_domain = str(data.get('sender_domain', '')).lower().strip().lstrip('@')
    available_domains = [str(row.get('domain', '')).lower() for row in sheets_db.get_gmail_sender_domains(workspace_admin_id)]
    if not available_domains:
        return jsonify({'error': 'Add at least one approved sender domain before syncing Gmail.'}), 400
    if sender_domain and sender_domain not in available_domains:
        return jsonify({'error': 'Sender domain is not approved for this workspace.'}), 400
    try:
        gmail = _gmail_service_for_workspace(workspace_admin_id)
        message_summaries = _list_unread_gmail_messages(gmail)
    except Exception as exc:
        return jsonify({'error': f'Gmail sync could not start: {exc}'}), 502

    summary = {'created': 0, 'merged': 0, 'skipped': 0, 'failed': 0}
    for summary_item in message_summaries:
        message_id = summary_item.get('id')
        if not message_id:
            continue
        if sheets_db.get_gmail_sync_message(message_id):
            summary['skipped'] += 1
            continue
        try:
            message = gmail.users().messages().get(userId='me', id=message_id, format='full').execute()
            payload = message.get('payload') or {}
            sender, subject, received_at = _gmail_headers(payload)
            actual_sender_domain = sender.rsplit('@', 1)[-1] if '@' in sender else ''
            if actual_sender_domain not in available_domains or (sender_domain and actual_sender_domain != sender_domain):
                summary['skipped'] += 1
                continue
            email_text = _extract_gmail_text(payload)
            parsed = _parse_claim_intimation_with_gemini(email_text, subject, current_user)
            claim_no = parsed.get('claim_no', '')
            if not claim_no:
                sheets_db.record_gmail_sync_message(message_id, workspace_admin_id, sender_email=sender,
                    subject=subject, received_at=received_at, parse_data=parsed,
                    sync_status='failed', error_message='Claim number was not extracted.')
                summary['failed'] += 1
                continue
            existing = sheets_db.find_workspace_report_by_claim_no(workspace_admin_id, claim_no)
            if existing:
                report_data = existing.get('report_data_json') or {}
                if isinstance(report_data, str):
                    report_data = json.loads(report_data or '{}')
                survey = report_data.setdefault('survey_report', {})
                fields = {
                    'claim_no': parsed['claim_no'], 'vehicle_regn_no': parsed['vehicle_no'],
                    'insured': parsed['insured_name'], 'policy_no': parsed['policy_no'],
                    'insurer': parsed['insurer'], 'date_of_loss': parsed['date_of_loss'],
                }
                for key, value in fields.items():
                    if value and not survey.get(key):
                        survey[key] = value
                history = report_data.setdefault('gmail_intimations', [])
                history.append({'message_id': message_id, 'sender': sender, 'subject': subject,
                                'received_at': received_at.isoformat() if received_at else None})
                report_id = sheets_db.save_workspace_report(
                    current_user.id, workspace_admin_id, report_data, existing_report_id=existing.get('id'),
                    status=existing.get('status', 'documents_awaited'),
                    survey_type=existing.get('survey_type', 'final'))
                summary['merged'] += 1
            else:
                survey_type = 'spot' if re.search(r'\b(spot|preliminary)\b', f'{subject}\n{email_text}', re.I) else 'final'
                prefix = _report_prefix_for_insurer(parsed.get('insurer'))
                sequence = sheets_db.reserve_report_number(workspace_admin_id, prefix, str(datetime.now().year))
                if sequence is None:
                    raise ValueError('Could not reserve a report number.')
                report_data = {
                    'survey_report': {
                        'report_no': f'{prefix}/{datetime.now().year}/{sequence:02d}',
                        'claim_no': parsed['claim_no'], 'vehicle_regn_no': parsed['vehicle_no'],
                        'insured': parsed['insured_name'], 'policy_no': parsed['policy_no'],
                        'insurer': parsed['insurer'], 'date_of_loss': parsed['date_of_loss'],
                    },
                    'assessment': {'report_type': 'Spot Report' if survey_type == 'spot' else 'Final Survey Report'},
                    'photos': {},
                    'claim_meta': {'status': 'documents_awaited', 'survey_type': survey_type},
                    'gmail_intimations': [{'message_id': message_id, 'sender': sender, 'subject': subject,
                                           'received_at': received_at.isoformat() if received_at else None}],
                }
                report_id = sheets_db.save_workspace_report(
                    current_user.id, workspace_admin_id, report_data, status='documents_awaited',
                    survey_type=survey_type, gmail_message_id=message_id, email_received_date=received_at)
                if not report_id:
                    raise ValueError('Could not save the generated report draft.')
                summary['created'] += 1
            sheets_db.record_gmail_sync_message(message_id, workspace_admin_id, report_id=report_id,
                sender_email=sender, subject=subject, received_at=received_at, parse_data=parsed)
        except Exception as exc:
            sheets_db.record_gmail_sync_message(message_id, workspace_admin_id,
                sync_status='failed', error_message=str(exc))
            summary['failed'] += 1
    return jsonify(summary)

@app.route('/api/gmail/intimations', methods=['GET'])
@login_required
@gmail_sync_required
def get_pending_gmail_intimations():
    """Returns list of pending Gmail intimations waiting for user approval to add or cancel."""
    workspace_admin_id = workspace_admin_id_for(current_user)
    pending = sheets_db.get_pending_gmail_messages(workspace_admin_id)
    return jsonify({'intimations': pending})

@app.route('/api/gmail/intimation/<message_id>/add', methods=['POST'])
@login_required
@gmail_sync_required
def add_gmail_intimation_to_register(message_id):
    """Extracts and adds a specific Gmail intimation to the claim register upon user confirmation."""
    workspace_admin_id = workspace_admin_id_for(current_user)
    sync_record = sheets_db.get_gmail_sync_message(message_id)
    if not sync_record:
        return jsonify({'error': 'Gmail intimation record not found.'}), 404
        
    parse_data = sync_record.get('parse_data_json') or {}
    if isinstance(parse_data, str):
        try: parse_data = json.loads(parse_data)
        except Exception: parse_data = {}

    claim_no = parse_data.get('claim_no', '')
    sender = sync_record.get('sender_email', '')
    subject = sync_record.get('subject', '')
    received_at = sync_record.get('received_at')

    if not claim_no:
        return jsonify({'error': 'Claim number missing in intimation. Cannot add automatically.'}), 400

    existing = sheets_db.find_workspace_report_by_claim_no(workspace_admin_id, claim_no)
    if existing:
        report_data = existing.get('report_data_json') or {}
        if isinstance(report_data, str):
            report_data = json.loads(report_data or '{}')
        survey = report_data.setdefault('survey_report', {})
        fields = {
            'claim_no': parse_data.get('claim_no'), 'vehicle_regn_no': parse_data.get('vehicle_no'),
            'insured': parse_data.get('insured_name'), 'policy_no': parse_data.get('policy_no'),
            'insurer': parse_data.get('insurer'), 'date_of_loss': parse_data.get('date_of_loss'),
        }
        for key, value in fields.items():
            if value and not survey.get(key):
                survey[key] = value
        history = report_data.setdefault('gmail_intimations', [])
        history.append({'message_id': message_id, 'sender': sender, 'subject': subject,
                        'received_at': received_at.isoformat() if hasattr(received_at, 'isoformat') else str(received_at)})
        report_id = sheets_db.save_workspace_report(
            current_user.id, workspace_admin_id, report_data, existing_report_id=existing.get('id'),
            status=existing.get('status', 'documents_awaited'),
            survey_type=existing.get('survey_type', 'final'))
        action = 'merged'
    else:
        survey_type = 'spot' if re.search(r'\b(spot|preliminary)\b', f'{subject}\n{parse_data.get("snippet", "")}', re.I) else 'final'
        prefix = _report_prefix_for_insurer(parse_data.get('insurer'))
        sequence = sheets_db.reserve_report_number(workspace_admin_id, prefix, str(datetime.now().year))
        if sequence is None:
            return jsonify({'error': 'Could not reserve report number.'}), 500
        report_data = {
            'survey_report': {
                'report_no': f'{prefix}/{datetime.now().year}/{sequence:02d}',
                'claim_no': parse_data.get('claim_no'), 'vehicle_regn_no': parse_data.get('vehicle_no'),
                'insured': parse_data.get('insured_name'), 'policy_no': parse_data.get('policy_no'),
                'insurer': parse_data.get('insurer'), 'date_of_loss': parse_data.get('date_of_loss'),
            },
            'assessment': {'report_type': 'Spot Report' if survey_type == 'spot' else 'Final Survey Report'},
            'photos': {},
            'claim_meta': {'status': 'documents_awaited', 'survey_type': survey_type},
            'gmail_intimations': [{'message_id': message_id, 'sender': sender, 'subject': subject,
                                   'received_at': received_at.isoformat() if hasattr(received_at, 'isoformat') else str(received_at)}],
        }
        report_id = sheets_db.save_workspace_report(
            current_user.id, workspace_admin_id, report_data, status='documents_awaited',
            survey_type=survey_type, gmail_message_id=message_id, email_received_date=received_at)
        action = 'created'

    sheets_db.record_gmail_sync_message(message_id, workspace_admin_id, report_id=report_id,
        sender_email=sender, subject=subject, received_at=received_at, parse_data=parse_data, sync_status='processed')

    return jsonify({'success': True, 'action': action, 'report_id': report_id, 'claim_no': claim_no})

@app.route('/api/gmail/intimation/<message_id>/cancel', methods=['POST'])
@login_required
@gmail_sync_required
def cancel_gmail_intimation(message_id):
    """Cancel/dismiss a synced Gmail intimation so it is ignored."""
    success = sheets_db.cancel_gmail_sync_message(message_id)
    if not success:
        return jsonify({'error': 'Could not cancel Gmail intimation.'}), 400
    return jsonify({'success': True, 'message': 'Gmail intimation cancelled.'})

def generate_pending_documents_reminder_text(claim_data, pending_docs, reminder_count=1):
    """
    Generates notification text formatted according to client specification for 1st, 2nd, and 3rd reminders.
    """
    survey = claim_data.get('survey_report', {}) if isinstance(claim_data.get('survey_report'), dict) else {}
    claim_no = claim_data.get('claim_no') or survey.get('claim_no', '[Claim Number]')
    policy_no = claim_data.get('policy_no') or survey.get('policy_no', '[Policy Number]')
    insured_name = claim_data.get('insured_name') or survey.get('insured', '[Customer Name]')
    vehicle_no = claim_data.get('vehicle_no') or survey.get('vehicle_regn_no', '[Vehicle Number]')
    insurer_name = claim_data.get('insurer') or survey.get('insurer', '[Insurance Company Name]')

    docs_list_str = "\n".join([f"{idx+1}. {doc}" for idx, doc in enumerate(pending_docs)]) if pending_docs else "1. Policy copy\n2. Duly completed and signed claim form"

    body = f"Dear Sir/Madam,\n\nThis is regarding the motor insurance claim mentioned below:\n\n"
    body += f"Claim Number: {claim_no}\n"
    body += f"Policy Number: {policy_no}\n"
    body += f"Insured Name: {insured_name}\n"
    body += f"Vehicle Registration Number: {vehicle_no}\n"
    body += f"Insurance Company: {insurer_name}\n\n"
    body += f"During the scrutiny of the claim documents, it has been observed that the following documents are still pending:\n\n"
    body += f"{docs_list_str}\n\n"
    body += "You are requested to submit clear and legible copies of the above documents at the earliest so that the survey report and claim assessment process can be completed without further delay.\n\n"
    body += "Please mention the claim number and vehicle registration number while sending the documents.\n\n"

    if reminder_count == 2:
        body += "Kindly note that this is the second time reminder, so please treat this with high priority; otherwise we assume you are not interested in taking the claim, and the insurance company may close the claim without further notice.\n\n"
    elif reminder_count >= 3:
        body += "Kindly note that this is the third time reminder, so please treat this with high priority; otherwise we assume you are not interested in taking the claim, and the insurance company may close the claim without further notice.\n\n"
    else:
        body += "Kindly note that any delay in submitting the required documents may delay the processing of your claim.\n\n"

    body += "Regards,\nSk Anowar Ali\nMotor Surveyor & Loss Assessor\nLicence No.: SLA-121784\nMobile: 8777370714"
    return body

@app.route('/api/claims/<report_id>/pending_documents', methods=['GET', 'POST'])
@login_required
def manage_pending_documents(report_id):
    """Get or update pending documents checklist for a claim."""
    workspace_admin_id = workspace_admin_id_for(current_user)
    report = sheets_db.get_report_by_id(report_id, current_user.id)
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
        
    report_data = report.get('report_data_json') or {}
    if isinstance(report_data, str):
        report_data = json.loads(report_data or '{}')
        
    if request.method == 'GET':
        pending = report_data.get('pending_documents', [
            {'name': 'Claim Form', 'received': False},
            {'name': 'RC Copy', 'received': False},
            {'name': 'Driving License (DL)', 'received': False},
            {'name': 'Road Tax Permit', 'received': False},
            {'name': 'Permit A', 'received': False},
            {'name': 'Permit B', 'received': False},
            {'name': 'Repair Estimate Copy', 'received': False},
            {'name': 'Other Supporting Documents', 'received': False}
        ])
        reminder_info = sheets_db.get_claim_reminder(report_id) or {}
        return jsonify({'pending_documents': pending, 'reminder_info': reminder_info})
        
    data = request.get_json() or {}
    pending_docs = data.get('pending_documents', [])
    report_data['pending_documents'] = pending_docs
    
    sheets_db.save_workspace_report(
        current_user.id, workspace_admin_id, report_data, existing_report_id=report_id,
        status=report.get('status', 'documents_awaited'),
        survey_type=report.get('survey_type', 'final'))
        
    return jsonify({'success': True, 'pending_documents': pending_docs})

@app.route('/api/claims/<report_id>/send_reminder', methods=['POST'])
@login_required
def send_pending_documents_reminder(report_id):
    """Send document pending reminder notification (Email, WhatsApp/SMS text formatting, Claim Manager option)."""
    workspace_admin_id = workspace_admin_id_for(current_user)
    report = sheets_db.get_report_by_id(report_id, current_user.id)
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
        
    data = request.get_json() or {}
    claim_manager_email = data.get('claim_manager_email', '').strip()
    claim_manager_phone = data.get('claim_manager_phone', '').strip()
    
    reminder_info = sheets_db.get_claim_reminder(report_id) or {}
    current_count = reminder_info.get('reminder_count', 0)
    new_count = current_count + 1
    
    if new_count > 3:
        return jsonify({'error': 'Maximum 3 reminders already sent for this claim.'}), 400
        
    report_data = report.get('report_data_json') or {}
    if isinstance(report_data, str):
        report_data = json.loads(report_data or '{}')
        
    pending_docs_raw = report_data.get('pending_documents', [])
    pending_doc_names = [d.get('name') if isinstance(d, dict) else str(d) for d in pending_docs_raw if isinstance(d, dict) and not d.get('received')]
    if not pending_doc_names:
        pending_doc_names = [
            "Policy copy",
            "Duly completed and signed claim form",
            "Repairer's final tax invoice and payment receipt",
            "Clear bank details/cancelled cheque of the insured"
        ]
        
    formatted_message = generate_pending_documents_reminder_text(report, pending_doc_names, new_count)
    
    sheets_db.update_claim_reminder(
        report_id, workspace_admin_id, report.get('claim_no', ''),
        new_count, claim_manager_email=claim_manager_email, claim_manager_phone=claim_manager_phone
    )
    
    return jsonify({
        'success': True,
        'reminder_count': new_count,
        'message_text': formatted_message,
        'max_reached': new_count >= 3
    })

@app.route('/get_user_upload_url', methods=['POST'])
@login_required
def get_user_upload_url():
    """Get upload URL using user's OAuth token (for large files)."""
    from flask import session
    access_token = session.get('google_access_token')
    if not access_token:
        return jsonify({'error': 'Not connected to Google Drive. Please connect first.'}), 401
    
    data = request.get_json()
    filename = data.get('filename')
    mime_type = data.get('mime_type', 'application/pdf')
    
    if not filename:
        return jsonify({'error': 'Filename required'}), 400
    
    # Retrieve the origin from the frontend request to bind the Google CORS policy
    origin = request.headers.get('Origin')
    
    # Initiate resumable upload with user's token
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
        'X-Upload-Content-Type': mime_type,
    }
    
    if origin:
        headers['Origin'] = origin
    
    metadata = {
        'name': filename,
        'mimeType': mime_type
    }
    
    response = requests.post(
        'https://www.googleapis.com/upload/drive/v3/files?uploadType=resumable',
        headers=headers,
        json=metadata
    )
    
    if response.status_code == 200:
        upload_url = response.headers.get('Location')
        return jsonify({'url': upload_url, 'access_token': access_token})
    elif response.status_code == 401:
        # Token expired, try to refresh
        refresh_token = session.get('google_refresh_token')
        if refresh_token:
            # Attempt token refresh
            refresh_response = requests.post(
                'https://oauth2.googleapis.com/token',
                data={
                    'client_id': GOOGLE_OAUTH_CLIENT_ID,
                    'client_secret': GOOGLE_OAUTH_CLIENT_SECRET,
                    'refresh_token': refresh_token,
                    'grant_type': 'refresh_token'
                }
            )
            if refresh_response.status_code == 200:
                new_tokens = refresh_response.json()
                session['google_access_token'] = new_tokens.get('access_token')
                # Retry upload URL request
                if origin:
                    headers['Origin'] = origin
                headers['Authorization'] = f"Bearer {new_tokens.get('access_token')}"
                retry_response = requests.post(
                    'https://www.googleapis.com/upload/drive/v3/files?uploadType=resumable',
                    headers=headers,
                    json=metadata
                )
                if retry_response.status_code == 200:
                    upload_url = retry_response.headers.get('Location')
                    return jsonify({'url': upload_url, 'access_token': new_tokens.get('access_token')})
        
        session.pop('google_access_token', None)
        return jsonify({'error': 'Google Drive authorization expired. Please reconnect.'}), 401
    else:
        print(f"Failed to get user upload URL: {response.text}")
        return jsonify({'error': f'Failed to initiate upload: {response.status_code}'}), 500

# --- Main Application Routes ---
@app.route('/')
@login_required
def index():
    return render_template('index.html')



@app.route('/get_gemini_upload_url', methods=['POST'])
@login_required
def get_gemini_upload_url():
    data = request.get_json(silent=True) or {}
    filename = secure_filename(data.get('filename', 'document.pdf')) or 'document.pdf'
    mime_type = data.get('mime_type', 'application/pdf')
    try:
        size = int(data.get('size', 0))
    except (TypeError, ValueError):
        size = 0

    if mime_type != 'application/pdf' or size <= 0 or size > app.config['MAX_CONTENT_LENGTH']:
        return jsonify({'error': 'Only PDF uploads up to 100 MB are supported.'}), 400
    
    api_key = current_user.gemini_api_key or os.getenv('GEMINI_API_KEY')
    if not api_key:
        return jsonify({"error": "No API key"}), 500

    upload_session = sheets_db.create_upload_session(
        current_user.id, 'gemini', filename, mime_type, size
    )
    if not isinstance(upload_session, dict) or not upload_session.get('id'):
        return jsonify({'error': 'Unable to create an upload record.'}), 503
        
    origin = request.headers.get('Origin')
        
    headers = {
        'X-Goog-Upload-Protocol': 'resumable',
        'X-Goog-Upload-Command': 'start',
        'X-Goog-Upload-Header-Content-Length': str(size),
        'X-Goog-Upload-Header-Content-Type': mime_type,
        'Content-Type': 'application/json'
    }
    
    if origin:
        headers['Origin'] = origin
    
    metadata = {
        'file': {
            'display_name': filename
        }
    }
    
    url = f"https://generativelanguage.googleapis.com/upload/v1beta/files?key={api_key}"
    resp = requests.post(url, headers=headers, json=metadata)
    
    if resp.status_code != 200:
        return jsonify({"error": f"Failed to get upload URL: {resp.text}"}), 500
        
    upload_url = resp.headers.get('X-Goog-Upload-URL')
    if not upload_url:
        return jsonify({'error': 'Gemini did not return an upload URL.'}), 502
    return jsonify({"url": upload_url, "upload_id": str(upload_session['id'])})

def download_drive_file_with_token(access_token, file_id):
    """Downloads file content from Drive using the user's OAuth access token."""
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(
            f'https://www.googleapis.com/drive/v3/files/{file_id}?alt=media',
            headers=headers
        )
        if response.status_code == 200:
            return response.content
        else:
            print(f"Failed to download file {file_id} with user token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Error downloading file with user token: {e}")
        return None

def _find_or_create_drive_folder(access_token, folder_name, parent_id=None):
    """Finds a folder by name in Drive (under optional parent), or creates it."""
    query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"
    
    headers = {'Authorization': f'Bearer {access_token}'}
    search_resp = requests.get(
        'https://www.googleapis.com/drive/v3/files',
        headers=headers,
        params={'q': query, 'fields': 'files(id,name)', 'spaces': 'drive'}
    )
    
    if search_resp.status_code == 200:
        files = search_resp.json().get('files', [])
        if files:
            return files[0]['id']
    
    # Create the folder
    metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder'
    }
    if parent_id:
        metadata['parents'] = [parent_id]
    
    create_resp = requests.post(
        'https://www.googleapis.com/drive/v3/files',
        headers={**headers, 'Content-Type': 'application/json'},
        json=metadata
    )
    
    if create_resp.status_code == 200:
        return create_resp.json().get('id')
    
    print(f"Failed to create folder '{folder_name}': {create_resp.status_code} - {create_resp.text}")
    return None

def upload_pdf_to_drive(access_token, pdf_bytes, filename, folder_name):
    """
    Uploads a PDF to the user's Drive inside 'Survey Reports/{folder_name}/'.
    If a file with the same name exists in that folder, it is replaced.
    Returns the web view link on success, None on failure.
    """
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        
        # 1. Find or create "Survey Reports" root folder
        root_folder_id = _find_or_create_drive_folder(access_token, 'Survey Reports')
        if not root_folder_id:
            print("Failed to find/create 'Survey Reports' folder")
            return None
        
        # 2. Find or create vehicle subfolder
        vehicle_folder_id = _find_or_create_drive_folder(access_token, folder_name, root_folder_id)
        if not vehicle_folder_id:
            print(f"Failed to find/create '{folder_name}' folder")
            return None
        
        # 3. Check if file already exists in that folder (to update instead of duplicate)
        query = f"name='{filename}' and '{vehicle_folder_id}' in parents and trashed=false"
        search_resp = requests.get(
            'https://www.googleapis.com/drive/v3/files',
            headers=headers,
            params={'q': query, 'fields': 'files(id)', 'spaces': 'drive'}
        )
        
        existing_file_id = None
        if search_resp.status_code == 200:
            files = search_resp.json().get('files', [])
            if files:
                existing_file_id = files[0]['id']
        
        # 4. Upload or update the PDF
        if existing_file_id:
            # Update existing file content
            upload_resp = requests.patch(
                f'https://www.googleapis.com/upload/drive/v3/files/{existing_file_id}?uploadType=media&fields=id,webViewLink',
                headers={**headers, 'Content-Type': 'application/pdf'},
                data=pdf_bytes
            )
        else:
            # Create new file with multipart upload
            import json as json_module
            boundary = '----DriveUploadBoundary'
            metadata = json_module.dumps({
                'name': filename,
                'parents': [vehicle_folder_id],
                'mimeType': 'application/pdf'
            })
            
            body = (
                f'--{boundary}\r\n'
                f'Content-Type: application/json; charset=UTF-8\r\n\r\n'
                f'{metadata}\r\n'
                f'--{boundary}\r\n'
                f'Content-Type: application/pdf\r\n\r\n'
            ).encode('utf-8') + pdf_bytes + f'\r\n--{boundary}--'.encode('utf-8')
            
            upload_resp = requests.post(
                'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&fields=id,webViewLink',
                headers={**headers, 'Content-Type': f'multipart/related; boundary={boundary}'},
                data=body
            )
        
        if upload_resp.status_code in (200, 201):
            result = upload_resp.json()
            return result.get('webViewLink', f"https://drive.google.com/file/d/{result.get('id')}/view")
        else:
            print(f"Drive upload failed: {upload_resp.status_code} - {upload_resp.text}")
            return None
            
    except Exception as e:
        print(f"Error uploading PDF to Drive: {e}")
        import traceback; traceback.print_exc()
        return None

@app.route('/upload_report_to_drive', methods=['POST'])
@login_required
def upload_report_to_drive():
    """Upload the generated report PDF to Google Drive (service account)."""
    data = request.get_json()
    request_id = data.get('request_id')
    
    if not request_id or request_id not in generated_data_store:
        return jsonify({'error': 'Report not found. Please generate the report first.'}), 404
    
    report_data = generated_data_store[request_id]
    pdf_bytes = report_data.get('pdf_report')
    vehicle_no = report_data.get('vehicle_no', '').strip()
    report_no = report_data.get('report_no', 'SurveyReport')
    
    if not pdf_bytes:
        return jsonify({'error': 'No PDF data found for this report.'}), 400
    
    filename_base = "".join(c for c in vehicle_no if c.isalnum() or c in ('_', '-')).rstrip() if vehicle_no else report_no.replace(' ', '_').replace('/', '-')
    filename = f"{filename_base}.pdf"
    folder_name_to_use = "".join(c for c in vehicle_no if c.isalnum() or c in ('_', '-', ' ')).strip() if vehicle_no else 'Unknown_Vehicle'
    
    from flask import session
    access_token = session.get('google_access_token')
    
    drive_link = None
    if access_token:
        # Upload to user's personal drive
        drive_link = upload_pdf_to_drive(access_token, pdf_bytes, filename, folder_name_to_use)
        
    if not drive_link:
        # Fallback to service account
        drive_link = sheets_db.upload_report_pdf(pdf_bytes, filename, vehicle_no if vehicle_no else 'Unknown_Vehicle')
    
    if drive_link:
        return jsonify({'success': True, 'drive_link': drive_link, 'message': f'Report uploaded to Drive!'})
    else:
        return jsonify({'error': 'Failed to upload report to Google Drive. Check server logs.'}), 500

@app.route('/api/extract_fee_pdf', methods=['POST'])
@login_required
def extract_fee_pdf():
    """Extract billing fields (Insurer, Insured Name, Vehicle No, Policy No) from uploaded PDF for Survey Fee Register."""
    if 'pdf_file' not in request.files and 'fee_pdf_file' not in request.files:
        return jsonify({'error': 'No PDF file uploaded'}), 400

    file_obj = request.files.get('pdf_file') or request.files.get('fee_pdf_file')
    if not file_obj or file_obj.filename == '':
        return jsonify({'error': 'No PDF file selected'}), 400

    try:
        content = file_obj.read()
        pdf_part = {'mime_type': file_obj.mimetype or 'application/pdf', 'data': content}
        api_key = getattr(current_user, 'gemini_api_key', None) or os.getenv("GEMINI_API_KEY")
        user_model = getattr(current_user, 'gemini_model', None)

        from modules.gemini import execute_gemini_task
        extracted_result = execute_gemini_task(
            api_key=api_key,
            pdf_part=pdf_part,
            user_model=user_model,
            is_invoice=True
        )

        survey = extracted_result.get('survey_report') or extracted_result.get('extracted') or {}
        extracted = {
            'insurer': survey.get('insurer') or survey.get('insurance_company') or '',
            'insured': survey.get('insured') or survey.get('insured_name') or '',
            'vehicle_no': survey.get('vehicle_regn_no') or survey.get('vehicle_no') or '',
            'policy_no': survey.get('policy_no') or '',
            'claim_no': survey.get('claim_no') or '',
            'report_no': survey.get('report_no') or '',
            'invoice_date': survey.get('report_date') or datetime.now().strftime('%Y-%m-%d')
        }
        return jsonify({'success': True, 'extracted': extracted})
    except Exception as e:
        print(f"Error extracting fee PDF: {e}")
        return jsonify({'error': f'Failed to process PDF: {e}'}), 500


@app.route('/process_pdf', methods=['POST'])
@login_required
def process_pdf():
    """Accept a PDF and start async Gemini processing. Returns a task_id for polling."""
    # --- Extract everything from the request context NOW (before background thread) ---
    pdf_part = None
    user_id = current_user.id
    # Snapshot user object data for the background thread
    user_data_snapshot = {
        'id': current_user.id,
        'username': current_user.username,
        'password_hash': current_user.password_hash,
        'full_name': current_user.full_name,
        'qualifications': current_user.qualifications,
        'designation': current_user.designation,
        'license_no': current_user.license_no,
        'expiry_date': current_user.expiry_date,
        'membership_no': current_user.membership_no,
        'address_line_1': current_user.address_line_1,
        'address_line_2': current_user.address_line_2,
        'address_line_3': current_user.address_line_3,
        'contact_no': current_user.contact_no,
        'email': current_user.email,
        'gemini_api_key': current_user.gemini_api_key,
        'gemini_model': current_user.gemini_model,
    }

    if request.content_type == 'application/json':
        data = request.get_json()
        mime_type = data.get('mime_type', 'application/pdf')
        
        if 'drive_file_id' in data:
            drive_file_id = data['drive_file_id']
            from flask import session
            access_token = session.get('google_access_token')
            if not access_token:
                 return jsonify({"error": "Google Drive not connected. Please reconnect in Profile Settings."}), 401
                 
            print(f"Downloading file {drive_file_id} from user's Drive...")
            pdf_bytes = download_drive_file_with_token(access_token, drive_file_id)
            if not pdf_bytes:
                 return jsonify({"error": "Failed to download file from Google Drive. Ensure the file still exists."}), 500
                 
            print("Successfully downloaded from Drive. Uploading to Gemini...")
            import tempfile
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                temp_file.write(pdf_bytes)
                temp_filename = temp_file.name
                
            try:
                uploaded_file = genai.upload_file(path=temp_filename, mime_type=mime_type)
                print(f"Uploaded to Gemini URI: {uploaded_file.uri}")
                pdf_part = {
                    "file_data": {
                        "mime_type": mime_type,
                        "file_uri": uploaded_file.uri
                    }
                }
            except Exception as e:
                print(f"Error uploading to Gemini: {e}")
                return jsonify({"error": "Failed to upload file to AI limits."}), 500
            finally:
                if os.path.exists(temp_filename):
                    os.remove(temp_filename)
                    
        elif 'gemini_file_uri' in data:
            gemini_file_uri = data['gemini_file_uri']
            print(f"Using direct Gemini URI: {gemini_file_uri}")
            pdf_part = {
                "file_data": {
                    "mime_type": mime_type,
                    "file_uri": gemini_file_uri
                }
            }
        else:
            return jsonify({"error": "No drive_file_id or gemini_file_uri provided"}), 400
             
    elif 'pdf_file' in request.files:
        file = request.files['pdf_file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        if file and file.mimetype == 'application/pdf':
            pdf_content = file.read()
            pdf_part = {"mime_type": "application/pdf", "data": pdf_content}
        else:
             return jsonify({"error": "Invalid file type. Please upload a PDF."}), 400
    else:
         return jsonify({"error": "No file provided"}), 400

    # --- Dispatch to background thread ---
    task_id = _create_task()
    _task_executor.submit(_process_pdf_worker, task_id, pdf_part, user_data_snapshot, user_id)
    return jsonify({"task_id": task_id}), 202


def _process_pdf_worker(task_id, pdf_part, user_data_snapshot, user_id):
    """Background worker for Gemini PDF processing. Runs outside request context."""
    try:
        prompt = build_gemini_prompt()
        prompt_part = {"text": prompt}

        # Reconstruct a User object for get_generative_models
        user_obj = User(user_data_snapshot)
        dynamic_model, dynamic_secondary = get_generative_models(user_obj)
        try:
            response = dynamic_model.generate_content([prompt_part, pdf_part], stream=False)
        except ResourceExhausted as e:
            print(f"Primary model hit rate limit: {e}. Switching to secondary model.")
            response = dynamic_secondary.generate_content([prompt_part, pdf_part], stream=False)

        # Handle potential lack of response parts or blocked content
        if not response.parts:
                if response.prompt_feedback and response.prompt_feedback.block_reason:
                    reason = response.prompt_feedback.block_reason.name
                    print(f"Gemini response blocked. Reason: {reason}")
                    _fail_task(task_id, f"Content generation blocked due to safety settings ({reason}). Please check the input document.")
                    return
                else:
                    print("Gemini returned an empty response with no specific block reason.")
                    try:
                        response_text = response.text
                        if not response_text:
                            _fail_task(task_id, "Received an empty response from the AI model. Please try again or check the document.")
                            return
                        print("Received text despite no parts, attempting parse...")
                    except Exception:
                            _fail_task(task_id, "Received an empty or invalid response from the AI model. Please try again or check the document.")
                            return
        else:
                response_text = response.text

        # Parse the combined data
        combined_data = parse_gemini_response(response_text)

        # --- Apply Defaults to Survey Report Data if field is empty ---
        survey_data = combined_data.get('survey_report', {})
        if not survey_data.get('vehicle_pre_accident_condition'): survey_data['vehicle_pre_accident_condition'] = "Average"
        if not survey_data.get('dl_endorsement'): survey_data['dl_endorsement'] = "Not Known"
        if not survey_data.get('police_reported_to'): survey_data['police_reported_to'] = "Not Reported"
        if not survey_data.get('police_diary_case_no'): survey_data['police_diary_case_no'] = "N/A"
        if not survey_data.get('tp_details'): survey_data['tp_details'] = "No ( As Per Claim Form )"
        if not survey_data.get('damages_extent'): survey_data['damages_extent'] = "The Spare Parts which are included in Assessment column, found pressed/deformed/torn/ distorted &/or broken."
        if not survey_data.get('remark'): survey_data['remark'] = "The declaration of the accident appeared consistent with the nature of the damages sustained"
        # --- Inject Last Saved Surveyor Details ---
        last_surveyor_details = sheets_db.get_last_surveyor_details(user_id)
        if last_surveyor_details:
            if 'assessment' in combined_data:
                if 'page3_details' not in combined_data['assessment']:
                    combined_data['assessment']['page3_details'] = {}
                combined_data['assessment']['page3_details']['surveyor_details'] = last_surveyor_details

        _complete_task(task_id, combined_data)

    except ValueError as ve:
            print(f"Value Error during processing: {ve}")
            if "Failed to parse JSON response" in str(ve) or "unexpected error occurred during response parsing" in str(ve):
                _fail_task(task_id, "Failed to parse the AI response. Please try again or check the document.")
            else:
                _fail_task(task_id, "An error occurred while processing the document.")
    except genai.types.BlockedPromptException as bpe:
            print(f"Gemini API Error - Blocked Prompt: {bpe}")
            _fail_task(task_id, "Content generation blocked by API. Please check the document content.")
    except Exception as e:
        print(f"Error processing PDF with Gemini: {e}")
        import traceback
        traceback.print_exc()
        _fail_task(task_id, "An unexpected error occurred during AI processing. Please try again.")


@app.route('/process_pdf/status/<task_id>', methods=['GET'])
@login_required
def process_pdf_status(task_id):
    """Poll for the status of an async PDF processing task."""
    task = _get_task(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    
    if task["status"] == "completed":
        return jsonify({"status": "completed", "result": task["result"]})
    elif task["status"] == "error":
        return jsonify({"status": "error", "error": task["error"]})
    else:
        return jsonify({"status": "processing"})

# --- Depreciation Calculation Helper ---
def get_backend_depreciation_rate(part_type, vehicle_year_str):
    """
    Calculates depreciation rate based on part type and vehicle age in total months.
    The frontend will calculate total months and map it to a "year bucket" integer.
    """
    part_type = str(part_type).strip().upper()
    try:
        # The string now represents the "year bucket" from the rules (0, 1, 2, etc.)
        year_bucket = int(vehicle_year_str) if vehicle_year_str else 0
    except ValueError:
        year_bucket = 0 # Default to 0 if not a valid integer

    if part_type == 'G':
        return 0.0
    if part_type == 'P':
        return 50.0
    if part_type == 'M':
        if year_bucket <= 0: return 0.0   # Vehicle Age <= 6 months (Year 0)
        elif year_bucket == 1: return 5.0   # > 6 months to 1 year (Year 1)
        elif year_bucket == 2: return 10.0  # > 1 year to 2 years (Year 2)
        elif year_bucket == 3: return 15.0  # > 2 years to 3 years (Year 3)
        elif year_bucket == 4: return 25.0  # > 3 years to 4 years (Year 4)
        elif year_bucket == 5: return 35.0  # > 4 years to 5 years (Year 5)
        elif 6 <= year_bucket <= 10: return 40.0 # > 5 years to 10 years (Year 6-10)
        elif year_bucket > 10: return 50.0  # > 10 years (Year 11+)
        else: return 0.0 # Default for unexpected year
    return 0.0 # Default for unknown type

def _process_invoice_worker(task_id, pdf_part, user_data_snapshot, user_id):
    try:
        user_obj = User(user_data_snapshot)
        api_key = user_obj.gemini_api_key or os.getenv('GEMINI_API_KEY')
        user_model = user_obj.gemini_model
        
        from modules.gemini import execute_gemini_task
        invoice_parts_data = execute_gemini_task(
            api_key=api_key,
            pdf_part=pdf_part,
            user_model=user_model,
            is_invoice=True
        )
        _complete_task(task_id, invoice_parts_data)
    except Exception as e:
        print(f"Error processing invoice: {e}")
        import traceback
        traceback.print_exc()
        _fail_task(task_id, f"An unexpected error occurred during invoice processing: {e}")

@app.route('/process_invoice', methods=['POST'])
@login_required
def process_invoice():
    pdf_content = None
    pdf_part = None
    user_id = current_user.id
    user_data_snapshot = {
        'id': current_user.id,
        'username': current_user.username,
        'password_hash': current_user.password_hash,
        'full_name': current_user.full_name,
        'qualifications': current_user.qualifications,
        'designation': current_user.designation,
        'license_no': current_user.license_no,
        'expiry_date': current_user.expiry_date,
        'membership_no': current_user.membership_no,
        'address_line_1': current_user.address_line_1,
        'address_line_2': current_user.address_line_2,
        'address_line_3': current_user.address_line_3,
        'contact_no': current_user.contact_no,
        'email': current_user.email,
        'gemini_api_key': current_user.gemini_api_key,
        'gemini_model': current_user.gemini_model,
    }

    if request.content_type == 'application/json':
        data = request.get_json() or {}
        drive_file_id = data.get('drive_file_id')
        if not drive_file_id:
             return jsonify({"error": "No drive_file_id provided"}), 400
        
        pdf_content = sheets_db.get_file_content(drive_file_id)
        if not pdf_content:
             return jsonify({"error": "Failed to retrieve file content from Drive."}), 500
        pdf_part = {"mime_type": "application/pdf", "data": pdf_content}
             
    elif 'invoice_pdf_file' in request.files:
        file = request.files['invoice_pdf_file']
        if file.filename == '':
            return jsonify({"error": "No selected invoice file"}), 400
        if file and (file.mimetype == 'application/pdf' or file.filename.lower().endswith('.pdf')):
            pdf_content = file.read()
            pdf_part = {"mime_type": "application/pdf", "data": pdf_content}
        else:
             return jsonify({"error": "Invalid file type. Please upload a PDF for the invoice."}), 400
    else:
         return jsonify({"error": "No invoice file provided"}), 400

    task_id = _create_task()
    _task_executor.submit(_process_invoice_worker, task_id, pdf_part, user_data_snapshot, user_id)
    return jsonify({"task_id": task_id}), 202
    
def number_to_words_indian(number_val):
    """
    Converts a number to Indian English words (basic version: lakhs, thousands, hundreds).
    Example: 12345.67 -> "RUPEES TWELVE THOUSAND THREE HUNDRED FORTY FIVE AND PAISE SIXTY SEVEN ONLY"
    Handles up to 99,99,99,999.99
    """
    num = round(float(number_val), 2)
    if num == 0:
        return "RUPEES ZERO ONLY"
    if num > 999999999.99:
        return "RUPEES [Amount too large for words]"

    words = []
    units = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE", "TEN",
             "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]

    def convert_less_than_thousand(n_int):
        s = []
        if n_int >= 100:
            s.append(units[n_int // 100] + " HUNDRED")
            n_int %= 100
        if n_int >= 20:
            s.append(tens[n_int // 10])
            n_int %= 10
        if n_int > 0:
            s.append(units[n_int])
        return " ".join(filter(None, s))

    integer_part = int(num)
    decimal_part = int(round((num - integer_part) * 100))

    if integer_part == 0:
        words.append("ZERO")
    else:
        crores = integer_part // 10000000
        lakhs = (integer_part % 10000000) // 100000
        thousands = (integer_part % 100000) // 1000
        remainder = integer_part % 1000

        if crores > 0:
            words.append(convert_less_than_thousand(crores) + " CRORE")
        if lakhs > 0:
            words.append(convert_less_than_thousand(lakhs) + " LAKH")
        if thousands > 0:
            words.append(convert_less_than_thousand(thousands) + " THOUSAND")
        if remainder > 0:
            words.append(convert_less_than_thousand(remainder))

    rupees_words = " ".join(filter(None, words)).strip()
    result = f"RUPEES {rupees_words}"

    if decimal_part > 0:
        paise_words = convert_less_than_thousand(decimal_part)
        result += f" AND PAISE {paise_words}"

    result += " ONLY"
    return result.upper()

def _calculate_report_assessment_summary(assessment_data_raw, survey_data_raw):
    """
    Helper to recalculate key assessment figures from stored JSON data.
    Returns a dict with:
        page3_gross_total, page3_cgst, page3_sgst, page3_igst, assessed_amount (Net Liability)
    """
    summary = {
        'page3_gross_total': 0.0,
        'page3_cgst': 0.0,
        'page3_sgst': 0.0,
        'page3_igst': 0.0,
        'assessed_amount': 0.0,
        'estimated_amount': 0.0, 
        'customer_gstin': '' 
    }

    try:
        page3_details_raw = assessment_data_raw.get('page3_details', {})
        summary['customer_gstin'] = page3_details_raw.get('customer_gstin', '')
        summary['estimated_amount'] = float(str(page3_details_raw.get('estimated_amount', '0')).replace(',', '')) if page3_details_raw.get('estimated_amount') else 0.0

        p3_fee_items_raw = page3_details_raw.get('fee_items', [])
        p3_photo_copies_str = str(page3_details_raw.get('photo_copies_count', '0')).strip()
        labour_tax_type_main = assessment_data_raw.get('labour_tax_type', 'CGST/SGST') 

        p3_photo_copies_count = int(p3_photo_copies_str) if p3_photo_copies_str.isdigit() else 0
        p3_photo_total_charge = p3_photo_copies_count * 10.0
        
        p3_fees_subtotal = 0.0
        for item_raw in p3_fee_items_raw:
            amount_str = str(item_raw.get('amount', '0')).replace(',', '')
            try:
                amount = float(amount_str)
                if amount != 0.0:
                     p3_fees_subtotal += amount
            except (ValueError, TypeError):
                pass
        
        p3_total_before_gst = p3_fees_subtotal + p3_photo_total_charge
        p3_cgst_calc = 0.0; p3_sgst_calc = 0.0; p3_igst_calc = 0.0
        
        p3_apply_gst = page3_details_raw.get('apply_gst', True)
        if p3_apply_gst:
            if labour_tax_type_main == 'IGST': 
                p3_igst_calc = p3_total_before_gst * 0.18
            else: # CGST/SGST (survey fee GST is always 18%, even when labour tax is Zero)
                p3_cgst_calc = p3_total_before_gst * 0.09
                p3_sgst_calc = p3_total_before_gst * 0.09
        
        summary['page3_cgst'] = p3_cgst_calc
        summary['page3_sgst'] = p3_sgst_calc
        summary['page3_igst'] = p3_igst_calc
        summary['page3_gross_total'] = p3_total_before_gst + p3_cgst_calc + p3_sgst_calc + p3_igst_calc

        labour_paint_depn_input = float(assessment_data_raw.get('labour_paint_depn', 0.0))
        policy_type = assessment_data_raw.get('policy_type', 'NORMAL')
        header_vehicle_year = assessment_data_raw.get('header_vehicle_year', '')
        user_labour_rows = assessment_data_raw.get('user_labour_rows', [])
        
        # New: Labour IMT flag
        labour_imt_applied = assessment_data_raw.get('labour_imt_applied', False)

        # Recalculate Labour Totals
        total_removing = 0.0
        total_denting = 0.0
        total_painting = 0.0

        for row in user_labour_rows:
            def safe_float(val_str, default=0.0):
                try: return float(str(val_str).replace(',', ''))
                except (ValueError, TypeError): return default
            
            removing = safe_float(row.get('removing_refitting')); 
            denting = safe_float(row.get('denting_repairing')); 
            painting = safe_float(row.get('painting'))
            
            total_removing += removing
            total_denting += denting
            total_painting += painting

        # Paint Logic
        final_paint_depn_to_use = labour_paint_depn_input
        if policy_type in ('NIL_DEPN', 'NIL_DEPN_PLUS'):
            final_paint_depn_to_use = 0.0
        
        net_paint_after_dep = total_painting - final_paint_depn_to_use
        
        # IMT 23 Logic (50% on Net Paint)
        labour_imt_deduction = 0.0
        if labour_imt_applied and net_paint_after_dep > 0:
            labour_imt_deduction = net_paint_after_dep * 0.5
        
        net_paint_liability = net_paint_after_dep - labour_imt_deduction
        
        # Taxable Labour
        taxable_labour = total_removing + total_denting + net_paint_liability
        
        # GST Calculation
        labour_gst_amount = 0.0
        if labour_tax_type_main != 'Zero':
            labour_gst_amount = taxable_labour * 0.18
        
        labour_grand_total_final = taxable_labour + labour_gst_amount

        updated_parts = assessment_data_raw.get('parts', [])
        parts_net_amt_final_calc = 0.0
        for part in updated_parts:
            qty = float(part.get('qty', 1.0))
            part_amt = float(part.get('part_amt', 0.0))
            part_type = str(part.get('type_part', '')).strip()
            gst_applicable = part.get('gst_applicable', False)
            original_gst_pc = float(part.get('original_gst_pc', 0.0))
            imt_applied = part.get('imt_applied', False)
            
            # Base Assessed
            total_parts_amt = qty * part_amt
            
            # Depreciation
            depr_rate = get_backend_depreciation_rate(part_type, header_vehicle_year)
            
            # Handle user override for depr
            saved_depr_val_str = str(part.get('depr', '-1.0')).strip()
            try: saved_depr_val = float(saved_depr_val_str)
            except ValueError: saved_depr_val = -1.0
            
            final_depr_amount = 0.0
            if policy_type in ('NIL_DEPN', 'NIL_DEPN_PLUS'):
                 final_depr_amount = 0.0
            elif saved_depr_val >= 0:
                 final_depr_amount = saved_depr_val
            else:
                 final_depr_amount = total_parts_amt * (depr_rate / 100.0) if total_parts_amt > 0 else 0.0
            
            # Net Base (Assessed - Dep)
            net_base = total_parts_amt - final_depr_amount
            
            # GST on Net Base
            total_gst = net_base * (original_gst_pc / 100.0) if gst_applicable else 0.0
            
            # Gross Post-Dep (Net Base + GST)
            gross_post_dep = net_base + total_gst
            
            # IMT 23 on Gross Post-Dep
            imt_23_amt = 0.0
            if imt_applied:
                 imt_23_amt = gross_post_dep * 0.5

            # Final Net Amount
            net_amt = gross_post_dep - imt_23_amt
            parts_net_amt_final_calc += net_amt
            
        excess_final_calc = float(assessment_data_raw.get('deductibles', 1000.0))
        impose_excess_calc = float(assessment_data_raw.get('impose_excess', 0.0)) # New Field
        salvage_raw_calc = assessment_data_raw.get('salvage', '0')
        salvage_val_numeric_calc = 0.0
        try:
            salvage_val_numeric_calc = float(str(salvage_raw_calc).replace(',', ''))
        except (ValueError, TypeError):
            pass 
        
        summary['assessed_amount'] = (labour_grand_total_final + parts_net_amt_final_calc) - excess_final_calc - impose_excess_calc - salvage_val_numeric_calc
        
        # Apply ND deduction (only for NIL_DEPN policy)
        nd_deduction_amount_calc = float(assessment_data_raw.get('nd_deduction_amount', 0.0))
        towing_charges_calc = float(assessment_data_raw.get('towing_charges', 0.0))
        if policy_type == 'NIL_DEPN' and nd_deduction_amount_calc > 0:
            summary['assessed_amount'] -= nd_deduction_amount_calc
        if towing_charges_calc > 0:
            summary['assessed_amount'] += towing_charges_calc

    except Exception as e:
        print(f"Error in _calculate_report_assessment_summary: {e}")
        summary = {
            'page3_gross_total': 0.0, 'page3_cgst': 0.0, 'page3_sgst': 0.0, 'page3_igst': 0.0,
            'assessed_amount': 0.0, 'estimated_amount': 0.0, 'customer_gstin': ''
        }
    return summary

# --- Custom PDF Class with Page Numbers ---
class PDFWithPageNumbers(FPDF):
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Set font
        self.set_font('Helvetica', 'I', 8)
        # Page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')

@app.route('/get_user_profile', methods=['GET'])
@login_required
def get_user_profile():
    user = current_user
    return jsonify({
        "full_name": user.full_name,
        "qualifications": user.qualifications,
        "designation": user.designation,
        "license_no": user.license_no,
        "expiry_date": user.expiry_date,
        "membership_no": user.membership_no,
        "address_line_1": user.address_line_1,
        "address_line_2": user.address_line_2,
        "address_line_3": user.address_line_3,
        "contact_no": user.contact_no,
        "email": user.email,
        "gemini_api_key": user.gemini_api_key,
        "gemini_model": user.gemini_model,
        "role": user.role,
        "permissions": user.permissions,
        "must_change_password": user.must_change_password,
        "workspace_admin_id": workspace_admin_id_for(user)
    })


@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json() or {}
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    if len(new_password) < 8:
        return jsonify({'error': 'New password must be at least 8 characters.'}), 400
    if not bcrypt.check_password_hash(current_user.password_hash, current_password):
        return jsonify({'error': 'Current password is incorrect.'}), 403
    password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    if not sheets_db.change_user_password(current_user.id, password_hash):
        return jsonify({'error': 'Failed to change password.'}), 500
    current_user.password_hash = password_hash
    current_user.must_change_password = False
    return jsonify({'success': True})


@app.route('/api/admin/users', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_users():
    admin_id = workspace_admin_id_for(current_user)
    if request.method == 'GET':
        users = sheets_db.list_admin_users(admin_id)
        safe_users = []
        for user in users:
            safe_users.append({
                'id': user.get('id'), 'username': user.get('username'),
                'full_name': user.get('full_name'), 'email': user.get('email'),
                'is_locked': bool(user.get('is_locked')), 'permissions': user.get('permissions') or {},
                'must_change_password': bool(user.get('must_change_password')),
            })
        return jsonify(safe_users)

    data = request.get_json() or {}
    username = str(data.get('username', '')).strip()
    password = data.get('temporary_password', '')
    if not username or len(password) < 8:
        return jsonify({'error': 'Username and an 8-character temporary password are required.'}), 400
    if sheets_db.get_user_by_username(username):
        return jsonify({'error': 'That username already exists.'}), 409
    requested_permissions = data.get('permissions') or {}
    permissions = {'gmail_sync': bool(requested_permissions.get('gmail_sync', False))}
    user_id = sheets_db.create_user({
        'username': username,
        'password_hash': bcrypt.generate_password_hash(password).decode('utf-8'),
        'full_name': str(data.get('full_name', '')).strip(),
        'email': str(data.get('email', '')).strip(),
        'qualifications': '', 'designation': 'Surveyor & Loss Assessor', 'license_no': '',
        'expiry_date': '', 'membership_no': '', 'address_line_1': '', 'address_line_2': '',
        'address_line_3': '', 'contact_no': '', 'role': 'employee', 'admin_id': admin_id,
        'permissions': permissions, 'must_change_password': True,
    })
    if not user_id:
        return jsonify({'error': 'Failed to create employee.'}), 500
    return jsonify({'success': True, 'id': user_id}), 201


@app.route('/api/admin/users/<int:user_id>/lock', methods=['POST'])
@login_required
@admin_required
def admin_lock_user(user_id):
    data = request.get_json(silent=True) or {}
    managed = sheets_db.get_admin_user(workspace_admin_id_for(current_user), user_id)
    if not managed:
        return jsonify({'error': 'Employee not found.'}), 404
    locked = bool(data['is_locked']) if 'is_locked' in data else not bool(managed.get('is_locked'))
    if not sheets_db.set_user_locked(workspace_admin_id_for(current_user), user_id, locked):
        return jsonify({'error': 'Failed to update employee lock state.'}), 500
    return jsonify({'success': True, 'is_locked': locked})


@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    data = request.get_json() or {}
    temporary_password = data.get('temporary_password', '')
    if len(temporary_password) < 8:
        return jsonify({'error': 'Temporary password must be at least 8 characters.'}), 400
    if not sheets_db.reset_user_password(
        workspace_admin_id_for(current_user), user_id,
        bcrypt.generate_password_hash(temporary_password).decode('utf-8')):
        return jsonify({'error': 'Employee not found or reset failed.'}), 404
    return jsonify({'success': True})


@app.route('/api/admin/backup/download', methods=['GET'])
@login_required
@admin_required
def admin_download_backup():
    """Generate and return a downloadable JSON database snapshot for Admin disaster recovery."""
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403

    try:
        reports_page = sheets_db.get_workspace_reports_page(workspace_admin_id, '', 1, 1000)
        reports = reports_page.get('items', []) if isinstance(reports_page, dict) else []
        fee_bills = sheets_db.get_workspace_fee_bills(workspace_admin_id) or []

        backup_payload = {
            'backup_timestamp': datetime.now().isoformat(),
            'workspace_admin_id': workspace_admin_id,
            'reports_count': len(reports),
            'fee_bills_count': len(fee_bills),
            'reports': reports,
            'fee_bills': fee_bills
        }

        backup_json = json.dumps(backup_payload, indent=2, default=str)
        filename = f"Database_Backup_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"

        return send_file(
            io.BytesIO(backup_json.encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error generating admin backup: {e}")
        return jsonify({'error': f'Failed to generate database backup: {e}'}), 500



@app.route('/api/admin/users/<int:user_id>/permissions', methods=['POST'])
@login_required
@admin_required
def admin_update_permissions(user_id):
    data = request.get_json() or {}
    supplied = data.get('permissions') or {}
    permissions = {'gmail_sync': bool(supplied.get('gmail_sync', False))}
    if not sheets_db.update_user_permissions(workspace_admin_id_for(current_user), user_id, permissions):
        return jsonify({'error': 'Employee not found or permission update failed.'}), 404
    return jsonify({'success': True, 'permissions': permissions})


@app.route('/api/admin/gmail-domains', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_gmail_domains():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if request.method == 'GET':
        return jsonify(sheets_db.get_gmail_sender_domains(workspace_admin_id))
    data = request.get_json() or {}
    domain = str(data.get('domain', '')).lower().strip().lstrip('@')
    if not re.fullmatch(r'[a-z0-9][a-z0-9.-]*\.[a-z]{2,}', domain):
        return jsonify({'error': 'Enter a valid sender domain.'}), 400
    created = sheets_db.add_gmail_sender_domain(workspace_admin_id, domain)
    if not created:
        return jsonify({'error': 'Failed to save sender domain.'}), 500
    return jsonify(created), 201


@app.route('/api/admin/gmail-domains/<int:domain_id>', methods=['DELETE'])
@login_required
@admin_required
def admin_delete_gmail_domain(domain_id):
    if not sheets_db.delete_gmail_sender_domain(workspace_admin_id_for(current_user), domain_id):
        return jsonify({'error': 'Sender domain not found.'}), 404
    return jsonify({'success': True})

@app.route('/update_user_profile', methods=['POST'])
@login_required
def update_user_profile():
    data = request.get_json()
    try:
        success = sheets_db.update_user(current_user.id, data)
        if not success:
            return jsonify({"error": "Failed to update profile in database."}), 500
        
        # update current session user object
        current_user.full_name = data.get('full_name', current_user.full_name)
        current_user.qualifications = data.get('qualifications', current_user.qualifications)
        current_user.designation = data.get('designation', current_user.designation)
        current_user.license_no = data.get('license_no', current_user.license_no)
        current_user.expiry_date = data.get('expiry_date', current_user.expiry_date)
        current_user.membership_no = data.get('membership_no', current_user.membership_no)
        current_user.address_line_1 = data.get('address_line_1', current_user.address_line_1)
        current_user.address_line_2 = data.get('address_line_2', current_user.address_line_2)
        current_user.address_line_3 = data.get('address_line_3', current_user.address_line_3)
        current_user.contact_no = data.get('contact_no', current_user.contact_no)
        current_user.email = data.get('email', current_user.email)
        current_user.gemini_api_key = data.get('gemini_api_key', current_user.gemini_api_key)
        current_user.gemini_model = data.get('gemini_model', current_user.gemini_model)
        
        return jsonify({"success": True, "message": "Profile updated successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/upload_signature', methods=['POST'])
@login_required
def upload_signature():
    if 'signature' not in request.files:
        return jsonify({'error': 'No signature file provided'}), 400
    file = request.files['signature']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    project_root = os.path.dirname(os.path.abspath(__file__))
    uploads_dir = os.path.join(project_root, 'uploads')
    static_dir = os.path.join(project_root, 'static')
    os.makedirs(uploads_dir, exist_ok=True)
    os.makedirs(static_dir, exist_ok=True)

    user_id_str = str(current_user.id)
    sig_user_filename = f'signature_{user_id_str}.png'
    sig_path = os.path.join(uploads_dir, sig_user_filename)
    file.save(sig_path)
    
    # Also save as signature.png for fallback
    global_sig_path = os.path.join(uploads_dir, 'signature.png')
    
    # Mirror to static directory for direct URL serving
    static_sig_path = os.path.join(static_dir, sig_user_filename)
    global_static_sig_path = os.path.join(static_dir, 'signature.png')
    
    try:
        import shutil
        shutil.copy2(sig_path, static_sig_path)
        shutil.copy2(sig_path, global_sig_path)
        shutil.copy2(sig_path, global_static_sig_path)
    except Exception as copy_err:
        print(f"Signature copy error: {copy_err}")
    
    ts = int(_time.time())
    return jsonify({'success': True, 'message': 'Digital seal & signature uploaded successfully!', 'url': f'/static/{sig_user_filename}?v={ts}'})

@app.route('/signature_status', methods=['GET'])
@login_required
def get_signature_status():
    project_root = os.path.dirname(os.path.abspath(__file__))
    user_id_str = str(current_user.id)
    sig_user_filename = f'signature_{user_id_str}.png'
    
    per_user_uploads = os.path.join(project_root, 'uploads', sig_user_filename)
    per_user_static = os.path.join(project_root, 'static', sig_user_filename)
    global_uploads = os.path.join(project_root, 'uploads', 'signature.png')
    global_static = os.path.join(project_root, 'static', 'signature.png')
    
    if os.path.exists(per_user_uploads) or os.path.exists(per_user_static):
        sig_url = f'/static/{sig_user_filename}'
        sig_exists = True
    elif os.path.exists(global_uploads) or os.path.exists(global_static):
        sig_url = '/static/signature.png'
        sig_exists = True
    else:
        sig_url = None
        sig_exists = False
        
    ts = int(_time.time())
    return jsonify({'has_signature': sig_exists, 'url': f'{sig_url}?v={ts}' if sig_exists else None})

@app.route('/api/available_models', methods=['GET'])
@login_required
def available_models():
    # Allow passing custom API key as query parameter for validation/testing in real-time
    custom_key = request.args.get('api_key')
    key_to_use = custom_key if custom_key is not None else (current_user.gemini_api_key or os.getenv("GEMINI_API_KEY"))
    models = get_user_best_models(key_to_use)
    return jsonify(models)

# --- Photo Upload Route ---
@app.route('/upload_photo', methods=['POST'])
@login_required
def upload_photo():
    if 'photo' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file:
        filename = secure_filename(f"{uuid.uuid4()}_{file.filename}")
        # Read file content
        content = file.read()
        
        # 1. Try uploading to Drive
        try:
            result = sheets_db.upload_image_to_drive(content, filename, file.mimetype)
            if result and result.get('id'):
                # Return a proxy URL that will serve the image through the backend
                proxy_url = f"/proxy_image/{result.get('id')}"
                return jsonify({'success': True, 'url': proxy_url})
        except Exception as e:
            print(f"Drive upload exception: {e}")
            
        # 2. Local Fallback if Drive upload fails or returns None (quota exceeded)
        print("Drive upload failed or quota exceeded. Falling back to local VPS storage...")
        try:
            upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            local_path = os.path.join(upload_dir, filename)
            with open(local_path, 'wb') as f:
                f.write(content)
            
            proxy_url = f"/local_image/{filename}"
            return jsonify({'success': True, 'url': proxy_url, 'warning': 'Saved to local storage (Drive full)'})
        except Exception as local_err:
            print(f"Local storage fallback failed: {local_err}")
            return jsonify({'error': 'Failed to upload to Drive or local storage'}), 500

# --- Photo Proxy Route ---
@app.route('/proxy_image/<file_id>')
@login_required
def proxy_image(file_id):
    """Serves images from Google Drive through the backend proxy."""
    # SECURITY: Validate file_id format to prevent injection (VULN-06)
    import re as re_module
    if not re_module.match(r'^[a-zA-Z0-9_-]+$', file_id):
        abort(400)
    
    content = sheets_db.get_file_content(file_id)
    if content:
        # Detect image type from content (basic check for common types)
        mime_type = 'image/jpeg'  # default
        if content[:8] == b'\x89PNG\r\n\x1a\n':
            mime_type = 'image/png'
        elif content[:4] == b'GIF8':
            mime_type = 'image/gif'
        elif content[:4] == b'RIFF' and content[8:12] == b'WEBP':
            mime_type = 'image/webp'
        
        return send_file(io.BytesIO(content), mimetype=mime_type)
    else:
        abort(404)

# --- Local Photo Serve Route ---
@app.route('/local_image/<filename>')
@login_required
def serve_local_image(filename):
    """Serves locally stored backup images."""
    # Validate filename to prevent directory traversal
    import re as re_module
    if not re_module.match(r'^[a-zA-Z0-9_-]+\.[a-zA-Z0-9]+$', filename):
        abort(400)
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
    from flask import send_from_directory
    return send_from_directory(upload_dir, filename)

# --- File Generation Route ---
@app.route('/generate_files', methods=['POST'])
@login_required
def generate_files():
    """Accept report data and start async PDF generation. Returns a task_id for polling."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    # Snapshot request-context values needed by the worker thread
    user_full_name = current_user.full_name
    user_id = current_user.id
    from flask import session
    access_token = session.get('google_access_token')

    task_id = _create_task()
    _task_executor.submit(_generate_files_worker, task_id, data, user_full_name, user_id, access_token)
    return jsonify({"task_id": task_id}), 202


@app.route('/generate_files/status/<task_id>', methods=['GET'])
@login_required
def generate_files_status(task_id):
    """Poll for the status of an async file generation task."""
    task = _get_task(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    
    if task["status"] == "completed":
        return jsonify({"status": "completed", "result": task["result"]})
    elif task["status"] == "error":
        return jsonify({"status": "error", "error": task["error"]})
    else:
        return jsonify({"status": "processing"})


def _generate_files_worker(task_id, data, user_full_name, user_id, access_token):
    """Background worker for PDF generation. Runs outside request context."""
    try:
        # Load user object from database for PDF header (needs full profile)
        _user_data = sheets_db.get_user_by_id(user_id)
        _user_obj = User(_user_data) if _user_data else None

        # --- Data Extraction ---
        survey_data = data.get('survey_report', {})
        assessment_data = data.get('assessment', {})
        
        # Extract Photo Data
        photos_data = data.get('photos', {})

        page3_details_raw = assessment_data.get('page3_details', {})
        p3_customer_gstin_raw = page3_details_raw.get('customer_gstin', '')
        p3_company_gstin_raw = page3_details_raw.get('company_gstin', '') 
        p3_fee_items_raw = page3_details_raw.get('fee_items', [])
        p3_estimated_amount_str = str(page3_details_raw.get('estimated_amount', '0')).strip()
        
        # Surveyor Bank Details
        surveyor_details = page3_details_raw.get('surveyor_details', {})
        p3_apply_gst = page3_details_raw.get('apply_gst', True)

        p3_photo_copies_str = str(page3_details_raw.get('photo_copies_count', '0')).strip()
        p3_photo_copies_count = 0 
        if p3_photo_copies_str: 
            try:
                float_val = float(p3_photo_copies_str)
                if float_val.is_integer(): 
                    count_val = int(float_val)
                    if count_val >= 0: 
                        p3_photo_copies_count = count_val
            except ValueError:
                pass

        user_labour_rows = assessment_data.get('user_labour_rows', [])
        updated_parts = assessment_data.get('parts', [])
        header_gst_raw = assessment_data.get('header_gst', '') 
        header_vehicle_year_raw = assessment_data.get('header_vehicle_year', '')
        policy_type = assessment_data.get('policy_type', 'NORMAL') 
        report_type_raw = assessment_data.get('report_type', 'Final Survey Report')
        claim_type_raw = assessment_data.get('claim_type', 'Cashless')
        labour_tax_type_main = assessment_data.get('labour_tax_type', 'CGST/SGST')
        labour_imt_applied = assessment_data.get('labour_imt_applied', False) 
        
        reinspection_note_raw = assessment_data.get('reinspection_note', "As per instruction received from your office I had visited the insured repairer garage. I had taken some photographs of the insured's repaired vehicle and checked. The undersigned have been satisfied with the repairing job of insured vehicle.")
        enclosures_text_raw = assessment_data.get('enclosures_text', '')
        parts_table_note_raw = assessment_data.get('parts_table_note', '')
        spot_report_text_raw = assessment_data.get('spot_report_text', '') 
        spot_report_enclosures_raw = assessment_data.get('spot_report_enclosures', '') 

        labour_paint_depn_input = float(assessment_data.get('labour_paint_depn', 0.0))
        note_text_final_raw = assessment_data.get('note_text', "Note :- The subject policy covered with Depn. waiver")
        payment_to_text_final_raw = assessment_data.get('payment_to_text', "REPAIRER")
        salvage_raw_data = assessment_data.get('salvage', '0') 

        # Estimate Overrides
        est_labour_override = assessment_data.get('est_labour_override', '')

        # ND Deduction and Towing Charges
        nd_deduction_pc = float(assessment_data.get('nd_deduction_pc', 5))
        nd_deduction_amount = float(assessment_data.get('nd_deduction_amount', 0.0))
        towing_charges = float(assessment_data.get('towing_charges', 0.0))
        est_paint_override = assessment_data.get('est_paint_override', '')
        est_parts_override = assessment_data.get('est_parts_override', '')

        final_survey_data = {key: survey_data.get(key, '') for key in EXPECTED_FIELDS}
        
        def get_survey_val(key):
            raw_value = final_survey_data.get(key, '')
            return normalize_pdf_text_for_fpdf(raw_value)

        header_vehicle_year = normalize_pdf_text_for_fpdf(header_vehicle_year_raw)
        note_text_final = normalize_pdf_text_for_fpdf(note_text_final_raw)
        payment_to_text_final = normalize_pdf_text_for_fpdf(payment_to_text_final_raw)
        reinspection_note = normalize_pdf_text_for_fpdf(reinspection_note_raw)
        p3_customer_gstin = normalize_pdf_text_for_fpdf(p3_customer_gstin_raw)
        p3_company_gstin = normalize_pdf_text_for_fpdf(p3_company_gstin_raw)
        salvage_raw = normalize_pdf_text_for_fpdf(salvage_raw_data)
        report_type = normalize_pdf_text_for_fpdf(report_type_raw)
        claim_type = normalize_pdf_text_for_fpdf(claim_type_raw)
        # Fix: Normalize potentially long/special char texts
        enclosures_text = normalize_pdf_text_for_fpdf(enclosures_text_raw)
        parts_table_note = normalize_pdf_text_for_fpdf(parts_table_note_raw)
        spot_report_text = normalize_pdf_text_for_fpdf(spot_report_text_raw)
        spot_report_enclosures = normalize_pdf_text_for_fpdf(spot_report_enclosures_raw)

        try:
            gst_num_raw = str(header_gst_raw).replace('%','').strip()
            gst_num = float(gst_num_raw)
            header_gst_display = normalize_pdf_text_for_fpdf(f"{gst_num:.2f}%")
        except (ValueError, TypeError):
            header_gst_display = normalize_pdf_text_for_fpdf(header_gst_raw)

        # --- Labour Recalculation (Page 2) ---
        labour_sum_removing = 0.0
        labour_sum_denting = 0.0
        labour_sum_painting = 0.0 
        processed_user_labour_rows_for_pdf = []
        
        for row_data in user_labour_rows:
            def safe_float_from_string(val_str, default=0.0):
                try: return float(str(val_str).replace(',', ''))
                except (ValueError, TypeError): return default
            
            removing = safe_float_from_string(row_data.get('removing_refitting', 0))
            denting = safe_float_from_string(row_data.get('denting_repairing', 0))
            painting = safe_float_from_string(row_data.get('painting', 0))
            
            labour_sum_removing += removing
            labour_sum_denting += denting
            labour_sum_painting += painting 
            
            processed_user_labour_rows_for_pdf.append({
                'part_name': normalize_pdf_text_for_fpdf(row_data.get('part_name', '')),
                'removing_refitting': removing,
                'denting_repairing': denting,
                'painting': painting
            })
        
        # Logic: 
        # 1. Paint Dep (12.5% default)
        labour_paint_depn_final = labour_paint_depn_input 
        if policy_type in ('NIL_DEPN', 'NIL_DEPN_PLUS'):
            labour_paint_depn_final = 0.0
        
        net_paint_after_dep = labour_sum_painting - labour_paint_depn_final
        
        # 2. IMT 23 (50% on Net Paint)
        labour_imt_deduction = 0.0
        if labour_imt_applied and net_paint_after_dep > 0:
            labour_imt_deduction = net_paint_after_dep * 0.5
        
        net_paint_liability = net_paint_after_dep - labour_imt_deduction
        
        # 3. Taxable Labour
        labour_rr_dent_sum = labour_sum_removing + labour_sum_denting
        taxable_labour = labour_rr_dent_sum + net_paint_liability
        
        # 4. GST
        labour_gst_amount = 0.0
        if labour_tax_type_main != 'Zero':
            labour_gst_amount = taxable_labour * 0.18
        
        # 5. Final
        labour_grand_total_final = taxable_labour + labour_gst_amount
        
        # --- Parts Recalculation (Page 3) & Material Liability ---
        parts_total_base_final = 0.0; parts_total_gst_final = 0.0; parts_grand_total_final = 0.0; parts_net_amt_final = 0.0; parts_depr_sum_final = 0.0
        parts_total_estimate = 0.0; parts_total_bill = 0.0; parts_total_assessed = 0.0; parts_total_imt23 = 0.0
        
        liability_metal = 0.0; liability_glass = 0.0; liability_plastic = 0.0

        final_parts_calculated = []
        for part in updated_parts:
            try:
                qty = float(part.get('qty', 1.0)); part_amt = float(part.get('part_amt', 0.0)); part_type = str(part.get('type_part', '')).strip().upper()
                gst_applicable = part.get('gst_applicable', False); original_gst_pc = float(part.get('original_gst_pc', 0.0))
                imt_applied = part.get('imt_applied', False) 
                
                estimate_amt = float(part.get('estimate_amt', 0.0))
                bill_amt = float(part.get('bill_amt', 0.0))
                
                depr_amount_from_frontend_str = str(part.get('depr', '-1.0')).strip()
                depr_amount_from_frontend = -1.0
                try: depr_amount_from_frontend = float(depr_amount_from_frontend_str)
                except ValueError: pass 

                # Base Assessed
                total_parts_amt = qty * part_amt 
                
                # Depreciation
                final_depr_amount_to_use = 0.0
                if policy_type in ('NIL_DEPN', 'NIL_DEPN_PLUS'): final_depr_amount_to_use = 0.0
                elif depr_amount_from_frontend >= 0: final_depr_amount_to_use = depr_amount_from_frontend
                else: 
                    calculated_depr_rate = get_backend_depreciation_rate(part_type, header_vehicle_year)
                    final_depr_amount_to_use = total_parts_amt * (calculated_depr_rate / 100.0) if total_parts_amt > 0 else 0.0
                
                # Net Base (Assessed - Dep)
                net_base = total_parts_amt - final_depr_amount_to_use
                
                # GST on Net Base
                total_gst = net_base * (original_gst_pc / 100.0) if gst_applicable else 0.0
                
                # Gross Post-Dep (Net Base + GST)
                gross_post_dep = net_base + total_gst

                # IMT 23 on Gross Post-Dep
                imt_23_amt = 0.0
                if imt_applied:
                    imt_23_amt = gross_post_dep * 0.5
                
                # Final Net Amount
                net_amt = gross_post_dep - imt_23_amt

                if part_type == 'M': liability_metal += net_amt
                elif part_type == 'G': liability_glass += net_amt
                elif part_type == 'P': liability_plastic += net_amt

                part_name_display = normalize_pdf_text_for_fpdf(part.get('part_name', ''))
                
                output_part = part.copy() 
                output_part['total_parts_amt'] = total_parts_amt; output_part['total_gst'] = total_gst; output_part['gross_amt'] = gross_post_dep
                output_part['depr'] = final_depr_amount_to_use; output_part['net_amt'] = net_amt; output_part['part_name_display'] = part_name_display
                output_part['estimate_amt'] = estimate_amt; output_part['bill_amt'] = bill_amt; output_part['imt_23_amt'] = imt_23_amt
                output_part['net_base'] = net_base # Store for PDF
                
                output_part['salvage_produce'] = normalize_pdf_text_for_fpdf(part.get('salvage_produce', 'YES'))
                output_part['remarks'] = normalize_pdf_text_for_fpdf(part.get('remarks', 'REPLACED BY NEW'))

                final_parts_calculated.append(output_part)
                
                parts_total_base_final += total_parts_amt; parts_total_gst_final += total_gst; parts_grand_total_final += gross_post_dep; parts_net_amt_final += net_amt
                parts_depr_sum_final += final_depr_amount_to_use
                parts_total_estimate += estimate_amt; parts_total_bill += bill_amt; parts_total_assessed += total_parts_amt; parts_total_imt23 += imt_23_amt

            except (ValueError, TypeError) as e: print(f"Warning: Error processing part {part.get('sl_no')} for output. Skipping totals. Error: {e}")

        excess_final = float(assessment_data.get('deductibles', 1000.0))
        impose_excess_final = float(assessment_data.get('impose_excess', 0.0)) # New Field
        try: salvage_val_numeric = float(str(salvage_raw).replace(',', ''))
        except (ValueError, TypeError): salvage_val_numeric = 0.0
        
        net_liability_final = (labour_grand_total_final + parts_net_amt_final) - excess_final - impose_excess_final - salvage_val_numeric
        
        # Apply ND deduction (only for NIL_DEPN policy)
        if policy_type == 'NIL_DEPN' and nd_deduction_amount > 0:
            net_liability_final -= nd_deduction_amount
        
        # Apply Towing Charges (add)
        if towing_charges > 0:
            net_liability_final += towing_charges
        
        # --- Page 4 (Tax Invoice) Calculations ---
        p3_photo_total_charge = p3_photo_copies_count * 10.0
        p3_fees_subtotal = 0.0
        p3_valid_fee_items = []
        for item_raw in p3_fee_items_raw:
            name_raw = str(item_raw.get('name', '')).strip()
            name = normalize_pdf_text_for_fpdf(name_raw)
            amount_str = str(item_raw.get('amount', '0')).replace(',', '')
            try:
                amount = float(amount_str)
                if amount != 0.0:
                    p3_valid_fee_items.append({'name': name, 'amount': amount})
                    p3_fees_subtotal += amount
            except (ValueError, TypeError): pass
            
        p3_total_before_gst = p3_fees_subtotal + p3_photo_total_charge
        p3_cgst = 0.0; p3_sgst = 0.0; p3_igst = 0.0
        
        if p3_apply_gst:
            if labour_tax_type_main == 'IGST': 
                p3_igst = p3_total_before_gst * 0.18
            else: # CGST/SGST (survey fee GST is always 18%, even when labour tax is Zero)
                p3_cgst = p3_total_before_gst * 0.09; p3_sgst = p3_total_before_gst * 0.09
            
        p3_grand_total = p3_total_before_gst + p3_cgst + p3_sgst + p3_igst
        p3_grand_total_in_words_raw = number_to_words_indian(p3_grand_total)
        p3_grand_total_in_words = normalize_pdf_text_for_fpdf(p3_grand_total_in_words_raw)
        p3_estimated_amount = 0.0
        try: p3_estimated_amount = float(p3_estimated_amount_str.replace(',', ''))
        except ValueError: p3_estimated_amount = 0.0

        # --- PDF Setup ---
        pdf = PDFWithPageNumbers(orientation='P', unit='mm', format='A4')
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(auto=True, margin=15)
        base_font_size_page1 = 10; base_font_size_page2 = 10; base_font_size_page3 = 9 
        line_h_page1 = 5.5; line_h_page2 = 5.5; line_h_page3 = 5    
        table_cell_padding_y = 0.8
        header_image_path = os.path.join(app.static_folder, 'header.png')
        header_image_height = 30; header_bottom_margin = 10

        # --- Helper Functions ---
        def add_pdf_header(pdf_obj):
            if pdf_obj.page_no() == 1:
                # Use user object loaded at start of worker (no request context here)
                u = _user_obj
                
                # --- Left Side ---
                pdf_obj.set_y(10)
                pdf_obj.set_x(10)
                
                # Name (Red)
                pdf_obj.set_text_color(239, 68, 68) # Red
                pdf_obj.set_font('Helvetica', 'B', 16)
                pdf_obj.cell(0, 8, normalize_pdf_text_for_fpdf(u.full_name), 0, 1, 'L')
                
                # Qualifications (Black)
                pdf_obj.set_text_color(0, 0, 0)
                pdf_obj.set_font('Helvetica', '', 9)
                pdf_obj.cell(0, 5, normalize_pdf_text_for_fpdf(u.qualifications), 0, 1, 'L')
                
                # Designation (Red)
                pdf_obj.set_text_color(239, 68, 68)
                pdf_obj.set_font('Helvetica', 'B', 10)
                pdf_obj.cell(0, 5, normalize_pdf_text_for_fpdf(u.designation), 0, 1, 'L')
                
                # License & Expiry & Membership (Black)
                pdf_obj.set_text_color(0, 0, 0)
                pdf_obj.set_font('Helvetica', '', 9)
                pdf_obj.cell(0, 5, normalize_pdf_text_for_fpdf(f"Licence No: {u.license_no}"), 0, 1, 'L')
                pdf_obj.cell(0, 5, normalize_pdf_text_for_fpdf(f"Expiry on: {u.expiry_date}"), 0, 1, 'L')
                pdf_obj.cell(0, 5, normalize_pdf_text_for_fpdf(f"IIISLA Membership No-{u.membership_no}"), 0, 1, 'L')

                # --- Right Side (Address) ---
                # Helper for right alignment
                def right_text(txt, y_offset=0, color=(0,0,0), bold=False):
                    pdf_obj.set_y(10 + y_offset)
                    pdf_obj.set_x(120) # Start from middle-right
                    pdf_obj.set_font('Helvetica', 'B' if bold else '', 9)
                    pdf_obj.set_text_color(*color)
                    pdf_obj.cell(80, 5, normalize_pdf_text_for_fpdf(txt), 0, 0, 'R')

                right_text(u.address_line_1, 0)
                right_text(u.address_line_2, 5)
                right_text(u.address_line_3, 10)

                
                # Cell (Label Black, Value Red) - Dynamic positioning to remove gap
                cell_val = normalize_pdf_text_for_fpdf(u.contact_no)
                cell_lbl = "Cell: "
                pdf_obj.set_font('Helvetica', 'B', 9); val_w = pdf_obj.get_string_width(cell_val)
                pdf_obj.set_font('Helvetica', '', 9); lbl_w = pdf_obj.get_string_width(cell_lbl)
                start_x = 200 - (lbl_w + val_w) # 200 is right margin alignment
                
                pdf_obj.set_y(25)
                pdf_obj.set_x(start_x)
                pdf_obj.set_text_color(0,0,0); pdf_obj.cell(lbl_w, 5, cell_lbl, 0, 0, 'L')
                pdf_obj.set_text_color(239, 68, 68); pdf_obj.set_font('Helvetica', 'B', 9); pdf_obj.cell(val_w, 5, cell_val, 0, 1, 'L')

                # Email (Label Black, Value Red) - Dynamic positioning to remove gap
                email_val = normalize_pdf_text_for_fpdf(u.email)
                email_lbl = "Email: "
                pdf_obj.set_font('Helvetica', 'B', 9); val_w = pdf_obj.get_string_width(email_val)
                pdf_obj.set_font('Helvetica', '', 9); lbl_w = pdf_obj.get_string_width(email_lbl)
                start_x = 200 - (lbl_w + val_w)
                
                pdf_obj.set_y(30)
                pdf_obj.set_x(start_x)
                pdf_obj.set_text_color(0,0,0); pdf_obj.cell(lbl_w, 5, email_lbl, 0, 0, 'L')
                pdf_obj.set_text_color(239, 68, 68); pdf_obj.set_font('Helvetica', 'B', 9); pdf_obj.cell(val_w, 5, email_val, 0, 1, 'L')

                # --- Bottom Line (Blue) ---
                pdf_obj.set_draw_color(59, 130, 246) # Blue
                pdf_obj.set_line_width(0.5)
                pdf_obj.line(10, 42, 200, 42)
                
                # Reset
                pdf_obj.set_y(45)
                pdf_obj.set_text_color(0, 0, 0)
                pdf_obj.set_draw_color(0, 0, 0)
                pdf_obj.set_line_width(0.2)
            else:
                pdf_obj.set_y(10)

            if pdf_obj.page_no() > 1:
                current_y = pdf_obj.get_y()
                pdf_obj.set_font("Helvetica", 'B', 9)
                report_text = f"Report No: {get_survey_val('report_no')}"
                vehicle_text = f"Vehicle No: {get_survey_val('vehicle_regn_no')}"
                pdf_obj.set_xy(pdf_obj.l_margin, current_y)
                pdf_obj.cell(80, 5, normalize_pdf_text_for_fpdf(report_text), 0, 0, 'L')
                right_margin_x = pdf_obj.w - pdf_obj.r_margin
                pdf_obj.set_xy(right_margin_x - 80, current_y)
                pdf_obj.cell(80, 5, normalize_pdf_text_for_fpdf(vehicle_text), 0, 1, 'R')
                pdf_obj.ln(10)

        def draw_table_row(data, widths, height_per_line, border='TBLR', align='C', fill=False, text_color=(0,0,0), fill_color=(255,255,255), alignments=None, font_style='', is_header=False, current_font_size=base_font_size_page2 -1):
            pdf.set_fill_color(*fill_color); pdf.set_text_color(*text_color)
            base_table_font_size = current_font_size; header_font_size = current_font_size 
            max_lines = 1
            for i, item in enumerate(data):
                w = widths[i]
                temp_item_str_original = format_pdf_number(item) if isinstance(item, (int, float)) else str(item)
                temp_item_str = normalize_pdf_text_for_fpdf(temp_item_str_original)
                pdf.set_font("Helvetica", 'B' if is_header else (font_style if font_style else ''), base_table_font_size)
                lines = pdf.multi_cell(w, height_per_line, temp_item_str, border=0, align=alignments[i] if alignments else align, dry_run=True, output="LINES", max_line_height=height_per_line)
                max_lines = max(max_lines, len(lines))
            total_row_height = max_lines * height_per_line + table_cell_padding_y 
            if pdf.get_y() + total_row_height > pdf.h - pdf.b_margin:
                pdf.add_page(orientation=pdf.def_orientation); add_pdf_header(pdf); pdf.set_fill_color(*fill_color); pdf.set_text_color(*text_color)
            x_start = pdf.get_x(); y_start = pdf.get_y()
            if fill: pdf.rect(x_start, y_start, sum(widths), total_row_height, 'F')
            current_x = x_start
            for i, item in enumerate(data):
                w = widths[i]; cell_align = alignments[i] if alignments else align
                if isinstance(item, (int, float)): item_str = format_pdf_number(item)
                else: item_str = normalize_pdf_text_for_fpdf(str(item))
                cell_font_style = font_style if font_style else ('B' if is_header else '')
                cell_font_size = header_font_size if is_header else base_table_font_size
                pdf.set_font("Helvetica", cell_font_style, cell_font_size)
                pdf.set_xy(current_x, y_start + table_cell_padding_y / 2) 
                pdf.multi_cell(w, height_per_line, item_str, border=0, align=cell_align, padding=(0, 1), max_line_height=height_per_line) 
                current_x += w
            line_x = x_start; line_y = y_start; row_width = sum(widths); row_height = total_row_height
            is_full_border = (isinstance(border, int) and border == 1) or (isinstance(border, str) and border.upper() == 'TBLR')
            draw_top = is_full_border or (isinstance(border, str) and 'T' in border.upper()); draw_bottom = is_full_border or (isinstance(border, str) and 'B' in border.upper())
            draw_left = is_full_border or (isinstance(border, str) and 'L' in border.upper()); draw_right = is_full_border or (isinstance(border, str) and 'R' in border.upper())
            pdf.set_draw_color(0,0,0) 
            if draw_top: pdf.line(line_x, line_y, line_x + row_width, line_y)
            if draw_bottom: pdf.line(line_x, line_y + row_height, line_x + row_width, line_y + row_height)
            if draw_left: pdf.line(line_x, line_y, line_x, line_y + row_height)
            if draw_right: pdf.line(line_x + row_width, line_y, line_x + row_width, line_y + row_height)
            if is_full_border: 
                temp_x = x_start
                for i in range(len(widths) - 1): temp_x += widths[i]; pdf.line(temp_x, line_y, temp_x, line_y + row_height)
            pdf.set_y(y_start + total_row_height); pdf.set_x(x_start) 
            return total_row_height

        def calculate_height(texts, widths, font_size, line_height, padding, font_style=None):
            max_lines = 1; temp_pdf = FPDF(); temp_pdf.add_page()
            for i, text in enumerate(texts):
                current_style = font_style[i] if isinstance(font_style, list) and i < len(font_style) else (font_style if font_style else '')
                temp_pdf.set_font("Helvetica", current_style, font_size)
                normalized_text_for_calc = normalize_pdf_text_for_fpdf(str(text) if text is not None else '')
                lines = temp_pdf.multi_cell(widths[i], line_height, normalized_text_for_calc, dry_run=True, output="LINES", max_line_height=line_height)
                max_lines = max(max_lines, len(lines))
            del temp_pdf
            return max_lines * line_height + padding

        def add_plain_pair(label1, val1, label2, val2, label_width, val_width, font_size=base_font_size_page1, line_h=line_h_page1):
            start_y = pdf.get_y()
            
            # Normalize text
            norm_label1 = normalize_pdf_text_for_fpdf(label1)
            norm_val1 = normalize_pdf_text_for_fpdf(str(val1))
            norm_label2 = normalize_pdf_text_for_fpdf(label2) if label2 else ""
            norm_val2 = normalize_pdf_text_for_fpdf(str(val2)) if val2 else ""
            
            # Combine into single markdown strings: **Label:** Value
            # This removes the fixed column gap, allowing value to start immediately after label
            text1 = f"**{norm_label1}** {norm_val1}"
            text2 = f"**{norm_label2}** {norm_val2}" if label2 else ""
            
            # Total width for the column is the sum of the allocated label and value widths
            col_total_width = label_width + val_width
            
            # Calculate height for column 1 using the current PDF context (dry_run)
            # We use the main pdf object to ensure font/markdown calculations are accurate
            pdf.set_font("Helvetica", '', font_size)
            lines1 = pdf.multi_cell(col_total_width, line_h, text1, markdown=True, dry_run=True, output="LINES", max_line_height=line_h)
            h1 = len(lines1) * line_h
            
            # Calculate height for column 2
            h2 = 0
            if label2:
                 lines2 = pdf.multi_cell(col_total_width, line_h, text2, markdown=True, dry_run=True, output="LINES", max_line_height=line_h)
                 h2 = len(lines2) * line_h
            
            max_h = max(h1, h2) + table_cell_padding_y
            
            # Check page break
            if pdf.get_y() + max_h > pdf.page_break_trigger: 
                pdf.add_page(); add_pdf_header(pdf); start_y = pdf.get_y()
            
            # Render Column 1
            pdf.set_xy(pdf.l_margin, start_y)
            pdf.set_font("Helvetica", '', font_size)
            pdf.multi_cell(col_total_width, line_h, text1, markdown=True, border=0, align='L', max_line_height=line_h)
            
            # Render Column 2
            if label2:
                pdf.set_xy(pdf.l_margin + col_total_width, start_y)
                pdf.set_font("Helvetica", '', font_size)
                pdf.multi_cell(col_total_width, line_h, text2, markdown=True, border=0, align='L', max_line_height=line_h)
            
            pdf.set_y(start_y + max_h)
            return pdf.get_y()

        def add_section_header(text):
            # Check if enough space remains (e.g., 30mm) to avoid orphaned headers
            if pdf.get_y() > pdf.h - pdf.b_margin - 30:
                pdf.add_page(orientation=pdf.def_orientation)
                add_pdf_header(pdf)
            pdf.ln(2); pdf.set_font("Helvetica", 'B', base_font_size_page1)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.cell(0, line_h_page1 * 1.5, normalize_pdf_text_for_fpdf(text), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y()); pdf.ln(1); pdf.set_font("Helvetica", '', base_font_size_page1)

        def add_photo_section(pdf_obj, title, vehicle_no, photos_list, photos_per_page):
            if not photos_list:
                return
            margin = 10
            # Reduced page height calculation to shift photos upward and prevent footer overlap
            # Changed subtraction from -20 to -35 to reserve more space at the bottom
            page_width = 210 - (2 * margin)
            page_height = 297 - (2 * margin) - 35 
            
            cols = 2
            try: photos_per_page = int(photos_per_page)
            except: photos_per_page = 4
            if photos_per_page == 4: rows = 2
            elif photos_per_page == 6: rows = 3
            elif photos_per_page == 8: rows = 4
            else: rows = 2 
            
            img_width = (page_width - 5) / cols
            img_height = (page_height - 5) / rows 
            
            start_y = 0
            for i, photo_b64 in enumerate(photos_list):
                if i % photos_per_page == 0:
                    pdf_obj.add_page(orientation='P'); add_pdf_header(pdf_obj)
                    pdf_obj.set_font("Helvetica", 'B', 12)
                    pdf_obj.cell(0, 10, normalize_pdf_text_for_fpdf(f"{title} ({vehicle_no})"), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
                    # Reduced line break after title to pull photos up slightly
                    pdf_obj.ln(1) 
                    start_y = pdf_obj.get_y()
                
                pos_in_page = i % photos_per_page; row = pos_in_page // cols; col = pos_in_page % cols
                x = pdf_obj.l_margin + (col * (img_width + 5)); y = start_y + (row * (img_height + 5))
                x = pdf_obj.l_margin + (col * (img_width + 5)); y = start_y + (row * (img_height + 5))
                try:
                    img_stream = None
                    if photo_b64.startswith('/proxy_image/'):
                        # It is a proxy URL - extract file ID and fetch directly
                        file_id = photo_b64.replace('/proxy_image/', '')
                        img_data = sheets_db.get_file_content(file_id)
                        if img_data:
                            img_stream = io.BytesIO(img_data)
                        else:
                            print(f"Failed to fetch image from Drive: {file_id}")
                            pdf_obj.set_xy(x, y); pdf_obj.set_font("Helvetica", '', 8); pdf_obj.cell(img_width, img_height, "Error loading image", 1, 0, 'C')
                            continue
                    elif photo_b64.startswith('/local_image/'):
                        filename = photo_b64.replace('/local_image/', '')
                        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
                        local_path = os.path.join(upload_dir, filename)
                        if os.path.exists(local_path):
                            with open(local_path, 'rb') as f:
                                img_data = f.read()
                            img_stream = io.BytesIO(img_data)
                        else:
                            print(f"Failed to find local image file: {local_path}")
                            pdf_obj.set_xy(x, y); pdf_obj.set_font("Helvetica", '', 8); pdf_obj.cell(img_width, img_height, "Error loading image", 1, 0, 'C')
                            continue
                    elif photo_b64.startswith('http'):
                        # It is a URL (legacy Drive Link or external URL)
                        try:
                            headers = {'User-Agent': 'Mozilla/5.0'}
                            response = requests.get(photo_b64, headers=headers)
                            response.raise_for_status()
                            img_stream = io.BytesIO(response.content)
                        except Exception as e:
                            print(f"Failed to download image from URL: {photo_b64} - {e}")
                            pdf_obj.set_xy(x, y); pdf_obj.set_font("Helvetica", '', 8); pdf_obj.cell(img_width, img_height, "Error DL Image", 1, 0, 'C')
                            continue
                            
                    elif ',' in photo_b64: 
                        # Base64 with data URI prefix
                        photo_b64_data = photo_b64.split(',')[1]
                        img_data = base64.b64decode(photo_b64_data); img_stream = io.BytesIO(img_data)
                    else:
                        # Raw Base64 or cleanup
                        img_data = base64.b64decode(photo_b64); img_stream = io.BytesIO(img_data)
                    
                    if img_stream:
                        pdf_obj.image(img_stream, x=x, y=y, w=img_width, h=img_height)
                        
                except Exception as e:
                    pdf_obj.set_xy(x, y); pdf_obj.set_font("Helvetica", '', 8); pdf_obj.cell(img_width, img_height, "Error loading image", 1, 0, 'C')

        # --- Page 1: Survey Report (Portrait) ---
        pdf.add_page(); pdf.set_margins(10, 10, 10); add_pdf_header(pdf)
        
        # Spot Report Logic for Header
        is_spot_report = (report_type_raw == 'Spot Report')
        if is_spot_report:
            combined_heading = normalize_pdf_text_for_fpdf("Spot/Preliminary Survey Report")
        elif report_type_raw == 'Re-inspection Report':
            # User Request: Display Final Survey Report on first page for Re-inspection
            combined_heading = normalize_pdf_text_for_fpdf(f"Final Survey Report ({claim_type})")
        else:
            combined_heading = normalize_pdf_text_for_fpdf(f"{report_type} ({claim_type})")
            
        pdf.set_font("Helvetica", 'B', 12); pdf.cell(0, 8, combined_heading, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C'); pdf.ln(6)
        pdf.set_font("Helvetica", size=base_font_size_page1)
        usable_width = pdf.w - 2 * pdf.l_margin; col_width = usable_width / 2; label_w = 30; val_w = col_width - label_w
        add_plain_pair("Report No.:", get_survey_val('report_no'), "Date:", get_survey_val('report_date'), label_w, val_w)
        pdf.ln(1); pdf.set_font("Helvetica", size=base_font_size_page1 - 0.5)
        disclaimer = normalize_pdf_text_for_fpdf("This Motor Survey Report is issued without prejudice in respect of cause, nature & extent of Loss/damage & subject to to the terms and conditions of the insurance policy")
        pdf.multi_cell(0, line_h_page1, disclaimer, 0, 'L'); pdf.ln(2)
        add_plain_pair("Policy No.:", get_survey_val('policy_no'), "Insurer's Claim No.:", get_survey_val('claim_no'), label_w, val_w)
        add_plain_pair("Policy Type:", get_survey_val('policy_type_label'), "Validity:", get_survey_val('policy_validity'), label_w, val_w)
        idv_value = get_survey_val('idv')
        idv_display = f"Rs. {idv_value}" if idv_value else ""
        add_plain_pair("IDV:", idv_display, "Hypothecation:", get_survey_val('hypothecation'), label_w, val_w)
        add_section_header("1) Insurer & Insured Details")
        add_plain_pair("Insurer:", get_survey_val('insurer'), "", "", 30, usable_width - 30)
        add_plain_pair("Insured:", get_survey_val('insured'), "", "", 30, usable_width - 30)
        add_plain_pair("Insured Contact Name:", get_survey_val('insured_contact_name'), "Insured Contact No.:", get_survey_val('insured_contact_no'), 35, col_width - 35)
        add_section_header("2) Reported particulars of vehicle")
        label_w_veh = 35; val_w_veh = col_width - label_w_veh
        add_plain_pair("a. Regn. No.:", get_survey_val('vehicle_regn_no'), "b. Date of Regn.:", get_survey_val('vehicle_regn_date'), label_w_veh, val_w_veh)
        add_plain_pair("c. Chassis No.:", get_survey_val('vehicle_chassis_no'), "d. Engine No.:", get_survey_val('vehicle_engine_no'), label_w_veh, val_w_veh)
        add_plain_pair("e. Make & Model:", get_survey_val('vehicle_make_model'), "f. Type of body:", get_survey_val('vehicle_type_body'), label_w_veh, val_w_veh)
        add_plain_pair("g. C.F. Valid up to:", get_survey_val('vehicle_cf_validity'), "h. Seating capacity:", get_survey_val('vehicle_seating'), label_w_veh, val_w_veh)
        add_plain_pair("i. B.H.P./C.C.:", get_survey_val('vehicle_bhp_cc'), "j. Pre accident condition:", get_survey_val('vehicle_pre_accident_condition'), label_w_veh, val_w_veh)
        add_plain_pair("k. U.L.W.:", get_survey_val('vehicle_ulw'), "l. R.L.W.:", get_survey_val('vehicle_rlw'), label_w_veh, val_w_veh)
        add_plain_pair("m. Class of Vehicle:", get_survey_val('class_of_vehicle'), "n. Regn. Cert. No.:", get_survey_val('regn_cert_no'), label_w_veh, val_w_veh)
        add_plain_pair("o. Vehicle Colour:", get_survey_val('vehicle_colour'), "p. Odometer Reading:", get_survey_val('vehicle_odometer'), label_w_veh, val_w_veh)
        add_plain_pair("q. Tax Token No.:", get_survey_val('vehicle_tax_token'), "r. Tax Validity:", get_survey_val('vehicle_tax_validity'), label_w_veh, val_w_veh)
        add_plain_pair("s. Permit No:", get_survey_val('vehicle_permit_no'), "t. Permit Type:", get_survey_val('vehicle_permit_type'), label_w_veh, val_w_veh)
        add_plain_pair("u. Permit Validity:", get_survey_val('vehicle_permit_validity'), "v. Route area:", get_survey_val('vehicle_route_area'), label_w_veh, val_w_veh)
        add_section_header("3) Reported particulars of Driving Licence:")
        label_w_dl = 35; val_w_dl = col_width - label_w_dl
        add_plain_pair("a. Name:", get_survey_val('dl_name'), "b. Licence No.:", get_survey_val('dl_no'), label_w_dl, val_w_dl)
        add_plain_pair("c. Date of issue:", get_survey_val('dl_issue_date'), "d. Valid up to:", get_survey_val('dl_validity'), label_w_dl, val_w_dl)
        add_plain_pair("e. Issuing Authority:", get_survey_val('dl_issuing_authority'), "f. Endorsement:", get_survey_val('dl_endorsement'), label_w_dl, val_w_dl)
        add_plain_pair("g. Type of Licence:", get_survey_val('dl_type'), "h. DOB of Driver:", get_survey_val('dl_dob'), label_w_dl, val_w_dl)
        add_section_header("4) Documents Compared:")
        doc_items = [("a. Regn. Cert.:", get_survey_val('doc_regn_cert')), ("b. Driving Licence:", get_survey_val('doc_dl')), ("c. Tax Token:", get_survey_val('doc_tax_token')), ("d. Fitness Cert:", get_survey_val('doc_fitness_certificate')), ("e. Permit:", get_survey_val('doc_permit_compared')), ("f. Load Challan:", get_survey_val('doc_load_challan'))]
        doc_label_w = 35; doc_val_w = col_width - doc_label_w
        for i in range(0, len(doc_items), 2):
            item1 = doc_items[i]; item2 = doc_items[i+1] if i+1 < len(doc_items) else (None, None)
            add_plain_pair(item1[0], item1[1], item2[0] if item2[0] else "", item2[1] if item2[1] else "", doc_label_w, doc_val_w)
        add_section_header("5) Load & Goods Details:")
        load_label_w = 45; load_val_w = col_width - load_label_w
        add_plain_pair("a. Nature of Goods & Packing:", get_survey_val('load_nature_packing'), "b. Weight of Goods:", get_survey_val('load_weight_goods'), load_label_w, load_val_w)
        add_plain_pair("c. Origin & Destination:", get_survey_val('load_origin_destination'), "d. L.R./Invoice No.:", get_survey_val('load_lr_invoice_no'), load_label_w, load_val_w)
        add_plain_pair("e. Name of Transport:", get_survey_val('load_transport_name'), "f. Date:", get_survey_val('load_date'), load_label_w, load_val_w)
        add_section_header("6) Reported date & time of accident")
        label_w_acc = 45; val_w_acc = col_width - label_w_acc
        add_plain_pair("a. Date & time of accident:", get_survey_val('accident_datetime'), "b. Assign. Received:", get_survey_val('accident_assign_received'), label_w_acc, val_w_acc)
        add_plain_pair("c. Date of survey:", get_survey_val('accident_survey_date'), "d. Place of accident:", get_survey_val('accident_place'), label_w_acc, val_w_acc)
        add_plain_pair("e. Place of Survey:", get_survey_val('accident_survey_place'), "", "", label_w_acc, usable_width - label_w_acc)
        add_section_header("7) Reported Police particulars:")
        label_w_pol = 45; val_w_pol = col_width - label_w_pol
        add_plain_pair("a. Accident reported to:", get_survey_val('police_reported_to'), "b. Diary / Case No.:", get_survey_val('police_diary_case_no'), label_w_pol, val_w_pol)
        add_plain_pair("c. Date Reported:", get_survey_val('police_date_reported'), "", "", label_w_pol, usable_width - label_w_pol)
        add_section_header("8) Other Details:")
        label_w_other = 45; val_w_other = col_width - label_w_other
        add_plain_pair("Details of T. P:", get_survey_val('tp_details'), "T.P. Injury/Loss:", get_survey_val('tp_injury_loss'), label_w_other, val_w_other)
        add_plain_pair("Injury to Driver/Occupant:", get_survey_val('injury_driver_occupant'), "Are damages consistent?:", get_survey_val('damages_consistent'), label_w_other, val_w_other)
        add_plain_pair("Cause and nature of accident:", get_survey_val('accident_cause'), "", "", label_w_other, usable_width - label_w_other)
        pdf.ln(5)
        add_plain_pair("Extent of consistent damages:", get_survey_val('damages_extent'), "", "", label_w_other, usable_width - label_w_other)
        pdf.ln(2); pdf.set_font("Helvetica", 'B', base_font_size_page1); pdf.cell(15, line_h_page1, "Remark:", 0, 0)
        pdf.set_font("Helvetica", '', base_font_size_page1); pdf.multi_cell(0, line_h_page1, get_survey_val('remark'), 0, 'L')
        
        # Contd... logic: Only show if NOT a spot report
        pdf.ln(1); 
        if not is_spot_report:
            pdf.set_x(pdf.w - pdf.r_margin - 20); pdf.cell(20, line_h_page1, normalize_pdf_text_for_fpdf("contd..."), align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # --- BRANCH: Spot Report vs Final/Re-inspection ---
        if is_spot_report:
            # Spot Report Narrative (Continue on same page)
            pdf.ln(5) # Add some spacing after Remark
            
            # CHANGE: Use base_font_size_page1 (7.2) to match previous text
            pdf.set_font("Helvetica", '', base_font_size_page1)
            
            # 1. Spot Report Narrative
            if spot_report_text:
                pdf.multi_cell(0, 5, spot_report_text, 0, 'L')
            else:
                default_spot_text = "Since it is Spot/ Preliminary survey the above damages were observed without dismantling the vehicle. More damages may be unearthing after dismantling the vehicle & its parts.\n\nTotal N Nos. photographs of the insured accidentally damaged vehicle were snapped by the undersigned during the course of Spot/ Preliminary survey which are attached with my report."
                pdf.multi_cell(0, 5, normalize_pdf_text_for_fpdf(default_spot_text), 0, 'L')
            
            pdf.ln(10) # Gap between text and footer section
            
            y_footer_start = pdf.get_y()
            
            # Check for page break safety
            if y_footer_start > 250: 
                pdf.add_page(orientation='P'); add_pdf_header(pdf)
                y_footer_start = pdf.get_y()

            # 2. Enclosures (Left Side)
            pdf.set_xy(pdf.l_margin, y_footer_start)
            pdf.set_font("Helvetica", 'B', base_font_size_page1) # Bold, matching size
            pdf.cell(0, 5, "Enclosures:", 0, 1, 'L')
            
            pdf.set_font("Helvetica", '', base_font_size_page1) # Regular, matching size
            final_spot_enclosures = spot_report_enclosures if spot_report_enclosures else "1. Digital Photos\n2. Professional Bill"
            pdf.multi_cell(80, 5, final_spot_enclosures, 0, 'L')

            # 3. Signature (Right Side)
            # Align top of signature block roughly with Enclosures
            pdf.set_xy(pdf.w - pdf.r_margin - 60, y_footer_start + 5) 
            
            # Add space for Stamp/Signature image overlap
            pdf.ln(10) 
            
            pdf.set_x(pdf.w - pdf.r_margin - 60)
            pdf.set_font("Helvetica", 'B', 10) # Signature stays slightly larger/standard
            pdf.cell(60, 5, normalize_pdf_text_for_fpdf(user_full_name), 0, 1, 'C')
            
            pdf.set_x(pdf.w - pdf.r_margin - 60)
            pdf.set_font("Helvetica", '', 9)
            pdf.cell(60, 5, "( Surveyor and Loss Assessor )", 0, 1, 'C')
            
        else:
            # --- Page 2: Labour Charges (Portrait) ---
            pdf.add_page(); pdf.set_margins(10, 10, 10); add_pdf_header(pdf); pdf.set_font("Helvetica", size=base_font_size_page2); usable_width_page2 = pdf.w - pdf.l_margin - pdf.r_margin
            top_info_y = pdf.get_y() 
            pdf.set_font("Helvetica", 'B', base_font_size_page2)
            gst_text = normalize_pdf_text_for_fpdf(f"GST: {header_gst_display}"); year_text_val = normalize_pdf_text_for_fpdf(f"Vehicle Year: {header_vehicle_year}")
            pdf.set_xy(pdf.l_margin, top_info_y); pdf.cell(60, line_h_page2, gst_text, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_x(pdf.l_margin); pdf.cell(60, line_h_page2, year_text_val, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(line_h_page2 * 1.5)
            pdf.set_font("Helvetica", 'B', base_font_size_page1); pdf.cell(0, line_h_page2, normalize_pdf_text_for_fpdf("12) Allocation of Labour charges:"), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT); pdf.ln(line_h_page2 * 0.5)
            
            # Updated Columns: Name, R&R, Dent, Paint
            labour_widths = [80, 35, 35, 35]
            labour_headers_raw = ['Name of the parts', 'Removing/Refitting', 'Denting/Repairing', 'Painting']
            labour_headers = [normalize_pdf_text_for_fpdf(h) for h in labour_headers_raw]
            alignments_labour_headers = ['C', 'C', 'C', 'C']
            alignments_labour_rows = ['L', 'R', 'R', 'R']
            
            draw_table_row(labour_headers, labour_widths, line_h_page2, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(220,220,220), alignments=alignments_labour_headers, font_style='B', is_header=True, current_font_size=base_font_size_page2 -1)
            
            def safe_format_pdf_labour(val_str):
                try: num = float(str(val_str).replace(',', '')); return format_pdf_number(num)
                except (ValueError, TypeError): return normalize_pdf_text_for_fpdf(str(val_str))

            for row in processed_user_labour_rows_for_pdf:
                row_data = [
                    row.get('part_name', ''),
                    safe_format_pdf_labour(row.get('removing_refitting', '0')),
                    safe_format_pdf_labour(row.get('denting_repairing', '0')),
                    safe_format_pdf_labour(row.get('painting', '0'))
                ]
                estimated_row_height = calculate_height(row_data, labour_widths, base_font_size_page2 - 1, line_h_page2, table_cell_padding_y)
                if pdf.get_y() + estimated_row_height > pdf.page_break_trigger: pdf.add_page(); add_pdf_header(pdf); draw_table_row(labour_headers, labour_widths, line_h_page2, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(220,220,220), alignments=alignments_labour_headers, font_style='B', is_header=True, current_font_size=base_font_size_page2 -1)
                draw_table_row(row_data, labour_widths, line_h_page2, border=1, alignments=alignments_labour_rows, current_font_size=base_font_size_page2 -1)
            
            # Total Row for Columns
            total_row_data_base_labour = [
                normalize_pdf_text_for_fpdf('TOTAL, Rs'),
                format_pdf_number(labour_sum_removing),
                format_pdf_number(labour_sum_denting),
                format_pdf_number(labour_sum_painting)
            ]
            draw_table_row(total_row_data_base_labour, labour_widths, line_h_page2, border='T', alignments=alignments_labour_rows, font_style='B', current_font_size=base_font_size_page2 -1)
            
            # --- Appended Calculation Rows (Integrated Table) ---
            calc_widths = [150, 35] # Merged first 3 columns, Value in 4th
            calc_alignments = ['L', 'R']
            
            # Helper to draw calculation rows
            def draw_calc_row(label, value, bold=False):
                row_data = [normalize_pdf_text_for_fpdf(label), format_pdf_number(value)]
                font_style = 'B' if bold else ''
                draw_table_row(row_data, calc_widths, line_h_page2, border=1, alignments=calc_alignments, font_style=font_style, current_font_size=base_font_size_page2 -1)

            # 1. Less 12.5% Dep
            draw_calc_row("Less: 12.5% (on paint material)", labour_paint_depn_final)
            
            # 2. Less 50% IMT 23
            if labour_imt_deduction > 0:
                draw_calc_row("Less: 50% Liability (As per IMT 23 Norms)", labour_imt_deduction)

            # 3. Total Painting Charges - Net Liability for Paint
            draw_calc_row("Total Painting Charges", net_paint_liability, bold=True)
                
            # 4. Add Labour (R&R + Dent)
            draw_calc_row("Add: Labour (R&R + Dent)", labour_rr_dent_sum)
            
            # 5. Taxable Total (Bold)
            draw_calc_row("Total Taxable Labour", taxable_labour, bold=True)
            
            # 6. Add GST
            if labour_tax_type_main != 'Zero':
                draw_calc_row("Add: 18% GST", labour_gst_amount)
            
            # 7. Final Amount (Bold)
            draw_calc_row("Final Labour Liability", labour_grand_total_final, bold=True)
            
            pdf.ln(line_h_page2 * 1.5)
            
            # --- Page 3: Landscape Table 13 ---
            pdf.add_page(orientation='L'); pdf.set_margins(10, 10, 10); add_pdf_header(pdf)
            pdf.set_auto_page_break(auto=False, margin=10) 
            pdf.set_font("Helvetica", size=base_font_size_page2)
            pdf.set_font("Helvetica", 'B', base_font_size_page1); pdf.cell(0, line_h_page2, normalize_pdf_text_for_fpdf("14. Cost of Spare Parts at MRP. :"), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT); pdf.ln(line_h_page2 * 0.5)
            parts_widths = [8, 8, 8, 55, 15, 19, 19, 19, 11, 8, 15, 23, 10, 15, 8, 15, 25]
            parts_headers_raw = ['Sl', 'E\nNo', 'Bill\nSL', 'Parts Descriptions', 'HNS\nCODE', 'Estimate\nAmount', 'Bill\nAmount', 'Assessed\nAmount', 'Parts\nType', 'Dep\n%', 'Dep.\nAMT', 'Net.\nAmount', 'GST\n%', 'GST\nAmount', 'IMT\n23', 'IMT-23\nAMT', 'Net AMT\nIncl. GST']
            parts_headers = [normalize_pdf_text_for_fpdf(h) for h in parts_headers_raw]
            alignments_parts_headers = ['C'] * 17
            alignments_parts_rows = ['C', 'C', 'C', 'L', 'C', 'R', 'R', 'R', 'C', 'C', 'R', 'R', 'C', 'R', 'C', 'R', 'R']
            
            draw_table_row(parts_headers, parts_widths, line_h_page2, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(200,200,200), alignments=alignments_parts_headers, font_style='B', is_header=True, current_font_size=base_font_size_page2 -1)
            pdf.set_text_color(0,0,0) 

            for i, part in enumerate(final_parts_calculated):
                sl_no_str = normalize_pdf_text_for_fpdf(str(int(part.get('sl_no', 0))))
                est_sl_no_str = normalize_pdf_text_for_fpdf(part.get('est_sl_no', sl_no_str))
                bill_sl_no_str = normalize_pdf_text_for_fpdf(part.get('bill_sl_no', sl_no_str))
                dep_pc_display = "0%"
                if part.get('total_parts_amt', 0) > 0 and part.get('depr', 0) > 0:
                    rate = (part.get('depr', 0) / part.get('total_parts_amt', 0)) * 100
                    dep_pc_display = f"{rate:.0f}%"
                elif policy_type in ('NIL_DEPN', 'NIL_DEPN_PLUS'): dep_pc_display = "NIL"
                imt_23_display = "NIL"
                if part.get('imt_23_amt', 0) > 0: imt_23_display = "YES"

                part_data = [
                    sl_no_str, est_sl_no_str, bill_sl_no_str, part.get('part_name_display', ''), normalize_pdf_text_for_fpdf(part.get('hns_code', '')), 
                    format_pdf_number(part.get('estimate_amt', 0.0)), format_pdf_number(part.get('bill_amt', 0.0)), format_pdf_number(part.get('total_parts_amt', 0.0)),   
                    normalize_pdf_text_for_fpdf(part.get('type_part', '')), normalize_pdf_text_for_fpdf(dep_pc_display), format_pdf_number(part.get('depr', 0.0)),              
                    format_pdf_number(part.get('net_base', 0.0)), normalize_pdf_text_for_fpdf(f"{part.get('original_gst_pc', 0):.0f}%"), format_pdf_number(part.get('total_gst', 0.0)),         
                    normalize_pdf_text_for_fpdf(imt_23_display), format_pdf_number(part.get('imt_23_amt', 0.0)), format_pdf_number(part.get('net_amt', 0.0))          
                ]
                dummy_pdf = FPDF(); dummy_pdf.add_page(); dummy_pdf.set_font("Helvetica", '', base_font_size_page2 - 1)
                max_lines = 1
                for j, txt in enumerate(part_data):
                    lines = dummy_pdf.multi_cell(parts_widths[j], line_h_page2, str(txt), dry_run=True, output="LINES")
                    max_lines = max(max_lines, len(lines))
                row_h = max_lines * line_h_page2 + table_cell_padding_y
                is_last_row = (i == len(final_parts_calculated) - 1); needed_height = row_h
                if is_last_row: needed_height += (line_h_page2 + table_cell_padding_y) 
                if pdf.get_y() + needed_height > pdf.h - pdf.b_margin:
                    pdf.add_page(orientation='L'); add_pdf_header(pdf)
                    draw_table_row(parts_headers, parts_widths, line_h_page2, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(200,200,200), alignments=alignments_parts_headers, font_style='B', is_header=True, current_font_size=base_font_size_page2 -1)
                    pdf.set_text_color(0,0,0)
                draw_table_row(part_data, parts_widths, line_h_page2, border=1, alignments=alignments_parts_rows, current_font_size=base_font_size_page2 -1)
            
            # Total Row
            parts_total_row = ['', '', '', 'TOTAL', '', format_pdf_number(parts_total_estimate), format_pdf_number(parts_total_bill), format_pdf_number(parts_total_assessed), '', '', format_pdf_number(parts_depr_sum_final), format_pdf_number(parts_total_assessed - parts_depr_sum_final), '', format_pdf_number(parts_total_gst_final), '', format_pdf_number(parts_total_imt23), format_pdf_number(parts_net_amt_final)]        
            draw_table_row(parts_total_row, parts_widths, line_h_page2, border='T', alignments=alignments_parts_rows, font_style='B', current_font_size=base_font_size_page2 -1)
            pdf.ln(line_h_page2); pdf.set_auto_page_break(auto=True, margin=10)

            # --- Summary Section (Landscape) ---
            summary_start_y = pdf.get_y()
            # INCREASED THRESHOLD: Was 80, now 110 to ensure meaningful summary fits or start new page
            if pdf.h - summary_start_y < 110: 
                pdf.add_page(orientation='L'); add_pdf_header(pdf); summary_start_y = pdf.get_y()
            
            # Left Column: Estimates
            left_col_x = pdf.l_margin; left_col_width = 80
            pdf.set_xy(left_col_x, summary_start_y)
            pdf.set_font("Helvetica", 'B', base_font_size_page2)
            pdf.cell(left_col_width/2, line_h_page2, "Estimates", 1, 0, 'L'); pdf.cell(left_col_width/2, line_h_page2, "Amount", 1, 1, 'R')
            pdf.set_font("Helvetica", '', base_font_size_page2)
            
            # Determine values to use (Override or Calculated)
            est_labour_val = float(est_labour_override) if est_labour_override and is_number(est_labour_override) else labour_rr_dent_sum
            est_paint_val = float(est_paint_override) if est_paint_override and is_number(est_paint_override) else labour_sum_painting
            est_parts_val = float(est_parts_override) if est_parts_override and is_number(est_parts_override) else parts_total_estimate
            
            pdf.cell(left_col_width/2, line_h_page2, "Labour Charges", 1, 0, 'L'); pdf.cell(left_col_width/2, line_h_page2, format_pdf_number(est_labour_val), 1, 1, 'R')
            pdf.cell(left_col_width/2, line_h_page2, "Paint cost", 1, 0, 'L'); pdf.cell(left_col_width/2, line_h_page2, format_pdf_number(est_paint_val), 1, 1, 'R')
            pdf.cell(left_col_width/2, line_h_page2, "Cost of Parts", 1, 0, 'L'); pdf.cell(left_col_width/2, line_h_page2, format_pdf_number(est_parts_val), 1, 1, 'R') 
            pdf.set_font("Helvetica", 'B', base_font_size_page2)
            approx_total = est_labour_val + est_paint_val + est_parts_val
            pdf.cell(left_col_width/2, line_h_page2, "Approximate Total", 1, 0, 'L'); pdf.set_fill_color(255, 255, 0)
            pdf.cell(left_col_width/2, line_h_page2, format_pdf_number(approx_total), 1, 1, 'R', fill=True)
            left_col_end_y = pdf.get_y()

            # Right Column: NEW Particulars of Liability Table
            # Layout: Descriptions | Assessed Amount | Total GST on A/Amount | Liability Amount Inclu. GST
            right_col_x = pdf.l_margin + left_col_width + 5
            
            # UPDATED FIXED WIDTHS: Reduced Desc, Increased others to fit text
            # Total width: 55 + 35 + 42 + 45 = 177mm (Fits in remaining ~190mm space)
            cols_liability = [55, 35, 42, 45] 
            
            headers_liability = ["Particulars of Liability", "Assessed Amount", "Total GST on A/Amount", "Liability Amount Inclu. GST"]
            aligns_liability = ['L', 'R', 'R', 'R']
            
            pdf.set_xy(right_col_x, summary_start_y)
            
            # Table Headers
            current_x = right_col_x
            header_labels = ["Descriptions", "Assessed Amount", "Total GST on A/Amount", "Liability Amount Inclu. GST"]
            # Color header blue-ish like image
            pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255,255,255) # Blue background, White text
            
            # Use smaller font for headers to prevent overlap
            pdf.set_font("Helvetica", 'B', base_font_size_page2 - 1.5) 
            
            for i, h in enumerate(header_labels):
                pdf.set_xy(current_x, pdf.get_y())
                # Normalize text to handle special chars if any
                safe_h = normalize_pdf_text_for_fpdf(h)
                pdf.cell(cols_liability[i], line_h_page2 + 2, safe_h, 1, 0, 'C', fill=True)
                current_x += cols_liability[i]
            pdf.ln(line_h_page2 + 2); pdf.set_text_color(0,0,0)

            def add_new_summary_row(desc, assessed, gst, liability, bold=False, fill_color=None):
                pdf.set_x(right_col_x)
                if fill_color: pdf.set_fill_color(*fill_color)
                else: pdf.set_fill_color(255, 255, 255)
                
                # Reset font to normal size for data rows
                pdf.set_font("Helvetica", 'B' if bold else '', base_font_size_page2)
                
                # Draw cells
                c_x = right_col_x
                pdf.cell(cols_liability[0], line_h_page2, normalize_pdf_text_for_fpdf(desc), 1, 0, 'L', fill=bool(fill_color))
                c_x += cols_liability[0]; pdf.set_x(c_x)
                pdf.cell(cols_liability[1], line_h_page2, format_pdf_number(assessed) if assessed != '' else '', 1, 0, 'R', fill=bool(fill_color))
                c_x += cols_liability[1]; pdf.set_x(c_x)
                pdf.cell(cols_liability[2], line_h_page2, format_pdf_number(gst) if gst != '' else '', 1, 0, 'R', fill=bool(fill_color))
                c_x += cols_liability[2]; pdf.set_x(c_x)
                pdf.cell(cols_liability[3], line_h_page2, format_pdf_number(liability) if liability != '' else '', 1, 1, 'R', fill=bool(fill_color))

            # 1. Labour
            labour_assessed = labour_rr_dent_sum
            labour_gst = 0.0
            if labour_tax_type_main != 'Zero':
                labour_gst = labour_assessed * 0.18
            labour_liability = labour_assessed + labour_gst
            add_new_summary_row("TOTAL LABOUR", labour_assessed, labour_gst, labour_liability)

            # 2. Paint
            # Logic: "Net Paint Liability" (after Dep/IMT) is the assessed amount. 18% tax on that.
            paint_assessed = net_paint_liability
            paint_gst = 0.0
            if labour_tax_type_main != 'Zero':
                 paint_gst = paint_assessed * 0.18
            paint_liability = paint_assessed + paint_gst
            add_new_summary_row("TOTAL PAINT", paint_assessed, paint_gst, paint_liability)

            # 3. Parts (Metal, Glass, Plastic)
            cat_data = {'M': {'gst': 0.0, 'liability': 0.0}, 'G': {'gst': 0.0, 'liability': 0.0}, 'P': {'gst': 0.0, 'liability': 0.0}}
            
            for part in final_parts_calculated:
                p_type = str(part.get('type_part', 'M')).strip().upper()
                if p_type not in ['M', 'G', 'P']: p_type = 'M'
                
                p_net = float(part.get('net_amt', 0.0))
                gst_rate = float(part.get('original_gst_pc', 0.0))
                
                p_base_component = 0.0
                p_tax_component = 0.0
                
                # Split Liability back into Base and Tax components for display
                if gst_rate > 0:
                    p_base_component = p_net / (1 + (gst_rate / 100.0))
                    p_tax_component = p_net - p_base_component
                else:
                    p_base_component = p_net
                    p_tax_component = 0.0
                    
                cat_data[p_type]['liability'] += p_net
                cat_data[p_type]['gst'] += p_tax_component

            # Metal
            m_assessed = cat_data['M']['liability'] - cat_data['M']['gst']
            add_new_summary_row("TOTAL METAL PARTS", m_assessed, cat_data['M']['gst'], cat_data['M']['liability'])
            
            # Glass
            g_assessed = cat_data['G']['liability'] - cat_data['G']['gst']
            add_new_summary_row("TOTAL GLASS PARTS", g_assessed, cat_data['G']['gst'], cat_data['G']['liability'])
            
            # Plastic
            p_assessed = cat_data['P']['liability'] - cat_data['P']['gst']
            add_new_summary_row("TOTAL PLASTIC PARTS", p_assessed, cat_data['P']['gst'], cat_data['P']['liability'])

            # Total Row
            total_liability = labour_liability + paint_liability + cat_data['M']['liability'] + cat_data['G']['liability'] + cat_data['P']['liability']
            # Blue background row like image
            pdf.set_x(right_col_x)
            pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica", 'B', base_font_size_page2)
            pdf.cell(sum(cols_liability[:3]), line_h_page2, "Total :", 1, 0, 'R', fill=True)
            pdf.cell(cols_liability[3], line_h_page2, format_pdf_number(total_liability), 1, 1, 'R', fill=True)
            pdf.set_text_color(0,0,0)

            # Less Rows
            def add_less_row(label, val):
                pdf.set_x(right_col_x)
                pdf.set_fill_color(255, 255, 255); pdf.set_font("Helvetica", '', base_font_size_page2)
                pdf.cell(sum(cols_liability[:3]), line_h_page2, label, 1, 0, 'R')
                pdf.cell(cols_liability[3], line_h_page2, format_pdf_number(val), 1, 1, 'R')

            add_less_row("Less : Salvage", salvage_val_numeric)
            add_less_row("Less: Compulsory excess", excess_final)
            add_less_row("Less: Impose excess", impose_excess_final)
            
            # Net Settlement
            net_settlement = total_liability - salvage_val_numeric - excess_final - impose_excess_final
            add_less_row("Net settlement Amount :", net_settlement)
            
            # ND Deduction Row (only for NIL_DEPN policy)
            if policy_type == 'NIL_DEPN' and nd_deduction_amount > 0:
                nd_label = f"Less: {nd_deduction_pc:g}% on assessed amount as per ND policy norms"
                add_less_row(nd_label, nd_deduction_amount)
                net_settlement -= nd_deduction_amount
            
            # Towing Charges Row (only if non-zero)
            if towing_charges > 0:
                add_less_row("Add: Towing Charges", towing_charges)
                net_settlement += towing_charges
            
            # Round off (Blue Row) - SAFETY CHECK FOR SIGNATURE
            # Ensure this row + Signature height fits on the page. 
            # If not, break page here so this row travels with the signature.
            sig_height_block = 35 # Height for signature block
            
            # FIXED: Added logic to break page if not enough space for Round Off + Signature
            if pdf.get_y() + line_h_page2 + sig_height_block + 15 > pdf.page_break_trigger:
                pdf.add_page(orientation='L')
                add_pdf_header(pdf)
                # Keep X alignment on new page
            
            pdf.set_x(right_col_x)
            pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica", 'B', base_font_size_page2)
            pdf.cell(sum(cols_liability[:3]), line_h_page2, "Net settlement Amount Round off:", 1, 0, 'R', fill=True)
            pdf.cell(cols_liability[3], line_h_page2, format_pdf_number(round(net_settlement)), 1, 1, 'R', fill=True)
            pdf.set_text_color(0,0,0)

            # Optional Note & Enclosures
            final_y = pdf.get_y() + 5
            
            if parts_table_note:
                pdf.set_y(final_y)
                pdf.set_x(pdf.l_margin) # Reset X to left margin
                pdf.set_font("Helvetica", 'B', base_font_size_page2)
                pdf.multi_cell(0, line_h_page2, parts_table_note, 0, 'L')
                final_y = pdf.get_y() + 2

            if enclosures_text:
                # Calculate needed height for enclosures
                pdf.set_font("Helvetica", '', base_font_size_page2)
                pdf.set_x(pdf.l_margin) # Reset X explicitly for dry run
                lines = pdf.multi_cell(0, line_h_page2, enclosures_text, dry_run=True, output="LINES")
                enc_h = len(lines) * line_h_page2 + 10 # Title + Lines
                
                if final_y + enc_h + sig_height_block > pdf.page_break_trigger:
                    pdf.add_page(orientation='L')
                    add_pdf_header(pdf)
                    final_y = pdf.get_y() + 5
                
                pdf.set_y(final_y)
                pdf.set_x(pdf.l_margin)
                pdf.set_font("Helvetica", 'B', base_font_size_page2)
                pdf.cell(0, line_h_page2, "Enclosures:", 0, 1, 'L')
                pdf.set_font("Helvetica", '', base_font_size_page2)
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(0, line_h_page2, enclosures_text, 0, 'L')
                final_y = pdf.get_y() + 5
            
            # Print Signature (Guaranteed to have content above it now)
            # Add gap for Stamp (Requested by User)
            gap_stamp = 40
            sig_lines_height = line_h_page2 * 3
            
            if pdf.get_y() + gap_stamp + sig_lines_height > pdf.page_break_trigger:
                pdf.add_page(orientation='L')
                add_pdf_header(pdf)
                pdf.set_y(pdf.get_y() + 30) # Gap on new page
            else:
                pdf.set_y(pdf.get_y() + gap_stamp)
            pdf.set_x(pdf.w - pdf.r_margin - 60); pdf.cell(60, line_h_page2, normalize_pdf_text_for_fpdf(user_full_name), 0, 1, 'C')
            pdf.set_x(pdf.w - pdf.r_margin - 60); pdf.cell(60, line_h_page2, "( Surveyor and Loss Assessor )", 0, 1, 'C')

            # --- Re-inspection Report Page (Optional) ---
            if report_type_raw == 'Re-inspection Report':
                pdf.add_page(orientation='P'); pdf.set_margins(10, 10, 10); add_pdf_header(pdf); pdf.set_auto_page_break(auto=False, margin=10) 
                pdf.set_font("Helvetica", 'B', 14); pdf.cell(0, 10, "RE-INSPECTION REPORT", 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C'); pdf.ln(5)
                pdf.set_font("Helvetica", 'B', 10); pdf.cell(15, 5, "To,", 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", '', 10)
                start_y_addr = pdf.get_y(); insurer_address = get_survey_val('insurer'); pdf.set_xy(pdf.l_margin + 15, start_y_addr); pdf.multi_cell(0, 5, insurer_address, 0, 'L'); pdf.ln(5)
                pdf.set_font("Helvetica", 'B', 10); pdf.cell(15, 5, "Sub:", 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", '', 10); pdf.cell(0, 5, f"Re-inspection of repaired vehicle bearing Regn. No. {get_survey_val('vehicle_regn_no')}", 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT); pdf.ln(5)
                pdf.set_font("Helvetica", '', 9); col1_w = 40; col2_w = 60; col3_w = 40; col4_w = 50 
                def draw_grid_row(label1, val1, label2=None, val2=None):
                    h = 7; pdf.set_font("Helvetica", 'B', 9); pdf.cell(col1_w, h, label1, 1, 0, 'L'); pdf.set_font("Helvetica", '', 9); pdf.cell(col2_w, h, val1, 1, 0, 'L')
                    if label2: pdf.set_font("Helvetica", 'B', 9); pdf.cell(col3_w, h, label2, 1, 0, 'L'); pdf.set_font("Helvetica", '', 9); pdf.cell(col4_w, h, val2, 1, 1, 'L')
                    else: pdf.ln(h)
                draw_grid_row("Policy No. :", get_survey_val('policy_no'), "Claim No. :", get_survey_val('claim_no')); draw_grid_row("Date of Accident :", get_survey_val('accident_datetime'), "Date of Survey :", get_survey_val('accident_survey_date'))
                pdf.set_font("Helvetica", 'B', 9); pdf.cell(col1_w, 7, "Insured :", 1, 0, 'L'); x_val = pdf.get_x(); y_val = pdf.get_y()
                pdf.set_font("Helvetica", '', 9); insured_val = get_survey_val('insured'); width_val = col2_w + col3_w + col4_w; pdf.set_xy(x_val, y_val); pdf.multi_cell(width_val, 7, insured_val, 1, 'L'); pdf.set_x(pdf.l_margin)
                draw_grid_row("Chassis No. :", get_survey_val('vehicle_chassis_no'), "Engine No. :", get_survey_val('vehicle_engine_no')); pdf.ln(5)
                pdf.set_font("Helvetica", '', 10); intro_text = "As per instruction received from your office, I visited the repairer's workshop. I have inspected the subject vehicle after repairs and verified the replaced parts. The details are as follows:"; pdf.multi_cell(0, 5, intro_text, 0, 'L'); pdf.ln(5)
                re_cols = [12, 108, 30, 40]; re_headers = ["SL", "Name of the Parts", "Salvage", "Remarks"]; re_aligns = ['C', 'L', 'C', 'C']
                draw_table_row(re_headers, re_cols, 7, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(220,220,220), alignments=re_aligns, font_style='B', is_header=True, current_font_size=9); pdf.set_text_color(0,0,0)
                for part in final_parts_calculated:
                    sl = str(part.get('sl_no')); name = part.get('part_name_display', ''); salvage = part.get('salvage_produce', 'YES'); remarks = part.get('remarks', 'REPLACED BY NEW'); row_data = [sl, name, salvage, remarks]
                    if pdf.get_y() > 260: pdf.add_page(orientation='P'); add_pdf_header(pdf); draw_table_row(re_headers, re_cols, 7, border=1, align='C', fill=True, text_color=(0,0,0), fill_color=(220,220,220), alignments=re_aligns, font_style='B', is_header=True, current_font_size=9)
                    draw_table_row(row_data, re_cols, 6, border=1, alignments=re_aligns, current_font_size=9)
                pdf.ln(5); pdf.set_font("Helvetica", 'B', 10); pdf.cell(0, 5, "Observations / Remarks:", 0, 1, 'L'); pdf.set_font("Helvetica", '', 10); pdf.multi_cell(0, 5, reinspection_note, 0, 'L'); pdf.ln(10)
                if pdf.get_y() > 250: pdf.add_page(orientation='P'); add_pdf_header(pdf)
                pdf.set_x(pdf.w - pdf.r_margin - 60); pdf.set_font("Helvetica", 'B', 10); pdf.cell(60, 5, normalize_pdf_text_for_fpdf(user_full_name), 0, 1, 'C'); pdf.set_x(pdf.w - pdf.r_margin - 60); pdf.set_font("Helvetica", '', 9); pdf.cell(60, 5, "( Surveyor and Loss Assessor )", 0, 1, 'C'); pdf.set_auto_page_break(auto=True, margin=15)

            # --- Page 4: Survey Fee Bill (Portrait) ---
        pdf.add_page(orientation='P'); pdf.set_margins(10, 10, 10); add_pdf_header(pdf); pdf.set_font("Helvetica", size=base_font_size_page3); usable_width_page3 = pdf.w - pdf.l_margin - pdf.r_margin
        pdf.set_font("Helvetica", 'B', 14); pdf.cell(0, 10, normalize_pdf_text_for_fpdf("SURVEY FEE BILL"), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C'); pdf.ln(5)
        pdf.set_font("Helvetica", '', base_font_size_page3)
        ref_no_text_val = normalize_pdf_text_for_fpdf(f"Ref. No - {get_survey_val('report_no')}"); date_text_val = normalize_pdf_text_for_fpdf(f"Date- {get_survey_val('report_date')}"); date_width = pdf.get_string_width(date_text_val) + 2
        pdf.cell(usable_width_page3 - date_width, line_h_page3, ref_no_text_val, 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.cell(date_width, line_h_page3, date_text_val, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
        current_y_before_to = pdf.get_y() + line_h_page3 * 0.25; pdf.set_y(current_y_before_to); pdf.cell(10, line_h_page3, normalize_pdf_text_for_fpdf("To"), 0, new_x=XPos.RIGHT, new_y=YPos.TOP); address_start_x = pdf.l_margin + 10; pdf.set_xy(address_start_x, current_y_before_to) 
        insurer_address_full_raw = get_survey_val('insurer'); insurer_address_full = normalize_pdf_text_for_fpdf(insurer_address_full_raw); insurer_address_lines = [line.strip() for line in insurer_address_full.split(',') if line.strip()]
        for i, line in enumerate(insurer_address_lines):
            if pdf.get_y() + line_h_page3 > pdf.page_break_trigger - 10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3); pdf.set_xy(address_start_x, pdf.get_y())
            pdf.multi_cell(usable_width_page3 - (address_start_x - pdf.l_margin), line_h_page3, line, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L') 
            if i < len(insurer_address_lines) -1 : pdf.set_x(address_start_x)
        
        # Company GSTIN under address
        if p3_company_gstin:
            pdf.set_x(address_start_x)
            pdf.set_font("Helvetica", 'B', base_font_size_page3)
            pdf.cell(0, line_h_page3, normalize_pdf_text_for_fpdf(f"GSTIN: {p3_company_gstin}"), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        pdf.set_x(pdf.l_margin); pdf.ln(line_h_page3 * 0.5)
        if p3_customer_gstin:
            pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.cell(pdf.get_string_width("INSURED GST NO- ") + 1, line_h_page3, normalize_pdf_text_for_fpdf("INSURED GST NO- "), 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", '', base_font_size_page3); pdf.multi_cell(0, line_h_page3, p3_customer_gstin, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L'); pdf.ln(line_h_page3 * 0.25) 
        table_data_p3_header = [("Policy No:", get_survey_val('policy_no')), ("Claim No.-", get_survey_val('claim_no')), ("Regd No:", get_survey_val('vehicle_regn_no')), ("Insured :", get_survey_val('insured'))]
        label_col_width_p3_info_table = 30; value_col_width_p3_info_table = usable_width_page3 - label_col_width_p3_info_table
        for label_raw, value in table_data_p3_header:
            label = normalize_pdf_text_for_fpdf(label_raw) 
            if value:
                row_height_est = calculate_height([label, value], [label_col_width_p3_info_table, value_col_width_p3_info_table], base_font_size_page3, line_h_page3, table_cell_padding_y/2, ['', 'B'])
                if pdf.get_y() + row_height_est > pdf.page_break_trigger -10 : pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
                pdf.set_font("Helvetica", '', base_font_size_page3); pdf.cell(label_col_width_p3_info_table, line_h_page3, label, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.multi_cell(value_col_width_p3_info_table, line_h_page3, value, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L', max_line_height=line_h_page3) 
        pdf.ln(line_h_page3); pdf.set_font("Helvetica", '', base_font_size_page3); fee_name_width = usable_width_page3 * 0.75; fee_amount_width = usable_width_page3 * 0.25
        for item in p3_valid_fee_items: 
            if pdf.get_y() + line_h_page3 > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
            pdf.cell(fee_name_width, line_h_page3, item['name'], border=1); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(item['amount']), border=1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        if pdf.get_y() + line_h_page3 > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
        photo_desc_raw = f"{p3_photo_copies_count} photograph copies @ Rs 10/- per Photograph"; photo_desc = normalize_pdf_text_for_fpdf(photo_desc_raw)
        pdf.cell(fee_name_width, line_h_page3, photo_desc, border=1); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_photo_total_charge), border=1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        if p3_total_before_gst != 0:
            pdf.set_font("Helvetica", '', base_font_size_page3)
            if p3_apply_gst:
                gst_lines_needed = 3 if labour_tax_type_main != 'IGST' else 2
                if pdf.get_y() + (line_h_page3 * gst_lines_needed) > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
                pdf.cell(fee_name_width, line_h_page3, normalize_pdf_text_for_fpdf("Sub Total"), border='LR', align='R'); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_total_before_gst), border='LR', align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                if labour_tax_type_main == 'IGST':
                    if p3_igst != 0: pdf.cell(fee_name_width, line_h_page3, normalize_pdf_text_for_fpdf("Add, 18% IGST"), border='LR', align='R'); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_igst), border='LR', align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                else: # Survey fee GST is always 18% (CGST/SGST), even when labour tax is Zero
                    if p3_cgst != 0: pdf.cell(fee_name_width, line_h_page3, normalize_pdf_text_for_fpdf("Add, 9% CGST"), border='LR', align='R'); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_cgst), border='LR', align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    if p3_sgst != 0: pdf.cell(fee_name_width, line_h_page3, normalize_pdf_text_for_fpdf("Add, 9% SGST"), border='LR', align='R'); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_sgst), border='LR', align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            else:
                if pdf.get_y() + line_h_page3 > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
            
            pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.cell(fee_name_width, line_h_page3, normalize_pdf_text_for_fpdf("Total Rupees"), border=1, align='R'); pdf.cell(fee_amount_width, line_h_page3, format_pdf_number(p3_grand_total), border=1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(line_h_page3 * 0.5)
        
        # --- Surveyor Bank Details (Bottom of Page 4) ---
        if surveyor_details:
            # Check space for bank details (approx 8 lines)
            if pdf.get_y() + (line_h_page3 * 8) > pdf.page_break_trigger: pdf.add_page(orientation='P'); add_pdf_header(pdf)
            
            pdf.set_font("Helvetica", 'B', base_font_size_page3)
            
            # Left Column: GSTIN, PAN, Bank Name, A/c, MICR, IFSC
            start_x = pdf.l_margin
            col1_w = 25; col2_w = 60
            
            def add_bank_row(label, value):
                pdf.set_x(start_x)
                pdf.cell(col1_w, line_h_page3, normalize_pdf_text_for_fpdf(label), 0, 0, 'L')
                pdf.cell(col2_w, line_h_page3, normalize_pdf_text_for_fpdf(value), 0, 1, 'L')

            add_bank_row("GSTIN :", surveyor_details.get('gstin', ''))
            add_bank_row("PAN :", surveyor_details.get('pan', ''))
            add_bank_row("Bank Name :", surveyor_details.get('bank_name', ''))
            add_bank_row("A/c NO. :", surveyor_details.get('account_no', ''))
            add_bank_row("MICR No. :", surveyor_details.get('micr', ''))
            add_bank_row("IFS Code :", surveyor_details.get('ifsc', ''))
            
            # Right Column: Surveyor Code & State
            state_val = surveyor_details.get('state_code', '(19)')
            code_val = surveyor_details.get('surveyor_code', '2075995')

            # Position for Right Column (adjust Y to align with top of bank details)
            pdf.set_xy(start_x + col1_w + col2_w + 10, pdf.get_y() - (line_h_page3 * 6)) 
            
            pdf.cell(50, line_h_page3, normalize_pdf_text_for_fpdf(f"State : {state_val}"), 0, 1, 'L')
            
            pdf.set_x(start_x + col1_w + col2_w + 10)
            pdf.cell(50, line_h_page3, normalize_pdf_text_for_fpdf(f"Insurer's Surveyor Code No.: {code_val}"), 0, 1, 'L')
            
            # Reset Y to below bank details
            pdf.set_y(pdf.get_y() + (line_h_page3 * 5))

        pdf.ln(line_h_page3 * 0.5) 
        if p3_grand_total_in_words and p3_grand_total != 0 : 
            words_height_est = calculate_height([p3_grand_total_in_words], [usable_width_page3 - (pdf.get_string_width("Rupees ( In Words)- ") + 1)], base_font_size_page3, line_h_page3, table_cell_padding_y/2, ['B'])
            if pdf.get_y() + words_height_est > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf)
            pdf.set_font("Helvetica", '', base_font_size_page3); label_text_raw = "Rupees ( In Words)-"; label_text = normalize_pdf_text_for_fpdf(label_text_raw); current_label_width = pdf.get_string_width(label_text + " ") + 1
            pdf.cell(current_label_width, line_h_page3, label_text, 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.multi_cell(usable_width_page3 - current_label_width, line_h_page3, p3_grand_total_in_words, 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
        pdf.ln(line_h_page3 * 0.5); pdf.set_font("Helvetica", '', base_font_size_page3)
        if pdf.get_y() + line_h_page3 > pdf.page_break_trigger-10: pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
        label_text_raw = "Estimated Amount = Rs."; label_text = normalize_pdf_text_for_fpdf(label_text_raw); current_label_width = pdf.get_string_width(label_text + " ") + 1
        pdf.cell(current_label_width, line_h_page3, label_text, 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.cell(usable_width_page3 - current_label_width, line_h_page3, format_pdf_number(p3_estimated_amount), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT); pdf.set_font("Helvetica", '', base_font_size_page3); pdf.ln(line_h_page3 * 0.25) 
        # --- Final Assessed Amount & Signature Block ---
        # Calculate combined height for "Assessed Amount" + Gap + Signature to keep them together
        sig_block_width = usable_width_page3 * 0.5
        sig_start_x = pdf.w - pdf.r_margin - sig_block_width
        sig_height_est = line_h_page3 * 4 # Signature block height approx
        
        # Check if we need to print "Assessed Amount"
        assessed_amt_height = line_h_page3 if net_liability_final != 0 else 0
        
        # Total logic block height: Assessed Amt (if any) + Gap (2 lines) + Signature
        gap_lines = 2
        total_block_needed = assessed_amt_height + (line_h_page3 * gap_lines) + sig_height_est
        
        # Check page break trigger for the WHOLE block
        if pdf.get_y() + total_block_needed > pdf.page_break_trigger: 
            pdf.add_page(orientation='P'); add_pdf_header(pdf); pdf.set_font("Helvetica", '', base_font_size_page3)
        
        # 1. Print Assessed Amount (if applicable)
        if net_liability_final != 0: 
            pdf.set_font("Helvetica", '', base_font_size_page3)
            
            # ND Deduction line (only for NIL_DEPN policy)
            if policy_type == 'NIL_DEPN' and nd_deduction_amount > 0:
                nd_label_raw = f"Less: {nd_deduction_pc:g}% on assessed amount as per ND policy norms"
                nd_label_text = normalize_pdf_text_for_fpdf(nd_label_raw)
                nd_label_width = pdf.get_string_width(nd_label_text + " ") + 1
                pdf.cell(nd_label_width, line_h_page3, nd_label_text, 0, new_x=XPos.RIGHT, new_y=YPos.TOP)
                pdf.set_font("Helvetica", 'B', base_font_size_page3)
                pdf.cell(usable_width_page3 - nd_label_width, line_h_page3, format_pdf_number(nd_deduction_amount), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", '', base_font_size_page3)
            
            # Towing Charges line (only if non-zero)
            if towing_charges > 0:
                tow_label_raw = "Add: Towing Charges"
                tow_label_text = normalize_pdf_text_for_fpdf(tow_label_raw)
                tow_label_width = pdf.get_string_width(tow_label_text + " ") + 1
                pdf.cell(tow_label_width, line_h_page3, tow_label_text, 0, new_x=XPos.RIGHT, new_y=YPos.TOP)
                pdf.set_font("Helvetica", 'B', base_font_size_page3)
                pdf.cell(usable_width_page3 - tow_label_width, line_h_page3, format_pdf_number(towing_charges), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", '', base_font_size_page3)
            
            label_text_raw = "Net settlement Amount Round off:"; label_text = normalize_pdf_text_for_fpdf(label_text_raw); current_label_width = pdf.get_string_width(label_text + " ") + 1
            pdf.cell(current_label_width, line_h_page3, label_text, 0, new_x=XPos.RIGHT, new_y=YPos.TOP); pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.cell(usable_width_page3 - current_label_width, line_h_page3, format_pdf_number(net_liability_final), 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # 2. Gap
        pdf.ln(line_h_page3 * gap_lines)
        
        # 3. Signature
        pdf.set_x(sig_start_x); pdf.set_font("Helvetica", 'B', base_font_size_page3); pdf.cell(sig_block_width, line_h_page3, normalize_pdf_text_for_fpdf(user_full_name), 0, 1, 'C')
        pdf.set_x(sig_start_x); pdf.set_font("Helvetica", '', base_font_size_page3); pdf.cell(sig_block_width, line_h_page3, "( Surveyor and Loss Assessor )", 0, 1, 'C')

        # --- Add Photo Pages ---
        add_photo_section(pdf, "First inspection photo", get_survey_val('vehicle_regn_no'), photos_data.get('first_inspection', {}).get('images', []), photos_data.get('first_inspection', {}).get('per_page', 4))
        add_photo_section(pdf, "Dismantling/follow up photo", get_survey_val('vehicle_regn_no'), photos_data.get('dismantling', {}).get('images', []), photos_data.get('dismantling', {}).get('per_page', 4))
        add_photo_section(pdf, "Re-inspection photo", get_survey_val('vehicle_regn_no'), photos_data.get('reinspection', {}).get('images', []), photos_data.get('reinspection', {}).get('per_page', 4))

        pdf_bytes = pdf.output()
        request_id = str(uuid.uuid4())
        vehicle_no_raw = final_survey_data.get('vehicle_regn_no', '')
        
        # Store metadata and user_id for download authorization and filename generation
        generated_data_store[request_id] = { 
            "pdf_report": pdf_bytes, 
            "user_id": user_id,
            "report_no": final_survey_data.get('report_no', 'SurveyReport'),
            "vehicle_no": vehicle_no_raw
        }

        # Write to disk so all worker processes (e.g. multi-process Gunicorn) can serve the file
        try:
            project_root = os.path.dirname(os.path.abspath(__file__))
            temp_pdf_dir = os.path.join(project_root, 'uploads', 'temp_pdfs')
            os.makedirs(temp_pdf_dir, exist_ok=True)
            temp_pdf_path = os.path.join(temp_pdf_dir, f"{request_id}.pdf")
            with open(temp_pdf_path, 'wb') as f:
                f.write(pdf_bytes)
        except Exception as disk_err:
            print(f"Warning: Could not write temp PDF to disk: {disk_err}")

        # Auto-upload to Google Drive (if user connected their personal Drive via Settings)
        drive_link = None
        try:
            filename_base = "".join(c for c in vehicle_no_raw if c.isalnum() or c in ('_', '-')).rstrip() if vehicle_no_raw.strip() else 'SurveyReport'
            filename_pdf = f"{filename_base}.pdf"
            
            # access_token was passed from the handler (no session in background thread)
            if access_token:
                folder_name_to_use = "".join(c for c in vehicle_no_raw if c.isalnum() or c in ('_', '-', ' ')).strip() if vehicle_no_raw else 'Unknown_Vehicle'
                drive_link = upload_pdf_to_drive(access_token, pdf_bytes, filename_pdf, folder_name_to_use)
                if drive_link:
                    print(f"Report auto-uploaded to User's Personal Drive: {drive_link}")
                else:
                    print("Warning: Auto-upload to User's Personal Drive failed.")
                    
            if not drive_link:
                # Fallback to service account Drive (if available/working)
                drive_link = sheets_db.upload_report_pdf(pdf_bytes, filename_pdf, vehicle_no_raw)
                if drive_link:
                    print(f"Report auto-uploaded to Service Account Drive: {drive_link}")
                else:
                    print("Warning: Auto-upload to Service Account Drive failed (non-critical).")
        except Exception as drive_err:
            print(f"Warning: Drive auto-upload error (non-critical): {drive_err}")

        _complete_task(task_id, {"request_id": request_id, "drive_link": drive_link})
    except FPDFException as fpdf_err:
        print(f"FPDF Error generating files: {fpdf_err}")
        import traceback; traceback.print_exc()
        _fail_task(task_id, f"An error occurred during PDF generation: {fpdf_err}")
    except Exception as e:
        print(f"Error generating files: {e}")
        import traceback; traceback.print_exc()
        _fail_task(task_id, f"An unexpected error occurred during file generation: {e}")


# --- Download Route ---
@app.route('/download/<file_type>/<request_id>')
@login_required
def download_file(file_type, request_id):
    # Enforce job/asset owner verification on download routes safely
    job = sheets_db.get_job_by_request_id(request_id)
    if job and job.get('user_id') is not None:
        if str(job.get('user_id')) != str(current_user.id):
            abort(403, description="Access denied. You do not own this report.")
    
    pdf_bytes = None
    vehicle_no = ''
    report_no = 'SurveyReport'

    # Priority 1: Check in-memory store
    if request_id in generated_data_store:
        data = generated_data_store[request_id]
        if data.get('user_id') is not None and str(data.get('user_id')) != str(current_user.id):
            abort(403, description="Access denied. You do not own this report.")
        pdf_bytes = data['pdf_report']
        vehicle_no = data.get('vehicle_no', '').strip()
        report_no = data.get('report_no', 'SurveyReport')
    else:
        # Priority 2: Check local temporary files (production cross-process)
        project_root = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(project_root, 'uploads', 'temp_pdfs', f"{request_id}.pdf")
        
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                pdf_bytes = f.read()
            
            if job and job.get('result_json'):
                res_data = job['result_json']
                if isinstance(res_data, str):
                    try: res_data = json.loads(res_data)
                    except Exception: res_data = {}
                if isinstance(res_data, dict):
                    vehicle_no = res_data.get('vehicle_no', '').strip()
                    report_no = res_data.get('report_no', 'SurveyReport')
        else:
            # Priority 3: Check if request_id is a saved report_id in database
            workspace_admin_id = workspace_admin_id_for(current_user)
            if workspace_admin_id:
                saved_report = sheets_db.get_accessible_report_by_id(
                    request_id, workspace_admin_id, current_user.id)
            else:
                saved_report = sheets_db.get_report_by_id(request_id, current_user.id)

            if saved_report:
                report_data_json = saved_report.get('report_data_json', '{}')
                if isinstance(report_data_json, str):
                    try: report_data_dict = json.loads(report_data_json)
                    except Exception: report_data_dict = {}
                else:
                    report_data_dict = report_data_json or {}

                if not is_admin_user(current_user):
                    report_data_dict = redact_financial_report_data(report_data_dict)

                from modules.pdf import render_report
                user_snapshot = {
                    'full_name': current_user.full_name or 'Surveyor',
                    'qualifications': current_user.qualifications or '',
                    'designation': current_user.designation or '',
                    'license_no': current_user.license_no or '',
                    'expiry_date': current_user.expiry_date or '',
                    'membership_no': current_user.membership_no or '',
                    'address_line_1': current_user.address_line_1 or '',
                    'address_line_2': current_user.address_line_2 or '',
                    'address_line_3': current_user.address_line_3 or '',
                    'contact_no': current_user.contact_no or '',
                    'email': current_user.email or ''
                }
                pdf_res = render_report(report_data_dict, user_snapshot, current_user.id)
                pdf_bytes = pdf_res['pdf_bytes']
                survey_info = report_data_dict.get('survey_report', {})
                vehicle_no = survey_info.get('vehicle_regn_no', '').strip()
                report_no = survey_info.get('report_no', 'SurveyReport')
            else:
                abort(404, description="Report PDF file expired or not found.")

    if not pdf_bytes:
        abort(404, description="Report PDF file expired or not found.")

    if vehicle_no:
        filename_base = "".join(c for c in vehicle_no if c.isalnum() or c in ('_', '-')).rstrip()
    else:
        report_no_clean = report_no.replace(' ', '_').replace('/', '-')
        filename_base = "".join(c for c in report_no_clean if c.isalnum() or c in ('_', '-')).rstrip() or 'SurveyReport'

    if file_type == 'report_pdf':
        filename = f"{filename_base}.pdf"
        mimetype = 'application/pdf'
        file_content = io.BytesIO(pdf_bytes)
    else:
        abort(400, description="Invalid file type requested.")

    return send_file(
        file_content,
        mimetype=mimetype,
        as_attachment=False if request.args.get('preview') else True, # Allow inline for preview
        download_name=filename
    )

# --- Database Interaction Routes ---
@app.route('/save_report', methods=['POST'])
@login_required
def save_report():
    try:
        data = request.get_json()
        if not data or 'survey_report' not in data or 'assessment' not in data:
            return jsonify({"error": "Invalid data format received"}), 400

        survey_report = data.get('survey_report', {})
        report_no = survey_report.get('report_no', '').strip()
        
        if not report_no:
            return jsonify({"error": "Report Number cannot be empty"}), 400

        # Extract the UUID of the currently-loaded report (if editing an existing one).
        # This ensures we UPDATE the correct row instead of matching by report_no string.
        existing_report_id = data.get('_current_report_id') or None
        workspace_admin_id = workspace_admin_id_for(current_user)

        try:
            existing = None
            if existing_report_id:
                existing = (sheets_db.get_accessible_report_by_id(
                    existing_report_id, workspace_admin_id, current_user.id)
                    if workspace_admin_id else sheets_db.get_report_by_id(existing_report_id, current_user.id))
                if not existing:
                    return jsonify({'error': 'Report not found or access denied.'}), 404
                previous_data = existing.get('report_data_json', {})
                if isinstance(previous_data, str):
                    previous_data = json.loads(previous_data or '{}')
                if not is_admin_user(current_user):
                    data = preserve_financial_report_data(data, previous_data)

                # Historical rows stay private to their original owner even after workspace rollout.
                if existing.get('workspace_admin_id') is None:
                    saved_id = sheets_db.save_report(current_user.id, data, existing_report_id=existing_report_id)
                    if saved_id:
                        return jsonify({"success": True, "message": f'Report "{report_no}" saved successfully.', "report_id": saved_id})
                    return jsonify({"error": "Failed to save legacy report."}), 500

            if workspace_admin_id:
                claim_meta = data.get('claim_meta') or {}
                status = claim_meta.get('status') or data.get('status') or 'new_appointment'
                if status not in VALID_CLAIM_STATUSES:
                    return jsonify({'error': 'Invalid claim status.'}), 400
                requested_type = (claim_meta.get('survey_type') or data.get('survey_type') or '').lower()
                survey_type = 'spot' if requested_type == 'spot' or 'spot' in str(data.get('assessment', {}).get('report_type', '')).lower() else 'final'
                saved_id = sheets_db.save_workspace_report(
                    current_user.id, workspace_admin_id, data, existing_report_id=existing_report_id,
                    status=status, survey_type=survey_type)
            else:
                saved_id = sheets_db.save_report(current_user.id, data, existing_report_id=existing_report_id)
            if saved_id:
                return jsonify({"success": True, "message": f'Report "{report_no}" saved successfully.', "report_id": saved_id})
            else:
                return jsonify({"error": "Failed to save to Database (unknown error)."}), 500
        except Exception as sheet_error:
            print(f"Database Error: {sheet_error}")
            import traceback; traceback.print_exc()
            return jsonify({"error": f"Failed to save to Database: {str(sheet_error)}"}), 500

    except Exception as e:
        print(f"Error saving report: {e}")
        import traceback; traceback.print_exc()
        return jsonify({"error": f"An unexpected error occurred while saving: {e}"}), 500
    

@app.route('/get_saved_reports', methods=['GET'])
@login_required
def get_saved_reports():
    try:
        search_query = request.args.get('q', request.args.get('query', ''))
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 50, type=int)
        
        workspace_admin_id = workspace_admin_id_for(current_user)
        if workspace_admin_id:
            page_data = sheets_db.get_accessible_reports_page(
                workspace_admin_id, current_user.id, search_query, page, page_size)
        else:
            page_data = sheets_db.get_user_reports_page(current_user.id, search_query, page, page_size)
        reports = page_data.get('items', []) if isinstance(page_data, dict) else []
        
        reports_list = [
            {
                'id': r.get('id'),
                'report_no': r.get('report_no'),
                'insured_name': r.get('insured_name'),
                'vehicle_no': r.get('vehicle_no'),
                'claim_no': r.get('claim_no'),
                'status': r.get('status', 'new_appointment'),
                'survey_type': r.get('survey_type', 'final'),
                'saved_at': datetime.fromisoformat(r.get('saved_at')).strftime('%Y-%m-%d %H:%M:%S') if r.get('saved_at') else 'N/A'
            } for r in reports
        ]
        return jsonify({
            'items': reports_list,
            'page': page_data.get('page', page),
            'page_size': page_data.get('page_size', page_size),
            'total': page_data.get('total', 0)
        })
    except Exception as e:
        print(f"Error fetching saved reports: {e}")
        return jsonify({"error": f"Failed to fetch reports: {e}"}), 500
    

@app.route('/load_report/<report_id>', methods=['GET'])
@login_required
def load_report(report_id):
    try:
        # report_id is passed as int, but stored as whatever in sheets (int probably)
        # We need to find the specific report in the user's list or by ID
        # Since API doesn't have direct "get by id", we can reuse get_user_reports and filter 
        # OR implement get_report_by_id in sheets_db. 
        # For MVP, filtering user reports is safe enough for small data.
        
        workspace_admin_id = workspace_admin_id_for(current_user)
        if workspace_admin_id:
            target_report = sheets_db.get_accessible_report_by_id(
                report_id, workspace_admin_id, current_user.id)
        else:
            target_report = sheets_db.get_report_by_id(report_id, current_user.id)
        
        if target_report:
            try:
                report_data = json.loads(target_report.get('report_data_json'))
            except (json.JSONDecodeError, TypeError, ValueError):
                # Fallback if json string is malformed or empty
                report_data = {} 
            if workspace_admin_id:
                report_data['claim_meta'] = {
                    'status': target_report.get('status', 'new_appointment'),
                    'survey_type': target_report.get('survey_type', 'final'),
                    'email_received_date': target_report.get('email_received_date'),
                }
            if not is_admin_user(current_user):
                report_data = redact_financial_report_data(report_data)
            return jsonify(report_data)
        else:
            return jsonify({"error": "Report not found or access denied"}), 404
    except Exception as e:
        print(f"Error loading report {report_id}: {e}")
        return jsonify({"error": f"Failed to load report: {e}"}), 500


def _report_prefix_for_insurer(insurer_name):
    insurer = (insurer_name or '').upper()
    if 'NATIONAL INSURANCE' in insurer:
        return 'NIC'
    if 'NEW INDIA' in insurer:
        return 'NIA'
    if 'ORIENTAL' in insurer:
        return 'OIC'
    if 'UNITED INDIA' in insurer:
        return 'UIIC'
    return 'REP'


@app.route('/api/dashboard', methods=['GET'])
@login_required
def get_dashboard():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    date_range = request.args.get('range') or None
    dashboard = sheets_db.get_workspace_dashboard(workspace_admin_id, date_range=date_range)
    if not is_admin_user(current_user):
        for key in ('total_invoiced', 'amount_received', 'outstanding_fees', 'overdue_count'):
            dashboard.pop(key, None)
    return jsonify(dashboard)



@app.route('/api/claims', methods=['GET', 'POST'])
@login_required
def claims_register():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    if request.method == 'GET':
        status = request.args.get('status') or None
        month = request.args.get('month') or None
        insurer = request.args.get('insurer') or None
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 50, type=int)
        query = request.args.get('q', request.args.get('query', ''))
        return jsonify(sheets_db.get_workspace_reports_page(
            workspace_admin_id, query, page, page_size, status=status, month=month, insurer=insurer))

    data = request.get_json() or {}
    claim_no = str(data.get('claim_no', '')).strip()
    if not claim_no:
        return jsonify({'error': 'Claim number is required.'}), 400
    insurer = str(data.get('insurer', '')).strip()
    prefix = _report_prefix_for_insurer(insurer)
    sequence = sheets_db.reserve_report_number(workspace_admin_id, prefix, str(datetime.now().year))
    if sequence is None:
        return jsonify({'error': 'Failed to reserve a report number.'}), 500
    survey_type = 'spot' if str(data.get('survey_type', '')).lower() == 'spot' else 'final'
    report_type = 'Spot Report' if survey_type == 'spot' else 'Final Survey Report'
    report_data = {
        'survey_report': {
            'report_no': f'{prefix}/{datetime.now().year}/{sequence:02d}',
            'claim_no': claim_no,
            'vehicle_regn_no': str(data.get('vehicle_no', '')).strip(),
            'insured': str(data.get('insured_name', '')).strip(),
            'policy_no': str(data.get('policy_no', '')).strip(),
            'insurer': insurer,
            'date_of_loss': str(data.get('date_of_loss', '')).strip(),
        },
        'assessment': {'report_type': report_type},
        'photos': {},
        'claim_meta': {'status': data.get('status', 'new_appointment'), 'survey_type': survey_type},
    }
    status = report_data['claim_meta']['status']
    if status not in VALID_CLAIM_STATUSES:
        return jsonify({'error': 'Invalid claim status.'}), 400
    report_id = sheets_db.save_workspace_report(
        current_user.id, workspace_admin_id, report_data, status=status, survey_type=survey_type)
    if not report_id:
        return jsonify({'error': 'Failed to create claim workspace.'}), 500
    return jsonify({'success': True, 'report_id': str(report_id), 'report_no': report_data['survey_report']['report_no']}), 201


@app.route('/api/claims/<report_id>', methods=['PATCH'])
@login_required
def update_claim(report_id):
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    data = request.get_json() or {}
    status = data.get('status')
    if status not in VALID_CLAIM_STATUSES:
        return jsonify({'error': 'Invalid claim status.'}), 400
    if not sheets_db.update_workspace_report_status(report_id, workspace_admin_id, current_user.id, status):
        return jsonify({'error': 'Claim not found or access denied.'}), 404
    return jsonify({'success': True, 'status': status})


@app.route('/api/reports/monthly', methods=['GET'])
@login_required
def monthly_reports():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    month = request.args.get('month') or datetime.now().strftime('%Y-%m')
    insurer = request.args.get('insurer') or None
    claims = sheets_db.get_workspace_reports_page(workspace_admin_id, '', 1, 100, month=month, insurer=insurer)
    response = {'month': month, 'claims': claims}
    if is_admin_user(current_user):
        response['fees'] = sheets_db.get_workspace_fee_bills(workspace_admin_id, month=month, insurer=insurer)
    return jsonify(response)

@app.route('/api/generate_report_no', methods=['POST'])
@login_required
def generate_report_no():
    try:
        data = request.get_json()
        insurer_name = data.get('insurer', '').upper()
        
        # Map known insurers to prefixes
        prefix = "REP"
        if "NATIONAL INSURANCE" in insurer_name:
            prefix = "NIC"
        elif "NEW INDIA" in insurer_name:
            prefix = "NIA"
        elif "ORIENTAL" in insurer_name:
            prefix = "OIC"
        elif "UNITED INDIA" in insurer_name:
            prefix = "UIIC"
        
        current_year = str(datetime.now().year)
        
        workspace_admin_id = workspace_admin_id_for(current_user)
        if workspace_admin_id:
            sequence = sheets_db.reserve_report_number(workspace_admin_id, prefix, current_year)
            if sequence is None:
                return jsonify({'error': 'Failed to reserve report number'}), 500
            return jsonify({"report_no": f"{prefix}/{current_year}/{sequence:02d}"}), 200

        # Legacy user-scoped numbering remains available for preserved records.
        reports_metadata = sheets_db.get_user_reports_metadata_only(current_user.id)
        
        max_seq = 0
        search_pattern = f"{prefix}/{current_year}/"
        
        for report in reports_metadata:
            report_num = str(report.get('report_no', ''))
            if report_num.startswith(search_pattern):
                try:
                    # Extract the sequence part (e.g., from "NIC/2026/01", extract "01")
                    seq_str = report_num.split('/')[-1]
                    seq_num = int(seq_str)
                    if seq_num > max_seq:
                        max_seq = seq_num
                except ValueError:
                    continue
        
        new_seq = max_seq + 1
        new_report_no = f"{prefix}/{current_year}/{new_seq:02d}"
        
        return jsonify({"report_no": new_report_no}), 200
        
    except Exception as e:
        print(f"Error generating report number: {e}")
        return jsonify({"error": "Failed to generate report number"}), 500

@app.route('/delete_report/<report_id>', methods=['DELETE'])
@login_required
def delete_report(report_id):
    try:
        data = request.get_json()
        password = data.get('password')
        
        if not password:
             return jsonify({"error": "Password is required to delete a report."}), 400
             
        # Check password against current user (stored in session/sheets)
        # current_user.password_hash came from sheets_db during login
        if not bcrypt.check_password_hash(current_user.password_hash, password):
             return jsonify({"error": "Incorrect password."}), 403

        # Deletion in Sheets is risky (shifting rows). 
        # For MVP, we will NOT delete to prevent data corruption.
        # Alternatively, we could clear the row content or mark as "deleted" column.
        workspace_admin_id = workspace_admin_id_for(current_user)
        success = (sheets_db.delete_accessible_report(report_id, workspace_admin_id, current_user.id)
                   if workspace_admin_id else sheets_db.delete_report(report_id, current_user.id))
        if success:
             return jsonify({"message": "Report deleted successfully"}), 200
        else:
             return jsonify({"error": "Failed to delete report or not found"}), 404
        
    except Exception as e:
        print(f"Error deleting report {report_id}: {e}")
        return jsonify({"error": f"Failed to delete report: {e}"}), 500

@app.route('/download_consolidated_csv', methods=['GET'])
@app.route('/download_gst_excel', methods=['GET'])
@login_required
@admin_required
def download_consolidated_csv():
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')

        if not from_date_str or not to_date_str:
            return jsonify({"error": "Both from_date and to_date are required."}), 400

        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
            to_date_end_of_day = datetime.combine(to_date_dt.date(), datetime.max.time())
            from_date_start_of_day = datetime.combine(from_date, datetime.min.time())
        except ValueError:
            return jsonify({"error": "Invalid date format. Please use YYYY-MM-DD."}), 400

        workspace_admin_id = workspace_admin_id_for(current_user)
        if workspace_admin_id:
            all_reports = sheets_db.get_workspace_reports_page(workspace_admin_id, '', 1, 100).get('items', [])
            # Export needs full report JSON, so resolve each workspace row before processing it.
            all_reports = [sheets_db.get_workspace_report_by_id(row.get('id'), workspace_admin_id) for row in all_reports]
            all_reports = [row for row in all_reports if row]
            legacy_reports = sheets_db.get_user_reports(current_user.id)
            report_ids = {str(row.get('id')) for row in all_reports}
            all_reports.extend(row for row in legacy_reports if str(row.get('id')) not in report_ids)
            all_fee_bills = sheets_db.get_workspace_fee_bills(workspace_admin_id)
            legacy_fees = sheets_db.get_user_fee_bills(current_user.id)
            fee_ids = {str(row.get('id')) for row in all_fee_bills}
            all_fee_bills.extend(row for row in legacy_fees if str(row.get('id')) not in fee_ids)
        else:
            all_reports = sheets_db.get_user_reports(current_user.id)
            all_fee_bills = sheets_db.get_user_fee_bills(current_user.id)

        export_rows = []

        for r in all_reports:
            saved_at_str = r.get('saved_at')
            if not saved_at_str: continue
            try:
                saved_at = datetime.fromisoformat(saved_at_str)
                if not (from_date_start_of_day <= saved_at <= to_date_end_of_day):
                    continue

                report_data = json.loads(r.get('report_data_json', '{}')) if isinstance(r.get('report_data_json'), str) else r.get('report_data_json', {})
                survey_data = report_data.get('survey_report', {})
                assessment_data = report_data.get('assessment', {})
                page3_details = assessment_data.get('page3_details', {})
                fee_items = page3_details.get('fee_items', [])

                taxable_sum = 0.0
                for item in fee_items:
                    try:
                        taxable_sum += float(item.get('amount', 0.0))
                    except (ValueError, TypeError):
                        pass

                if taxable_sum <= 0:
                    calc = _calculate_report_assessment_summary(assessment_data, survey_data)
                    taxable_sum = calc.get('page3_gross_total', 0.0)

                gst_pc = 18.0
                gst_amount = taxable_sum * (gst_pc / 100.0)
                total_amount = taxable_sum + gst_amount

                export_rows.append({
                    "date_obj": saved_at,
                    "insured_name": survey_data.get('insured', 'N/A'),
                    "insurer_name": survey_data.get('insurer', 'N/A'),
                    "policy_no": survey_data.get('policy_no', 'N/A'),
                    "claim_no": survey_data.get('claim_no', 'N/A'),
                    "vehicle_no": survey_data.get('vehicle_regn_no', 'N/A'),
                    "invoice_no": survey_data.get('report_no', 'N/A'),
                    "invoice_date": survey_data.get('report_date', saved_at.strftime('%Y-%m-%d')),
                    "assigned_date": str(r.get('created_at') or saved_at.strftime('%Y-%m-%d'))[:10],
                    "gst_pc": f"{gst_pc:g}%",
                    "gst_amount": f"{gst_amount:.2f}",
                    "taxable_amount": f"{taxable_sum:.2f}",
                    "total_amount": f"{total_amount:.2f}"
                })
            except Exception as e_rep:
                print(f"Error parsing survey report for CSV export: {e_rep}")

        for fb in all_fee_bills:
            created_at_str = fb.get('created_at') or fb.get('invoice_date')
            try:
                if 'T' in str(created_at_str):
                    dt_val = datetime.fromisoformat(str(created_at_str))
                else:
                    dt_val = datetime.strptime(str(created_at_str), '%Y-%m-%d')

                if not (from_date_start_of_day <= dt_val <= to_date_end_of_day):
                    continue

                taxable = float(fb.get('taxable_amount', 0.0))
                gst_pc = float(fb.get('gst_pc', 18.0))
                gst_amt = float(fb.get('gst_amount', taxable * (gst_pc / 100.0)))
                total_amt = float(fb.get('total_amount', taxable + gst_amt))

                export_rows.append({
                    "date_obj": dt_val,
                    "insured_name": fb.get('insured_name', 'N/A'),
                    "insurer_name": fb.get('insurer_name', fb.get('insurer', 'N/A')),
                    "policy_no": fb.get('policy_no', 'N/A'),
                    "claim_no": fb.get('claim_no', 'N/A'),
                    "vehicle_no": fb.get('vehicle_no', 'N/A'),
                    "invoice_no": fb.get('invoice_no', 'N/A'),
                    "invoice_date": fb.get('invoice_date', dt_val.strftime('%Y-%m-%d')),
                    "assigned_date": str(fb.get('created_at') or dt_val.strftime('%Y-%m-%d'))[:10],
                    "gst_pc": f"{gst_pc:g}%",
                    "gst_amount": f"{gst_amt:.2f}",
                    "taxable_amount": f"{taxable:.2f}",
                    "total_amount": f"{total_amt:.2f}"
                })
            except Exception as e_fb:
                print(f"Error parsing fee bill for CSV export: {e_fb}")

        export_rows.sort(key=lambda x: x['date_obj'])

        output = io.StringIO()
        csv_writer = csv.writer(output)

        headers = [
            "Insured Name",
            "Insurer Company Name",
            "Policy number",
            "Claim number",
            "Vehicle number",
            "Invoice no",
            "Invoice date",
            "Assigned Date",
            "Gst %",
            "Gst amount",
            "Taxable amount",
            "Total amount (including GST)"
        ]
        csv_writer.writerow(headers)

        for row in export_rows:
            csv_writer.writerow([
                row["insured_name"],
                row["insurer_name"],
                row["policy_no"],
                row["claim_no"],
                row["vehicle_no"],
                row["invoice_no"],
                row["invoice_date"],
                row["assigned_date"],
                row["gst_pc"],
                row["gst_amount"],
                row["taxable_amount"],
                row["total_amount"]
            ])

        output.seek(0)
        filename = f"Consolidated_GST_Report_{from_date_str}_to_{to_date_str}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error generating consolidated GST CSV: {e}")
        return jsonify({"error": f"Failed to generate consolidated report: {e}"}), 500


@app.route('/download_fees_excel', methods=['GET'])
@login_required
@admin_required
def download_fees_excel():
    """Monthly CA-friendly Survey Fee Register as a real XLSX workbook."""
    month = request.args.get('month') or datetime.now().strftime('%Y-%m')
    insurer = request.args.get('insurer') or None
    if not re.fullmatch(r'\d{4}-\d{2}', month):
        return jsonify({'error': 'month must use YYYY-MM format.'}), 400
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        bills = sheets_db.get_workspace_fee_bills(workspace_admin_id, month=month, insurer=insurer)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Survey Fee Register'
        headers = [
            'Invoice Date', 'Invoice No', 'Insurer', 'Insured Name', 'Claim No', 'Policy No',
            'Vehicle No', 'Professional Fee', 'GST %', 'GST Amount', 'Gross Invoice Value',
            'TDS Amount', 'Cash Received', 'Outstanding Amount', 'Due Date', 'Payment Status', 'Invoice Status'
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
        for bill in bills:
            sheet.append([
                bill.get('invoice_date', ''), bill.get('invoice_no', ''), bill.get('insurer_name', ''),
                bill.get('insured_name', ''), bill.get('claim_no', ''), bill.get('policy_no', ''),
                bill.get('vehicle_no', ''), float(bill.get('professional_fee', bill.get('taxable_amount', 0)) or 0),
                float(bill.get('gst_pc', 0) or 0), float(bill.get('gst_amount', 0) or 0),
                float(bill.get('gross_invoice_value', bill.get('total_amount', 0)) or 0),
                float(bill.get('tds_amount', 0) or 0), float(bill.get('amount_received', 0) or 0),
                float(bill.get('outstanding_amount', 0) or 0), bill.get('due_date', ''),
                bill.get('payment_status', ''), bill.get('invoice_status', ''),
            ])
        for column in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = min(28, max(13, len(headers[column - 1]) + 2))
        for row in sheet.iter_rows(min_row=2, min_col=8, max_col=14):
            for cell in row:
                cell.number_format = '#,##0.00'
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        suffix = f'_{insurer}' if insurer else ''
        safe_suffix = re.sub(r'[^A-Za-z0-9_-]+', '_', suffix)
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'Survey_Fee_Register_{month}{safe_suffix}.xlsx')
    except Exception as exc:
        print(f'Error generating fee XLSX: {exc}')
        return jsonify({'error': 'Failed to generate Survey Fee Register.'}), 500


@app.route('/api/next_invoice_no', methods=['GET'])
@login_required
@admin_required
def get_next_invoice_no():
    insurer = request.args.get('insurer', 'Company')
    date_val = request.args.get('date')
    inv_no = sheets_db.get_next_invoice_number(
        current_user.id, insurer, date_val, workspace_admin_id=workspace_admin_id_for(current_user))
    return jsonify({"invoice_no": inv_no})


@app.route('/api/fees_summary', methods=['GET'])
@login_required
@admin_required
def fees_summary():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    dashboard = sheets_db.get_workspace_dashboard(workspace_admin_id)
    return jsonify({key: dashboard[key] for key in ('total_invoiced', 'amount_received', 'outstanding_fees', 'overdue_count')})


@app.route('/api/fee_bills', methods=['GET', 'POST'])
@login_required
@admin_required
def handle_fee_bills():
    workspace_admin_id = workspace_admin_id_for(current_user)
    if not workspace_admin_id:
        return jsonify({'error': 'Your account is not assigned to an admin workspace.'}), 403
    if request.method == 'POST':
        data = request.get_json() or {}
        report_id = data.get('report_id')
        if report_id and not sheets_db.get_workspace_report_by_id(report_id, workspace_admin_id):
            return jsonify({'error': 'The selected report does not belong to this workspace.'}), 404
        bill_id = sheets_db.save_fee_bill(current_user.id, data, workspace_admin_id=workspace_admin_id)
        if not bill_id:
            return jsonify({'error': 'Could not save the fee bill.'}), 500
        return jsonify({"success": True, "id": bill_id}), 201

    bills = sheets_db.get_workspace_fee_bills(
        workspace_admin_id, month=request.args.get('month') or None,
        insurer=request.args.get('insurer') or None, report_id=request.args.get('report_id') or None)
    return jsonify(bills), 200


@app.route('/api/fee_bills/<bill_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_fee_bill_route(bill_id):
    success = sheets_db.delete_fee_bill(bill_id, current_user.id, workspace_admin_id=workspace_admin_id_for(current_user))
    if success:
        return jsonify({"success": True}), 200
    return jsonify({"error": "Failed to delete fee bill"}), 400


@app.route('/generate_fee_pdf', methods=['POST'])
@login_required
@admin_required
def generate_fee_pdf_route():
    try:
        data = request.get_json() or {}
        include_sig = data.get('include_signature', True)
        from modules.pdf import render_fee_report
        user_snapshot = {
            'full_name': current_user.full_name or 'Surveyor',
            'qualifications': current_user.qualifications or '',
            'designation': current_user.designation or '',
            'license_no': current_user.license_no or '',
            'expiry_date': current_user.expiry_date or '',
            'membership_no': current_user.membership_no or '',
            'address_line_1': current_user.address_line_1 or '',
            'address_line_2': current_user.address_line_2 or '',
            'address_line_3': current_user.address_line_3 or '',
            'contact_no': current_user.contact_no or '',
            'email': current_user.email or ''
        }
        res = render_fee_report(data, user_snapshot, current_user.id, include_signature=include_sig)
        pdf_bytes = res['pdf_bytes']
        inv_no = res['invoice_no']
        safe_name = "".join(c for c in inv_no if c.isalnum() or c in ('_', '-')) or 'FeeBill'

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{safe_name}.pdf"
        )
    except Exception as e:
        print(f"Error generating fee PDF: {e}")
        return jsonify({"error": str(e)}), 500




# --- Initial User Creation Helper (Manual Trigger via API or Script recommended for Sheets) ---
# Since we removed DB init, users must be added to the Sheet manually or via a new CLI command.
@app.cli.command('create-default-user')
def create_default_user():
    """Create the default user in Google Sheets if not exists."""
    # SECURITY: Read credentials from environment variables (VULN-01)
    username = os.getenv('ADMIN_USERNAME')
    password = os.getenv('ADMIN_PASSWORD')
    
    if not username or not password:
        print("ERROR: ADMIN_USERNAME and ADMIN_PASSWORD must be set as environment variables.")
        return
    
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    
    existing = sheets_db.get_user_by_username(username)
    if not existing:
        user_data = {
            'username': username,
            'password_hash': hashed_password,
            'full_name': os.getenv('USER_FULL_NAME', 'Default User'),
            'qualifications': os.getenv('USER_QUALIFICATIONS', ''),
            'designation': os.getenv('USER_DESIGNATION', 'Surveyor & Loss Assessor'),
            'license_no': os.getenv('USER_LICENSE_NO', ''),
            'expiry_date': os.getenv('USER_EXPIRY_DATE', ''),
            'membership_no': os.getenv('USER_MEMBERSHIP_NO', ''),
            'address_line_1': os.getenv('USER_ADDRESS_1', ''),
            'address_line_2': os.getenv('USER_ADDRESS_2', ''),
            'address_line_3': os.getenv('USER_ADDRESS_3', ''),
            'contact_no': os.getenv('USER_CONTACT_NO', ''),
            'email': os.getenv('USER_EMAIL', ''),
            'role': 'admin',
            'admin_id': None,
            'permissions': {},
        }
        sheets_db.create_user(user_data)
        print(f"Default user '{username}' created in Sheets.")
    else:
        print(f"User '{username}' already exists in Sheets.")

@app.cli.command('create-user')
@click.argument('username')
@click.argument('password')
@click.option('--name', default='Employee')
def create_user_cli(username, password, name):
    """Create a new user account in PostgreSQL."""
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    existing = sheets_db.get_user_by_username(username)
    if not existing:
        user_data = {
            'username': username,
            'password_hash': hashed_password,
            'full_name': name,
            'qualifications': '',
            'designation': 'Surveyor & Loss Assessor',
            'license_no': '',
            'expiry_date': '',
            'membership_no': '',
            'address_line_1': '',
            'address_line_2': '',
            'address_line_3': '',
            'contact_no': '',
            'email': ''
        }
        sheets_db.create_user(user_data)
        print(f"User '{username}' created successfully.")
    else:
        print(f"Error: User '{username}' already exists.")


@app.cli.command('promote-admin')
@click.argument('username')
def promote_admin_cli(username):
    """Promote an existing account to an admin workspace owner."""
    if sheets_db.promote_user_to_admin(username):
        print(f"User '{username}' is now an admin workspace owner.")
    else:
        print(f"Error: Could not promote '{username}'. Confirm that the user exists.")


@app.cli.command('create-employee')
@click.argument('username')
@click.argument('temporary_password')
@click.option('--admin', 'admin_username', required=True, help='Username of the owning admin workspace.')
@click.option('--name', default='Employee')
@click.option('--email', default='')
@click.option('--gmail-sync', is_flag=True, default=False, help='Allow this employee to sync the shared Gmail mailbox.')
def create_employee_cli(username, temporary_password, admin_username, name, email, gmail_sync):
    """Create an employee in an existing admin-owned workspace."""
    if len(temporary_password) < 8:
        print('Error: temporary_password must contain at least 8 characters.')
        return
    if sheets_db.get_user_by_username(username):
        print(f"Error: User '{username}' already exists.")
        return
    admin = sheets_db.get_user_by_username(admin_username)
    if not admin or admin.get('role') != 'admin':
        print(f"Error: '{admin_username}' is not an admin workspace owner.")
        return
    user_id = sheets_db.create_user({
        'username': username,
        'password_hash': bcrypt.generate_password_hash(temporary_password).decode('utf-8'),
        'full_name': name,
        'qualifications': '', 'designation': 'Surveyor & Loss Assessor', 'license_no': '',
        'expiry_date': '', 'membership_no': '', 'address_line_1': '', 'address_line_2': '',
        'address_line_3': '', 'contact_no': '', 'email': email,
        'role': 'employee', 'admin_id': admin.get('id'),
        'permissions': {'gmail_sync': bool(gmail_sync)}, 'must_change_password': True,
    })
    if user_id:
        print(f"Employee '{username}' created in {admin_username}'s workspace.")
    else:
        print(f"Error: Could not create employee '{username}'.")

# --- App Factory Pattern ---
def create_app(db_adapter=None, task_executor=None):
    new_app = Flask(__name__)
    new_app.config.update(app.config)
    
    if db_adapter:
        new_app.config['DB_ADAPTER'] = db_adapter
    if task_executor:
        new_app.config['TASK_EXECUTOR'] = task_executor
        
    # Copy route mappings and middleware hooks from the global app instance
    new_app.url_map = app.url_map
    new_app.view_functions = app.view_functions.copy()
    new_app.before_request_funcs = app.before_request_funcs.copy()
    new_app.after_request_funcs = app.after_request_funcs.copy()
    new_app.teardown_request_funcs = app.teardown_request_funcs.copy()
    new_app.teardown_appcontext_funcs = app.teardown_appcontext_funcs.copy()
    new_app.error_handler_spec = app.error_handler_spec.copy()
    new_app.cli = app.cli
    
    bcrypt.init_app(new_app)
    login_manager.init_app(new_app)
    if hasattr(limiter, 'init_app'):
        limiter.init_app(new_app)
        
    return new_app

# Instantiate the default application instance for WSGI/Gunicorn servers
app = create_app()

# --- Run Application ---
if __name__ == '__main__':
    is_dev = os.getenv('FLASK_ENV') == 'development'
    app.run(debug=is_dev, host='0.0.0.0', port=5000)
